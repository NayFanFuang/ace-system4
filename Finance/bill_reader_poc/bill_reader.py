"""
ACE Bill Reader — Proof of Concept
==================================

อ่านบิลค่าโทรศัพท์ (AIS / True / DTAC) ที่เป็น PDF สแกน แล้วแปลงเป็น
Payment Voucher (ฟอร์ม PV.03) อัตโนมัติ

Pipeline:
    PDF (สแกน) --render--> รูปภาพต่อหน้า --AI Vision--> JSON ต่อรายการเบอร์
              --compute--> เติม WHT 3% + Net --fill--> ไฟล์ Excel PV.03

ค่าที่ "อ่านได้จากบิลจริง":  เบอร์, รอบบิล, ยอดก่อน VAT, VAT 7%
ค่าที่ "ระบบคำนวณเพิ่ม":     WHT 3% = ยอดก่อน VAT x 3%   (บิลไม่ได้พิมพ์มา)
                            Net    = ยอด + VAT - WHT

หมายเหตุ: เป็น PoC — ฟังก์ชัน AI Vision ต่อ Anthropic API ได้จริง
(ต้องตั้ง env ANTHROPIC_API_KEY) ถ้าไม่มี key จะ fallback ใช้ผลที่ cache ไว้
เพื่อสาธิต pipeline ส่วนที่เหลือ (render -> compute -> Excel) ให้ครบเส้น
"""

from __future__ import annotations

import base64
import datetime
import json
import os
import re
import sys
from dataclasses import dataclass, field, asdict
from pathlib import Path

import fitz  # pymupdf
import openpyxl
from copy import copy

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent.parent  # ACE_System4/
TEMPLATE = ROOT / "Finance" / "9.Ver05 Template payment voucher  Examples.xlsx"
TEMPLATE_SHEET = "Template PV.03 (Other Expense)"

WHT_RATE = 0.03  # หัก ณ ที่จ่าย 3% (ค่าบริการ/ค่าโทรศัพท์ นิติบุคคล)

# ค่าตั้งต้นสำหรับฟิลด์ที่ "อ่านจากบิลไม่ได้" -> ระบบเติมเอง (แก้ได้ที่นี่)
DEFAULTS = {
    "project": "NBTC2501 NBTC NSA/SA Benchmarking Project",
    "name":    "Ms.Anong Jantaraeng",   # ผู้ขอเบิก
    "issued":  "Ms.Anong Jantaraeng",   # ผู้จัดทำ (Account payable)
}

# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class BillLine:
    """หนึ่งบรรทัด = หนึ่งเบอร์/หนึ่งใบแจ้งหนี้"""
    phone: str            # เบอร์ (Reference No.2)  เช่น 0632297691
    period: str           # รอบบิล เช่น 01/03/2026-31/03/2026
    amount: float         # ยอดก่อน VAT (ค่าบริการรวมปรับปรุง)
    vat: float            # VAT 7%
    vendor: str = ""      # ผู้ให้บริการ
    source_page: int = 0  # หน้าใน PDF ที่อ่านมา

    # คำนวณเพิ่ม
    wht: float = field(default=0.0)
    net: float = field(default=0.0)

    def compute(self) -> "BillLine":
        self.wht = round(self.amount * WHT_RATE, 2)
        self.net = round(self.amount + self.vat - self.wht, 2)
        return self

    @property
    def phone_fmt(self) -> str:
        d = re.sub(r"\D", "", self.phone)
        if len(d) == 10:
            return f"{d[0:2]}-{d[2:6]}-{d[6:]}"
        return self.phone

    def desc(self) -> str:
        # จับชื่อแบรนด์สั้น ๆ จากชื่อผู้ให้บริการ
        tag = "Bill"
        for brand in ("AIS", "True", "DTAC", "NT", "TOT"):
            if brand.lower() in self.vendor.lower():
                tag = brand
                break
        return f"{tag} no.{self.phone_fmt} of period {self.period}"


# ---------------------------------------------------------------------------
# Step 1 — render PDF pages to images
# ---------------------------------------------------------------------------

def render_pdf(pdf_path: Path, out_dir: Path, dpi: int = 200) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    doc = fitz.open(pdf_path)
    paths = []
    for i in range(doc.page_count):
        pix = doc[i].get_pixmap(dpi=dpi)
        p = out_dir / f"page_{i + 1:02d}.png"
        pix.save(p)
        paths.append(p)
    doc.close()
    return paths


# ---------------------------------------------------------------------------
# Step 2 — AI Vision extraction (Anthropic Claude)
# ---------------------------------------------------------------------------

VISION_PROMPT = """คุณเป็นผู้ช่วยบัญชี อ่าน "ใบแจ้งค่าใช้บริการ" (บิลค่าโทรศัพท์) ที่แนบมา
ดึงข้อมูลต่อไปนี้ออกมาเป็น JSON เท่านั้น ห้ามมีข้อความอื่น:

{
  "vendor": "ชื่อผู้ให้บริการ เช่น AIS / True / DTAC",
  "phone": "เบอร์โทร (Reference No.2 / เลขหมาย) ตัวเลขล้วน",
  "period": "รอบบิล (Bill Cycle) รูปแบบ DD/MM/YYYY-DD/MM/YYYY แปลง พ.ศ.เป็น ค.ศ.",
  "amount": ยอดค่าบริการก่อน VAT (ค่าบริการรวมปรับปรุง) เป็นตัวเลข,
  "vat": ภาษีมูลค่าเพิ่ม 7% เป็นตัวเลข
}

กฎ: ถ้าปีเป็น พ.ศ. (เช่น 2569) ให้ลบ 543 เป็น ค.ศ. (2026)
ห้ามคำนวณ WHT ระบบจะคำนวณเอง ตอบเฉพาะตัวเลขที่พิมพ์อยู่บนบิลจริง"""


def extract_with_vision(image_path: Path, model: str = "claude-opus-4-8") -> dict:
    """ส่งรูปบิล 1 หน้าให้ Claude vision -> dict {vendor, phone, period, amount, vat}"""
    import anthropic  # import here so the rest of the PoC runs without the SDK

    client = anthropic.Anthropic()  # ใช้ env ANTHROPIC_API_KEY
    b64 = base64.standard_b64encode(image_path.read_bytes()).decode()
    msg = client.messages.create(
        model=model,
        max_tokens=512,
        messages=[{
            "role": "user",
            "content": [
                {"type": "image",
                 "source": {"type": "base64", "media_type": "image/png", "data": b64}},
                {"type": "text", "text": VISION_PROMPT},
            ],
        }],
    )
    text = "".join(b.text for b in msg.content if b.type == "text")
    text = re.sub(r"^```(json)?|```$", "", text.strip(), flags=re.MULTILINE).strip()
    return json.loads(text)


def read_bill(pdf_path: Path, work_dir: Path, cache: Path | None = None) -> list[BillLine]:
    """อ่านบิลทั้งไฟล์ -> list[BillLine] (พร้อม fallback cache เมื่อไม่มี API key)"""
    images = render_pdf(pdf_path, work_dir / "pages")

    use_cache = cache and cache.exists() and not os.environ.get("ANTHROPIC_API_KEY")
    if use_cache:
        print(f"[i] ไม่มี ANTHROPIC_API_KEY -> ใช้ผล vision ที่ cache: {cache.name}")
        raw = json.loads(cache.read_text(encoding="utf-8"))
        records = raw["lines"]
    else:
        print(f"[i] อ่านด้วย AI Vision ({len(images)} หน้า)...")
        records = []
        for i, img in enumerate(images, 1):
            d = extract_with_vision(img)
            d["source_page"] = i
            records.append(d)

    lines = [BillLine(
        phone=str(r["phone"]), period=r["period"], amount=float(r["amount"]),
        vat=float(r["vat"]), vendor=r.get("vendor", ""),
        source_page=r.get("source_page", 0),
    ).compute() for r in records]
    return lines


# ---------------------------------------------------------------------------
# Step 3 — fill the PV.03 Excel template
# ---------------------------------------------------------------------------

FIRST_ROW = 12   # แถวแรกของตารางรายการ
LAST_ROW = 23    # แถวสุดท้าย (รวม 12 บรรทัด) ก่อน Total ที่แถว 24


def fill_pv(lines: list[BillLine], header: dict, out_path: Path) -> Path:
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb[TEMPLATE_SHEET]

    # หัวกระดาษ
    if header.get("pv_no"):   ws["M5"] = header["pv_no"]
    if header.get("item"):    ws["M6"] = header["item"]
    if header.get("project"): ws["C7"] = header["project"]
    if header.get("name"):    ws["C8"] = header["name"]
    if header.get("issued"):  ws["C9"] = header["issued"]
    if header.get("date"):    ws["M7"] = header["date"]

    # บรรทัดหัวเรื่องผู้ให้บริการ (เหมือนตัวอย่าง True: ชื่อบริษัทเป็นบรรทัดแรกในตาราง)
    r = FIRST_ROW
    vendor = lines[0].vendor if lines else ""
    if vendor:
        ws.cell(row=r, column=2, value=vendor)  # B
        r += 1

    for idx, ln in enumerate(lines, 1):
        if r > LAST_ROW:
            print(f"[!] เกิน {LAST_ROW - FIRST_ROW + 1} บรรทัด — ตัดที่ {idx-1} รายการ")
            break
        ws.cell(row=r, column=1, value=idx)            # A: Item.
        ws.cell(row=r, column=2, value=" - " + ln.desc())  # B: Desc
        ws.cell(row=r, column=10, value=ln.amount)     # J: Amount baht
        ws.cell(row=r, column=11, value=ln.vat)        # K: Vat 7%
        ws.cell(row=r, column=12, value=-ln.wht)       # L: WHT 3% (ติดลบ)
        ws.cell(row=r, column=13, value=ln.net)        # M: Net Total
        r += 1

    # J24..M24 เป็นสูตร SUM อยู่แล้ว ไม่ต้องแตะ
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# Auto-fill header — เติมฟิลด์ที่อ่านจากบิลไม่ได้ให้เองทั้งหมด
# ---------------------------------------------------------------------------

THAI_MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def build_header(pdf_path: Path, today: datetime.date | None = None) -> dict:
    """เติม PV No / Item / Date / Project / Name / Issued ให้ครบ
    - PV No, Item : แกะจากชื่อไฟล์ เช่น 'TF-04-2026 Item.43'
    - Date        : วันที่ปัจจุบัน (รูปแบบ DD-Mon-YYYY)
    - ที่เหลือ     : ค่าตั้งต้นใน DEFAULTS
    ฟิลด์ที่เดาไม่ได้จะ fallback อัตโนมัติ ไม่มี placeholder ค้าง
    """
    today = today or datetime.date.today()
    stem = pdf_path.stem

    m = re.search(r"TF-?(\d{1,2})-?(\d{4})", stem, re.I)
    pv_no = f"TF-{int(m.group(1)):02d}/{m.group(2)}" if m else f"TF-XX/{today.year}"

    mi = re.search(r"Item\.?\s*(\d+)", stem, re.I)
    item = mi.group(1) if mi else ""

    date_str = f"{today.day:02d}-{THAI_MONTHS[today.month - 1]}-{today.year}"

    return {
        "pv_no": pv_no,
        "item": item,
        "date": date_str,
        "project": DEFAULTS["project"],
        "name": DEFAULTS["name"],
        "issued": DEFAULTS["issued"],
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def print_table(lines: list[BillLine]) -> None:
    print(f"\n{'#':>2}  {'Desc':45} {'Amount':>10} {'VAT':>8} {'WHT':>8} {'Net':>10}")
    print("-" * 88)
    tA = tV = tW = tN = 0.0
    for i, ln in enumerate(lines, 1):
        print(f"{i:>2}  {ln.desc():45} {ln.amount:>10,.2f} {ln.vat:>8,.2f} "
              f"{-ln.wht:>8,.2f} {ln.net:>10,.2f}")
        tA += ln.amount; tV += ln.vat; tW += ln.wht; tN += ln.net
    print("-" * 88)
    print(f"{'':>2}  {'TOTAL':45} {tA:>10,.2f} {tV:>8,.2f} {-tW:>8,.2f} {tN:>10,.2f}")


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print("usage: python bill_reader.py <bill.pdf> [--cache cache.json] "
              "[--out PV.xlsx]")
        return 1
    pdf = Path(argv[1])
    cache = None
    out = HERE / "output" / (pdf.stem + "_PV.xlsx")
    if "--cache" in argv:
        cache = Path(argv[argv.index("--cache") + 1])
    if "--out" in argv:
        out = Path(argv[argv.index("--out") + 1])

    work = HERE / "work" / pdf.stem
    lines = read_bill(pdf, work, cache=cache)
    print_table(lines)

    header = build_header(pdf)
    print(f"[i] หัวกระดาษ (เติมเอง): PV {header['pv_no']} · Item {header['item']} · "
          f"{header['date']} · {header['project']}")
    saved = fill_pv(lines, header, out)
    print(f"\n[OK] สร้าง Payment Voucher: {saved}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
