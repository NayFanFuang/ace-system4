import json
import re
import secrets
import string
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import base64
import uuid

from fastapi import APIRouter, Depends, HTTPException, Query, Request, UploadFile, File, Form
from pydantic import BaseModel, EmailStr
from sqlalchemy import func, or_, select, text, update
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.deps import ROLE_LABELS, ROLE_SCOPES, get_current_user, require_hr_read_user, require_hr_user, require_po_finance_action_user, require_po_tracking_user, require_project_user, require_self_or_admin
from app.models.auth_user import AuthUser
from app.models.billing_plan import BillingPlan
from app.models.clock import ClockSite, ClockSession
from app.models.employee import Employee, EmployeeRelocation, HWImportLog, ProjectAssignment, ProjectCatalog, ProjectPO, ProjectSite
from app.services.auth_service import hash_password, validate_password_policy
from app.services.audit_service import write_audit_log
from app.services.email_service import queue_and_send_email, welcome_email


router = APIRouter(prefix="/api", tags=["Employees"])

PROJECT_ROLE_LABELS = {
    "DTE": "Drive Test Engineer",
    "DTA": "Drive Test Analysis Engineer",
    "REPORT_PREP": "Report Preparation Engineer",
    "TE": "TE Engineer",
    "TEAM_LEAD": "Team Lead",
    "SUPERVISOR": "Supervisor",
    "OTHER": "Project Support",
}


def _project_team_from_role(role: str | None) -> str:
    if role == "TE":
        return "TE"
    return "RF"


def _project_position_from_role(role: str | None) -> str:
    return PROJECT_ROLE_LABELS.get(role or "", role or "Project Staff")


EMPLOYEE_HR_COLUMNS = {
    "first_name": "VARCHAR(80)",
    "last_name": "VARCHAR(80)",
    "preferred_name": "VARCHAR(80)",
    "personal_email": "VARCHAR(150)",
    "phone": "VARCHAR(40)",
    "work_phone": "VARCHAR(40)",
    "job_title": "VARCHAR(120)",
    "job_level": "VARCHAR(50)",
    "section_name": "VARCHAR(80)",
    "manager_name": "VARCHAR(150)",
    "cost_center": "VARCHAR(80)",
    "work_location": "VARCHAR(150)",
    "employment_type": "VARCHAR(40)",
    "contract_type": "VARCHAR(40)",
    "probation_end_date": "DATE",
    "termination_date": "DATE",
    "date_of_birth": "DATE",
    "gender": "VARCHAR(20)",
    "nationality": "VARCHAR(80)",
    "id_card_no": "VARCHAR(40)",
    "address": "TEXT",
    "emergency_contact_name": "VARCHAR(150)",
    "emergency_contact_relation": "VARCHAR(80)",
    "emergency_contact_phone": "VARCHAR(40)",
    "bank_name": "VARCHAR(100)",
    "bank_account_no": "VARCHAR(60)",
    "bank_account_name": "VARCHAR(150)",
    "photo_url": "VARCHAR(300)",
    "notes": "TEXT",
}

PROJECT_ASSIGNMENT_COLUMNS = {
    "clock_type": "VARCHAR(20) NOT NULL DEFAULT 'DAILY'",
}

PROJECT_SITE_COLUMNS = {
    "full_on_air":      "DATE",
    "cluster_ready":    "DATE",
    "rf_cluster_name":  "VARCHAR(200)",
    "rf_cluster_owner": "VARCHAR(200)",
    "province":         "VARCHAR(100)",
    "district":         "VARCHAR(100)",
}

PROJECT_PO_COLUMNS = {
    "ace_project_code": "VARCHAR(30)",
    "hw_id": "VARCHAR(50)",
    "vendor": "VARCHAR(20) DEFAULT 'HW'",
    "hw_status_changed_at": "TIMESTAMP WITH TIME ZONE",
    "hw_prev_status": "VARCHAR(40)",
    "hw_first_seen_at": "TIMESTAMP WITH TIME ZONE",
    "ac1_plan_month": "VARCHAR(7)",
    "ac1_billed_at": "TIMESTAMP WITH TIME ZONE",
    "ac1_invoice_no": "VARCHAR(80)",
    "ac2_plan_month": "VARCHAR(7)",
    "ac2_billed_at": "TIMESTAMP WITH TIME ZONE",
    "ac2_invoice_no": "VARCHAR(80)",
    "workflow_status": "VARCHAR(40) NOT NULL DEFAULT 'NEW'",
    "mapping_confidence": "INTEGER NOT NULL DEFAULT 0",
    "mapping_rule": "VARCHAR(120)",
    "need_mapping_review": "BOOLEAN NOT NULL DEFAULT TRUE",
    "current_owner_role": "VARCHAR(40) NOT NULL DEFAULT 'FINANCE'",
    "current_owner_user": "VARCHAR(120)",
    "hold_reason": "TEXT",
    "expected_release_date": "DATE",
    "finance_checked_at": "TIMESTAMP WITH TIME ZONE",
    "sent_to_project_at": "TIMESTAMP WITH TIME ZONE",
    "project_accepted_at": "TIMESTAMP WITH TIME ZONE",
    "approved_at": "TIMESTAMP WITH TIME ZONE",
    "revision": "INTEGER NOT NULL DEFAULT 1",
    "locked": "BOOLEAN NOT NULL DEFAULT FALSE",
    "last_action_at": "TIMESTAMP WITH TIME ZONE",
}

PO_WORKFLOW_STATUSES = {
    "NEW",
    "AUTO_MAPPED",
    "NEED_REVIEW",              # Finance needs to fill project_code / work_type
    "NEED_MAPPING_REVIEW",
    "PENDING_SITE_MAP",       # Project to map PO → site_code
    "SITE_MAPPED",            # Site mapped, ready to plan
    "PLANNED",                # DTE assigned + scheduled
    "IN_PROGRESS",            # DTE executing in field
    "WORK_DONE",              # Field work complete, pending reconcile
    "PENDING_APPROVAL",       # Approver sign-off
    "APPROVED",               # Locked — ready to invoice
    "ON_HOLD",                # Paused (finance / project)
    "REJECTED",               # Approver rejected
    "RETURNED_TO_FINANCE",    # Returned from project/approver
    # Plan DTE
    "PLANNED",                # DTE assigned + scheduled
    "IN_PROGRESS",            # DTE executing in field
    "WORK_DONE",              # Field work complete
    # Leader Review
    "LEADER_CHECKING",        # Leader reviewing work + timesheet
    "LEADER_APPROVED",        # Leader approved — pending Finance pay + bill
    # Finance close-out (2 parallel tracks)
    "PENDING_PAYMENT",        # Finance to pay DTE (triggered by leader timesheet confirm)
    "PENDING_BILLING",        # Finance to bill HW (triggered by leader HW evidence confirm)
    "DTE_PAID",               # DTE payment done
    "HW_BILLED",              # HW invoice issued
    "CLOSED",                 # Both DTE_PAID + HW_BILLED
    # Legacy statuses (kept for existing rows)
    "FINANCE_CHECKING",
    "FINANCE_HOLD",
    "READY_TO_SEND_PROJECT",
    "WAITING_PROJECT_CHECK",
    "PROJECT_ADJUSTING",
    "PROJECT_ACCEPTED",
    "FINANCE_RECHECK",
    "RETURNED_TO_PROJECT",
    "LOCKED_FOR_WORK",
}

PO_WORK_TYPES = {"SSV", "PAC"}


class EmployeeIn(BaseModel):
    employee_code: str | None = None
    email: EmailStr | None = None
    full_name: str
    first_name: str | None = None
    last_name: str | None = None
    preferred_name: str | None = None
    personal_email: EmailStr | None = None
    phone: str | None = None
    work_phone: str | None = None
    department: str = "Project"
    section_name: str | None = None
    job_title: str | None = None
    job_level: str | None = None
    manager_name: str | None = None
    cost_center: str | None = None
    work_location: str | None = None
    project_team: str = "RF"
    project_role: str | None = None
    project_code: str | None = None
    project_name: str | None = None
    position: str | None = None
    status: str = "ACTIVE"
    employment_type: str | None = None
    contract_type: str | None = None
    contract_start_date: str | None = None
    contract_end_date: str | None = None
    contract_duration_months: int | None = None
    hire_date: str | None = None
    probation_end_date: str | None = None
    termination_date: str | None = None
    date_of_birth: str | None = None
    gender: str | None = None
    nationality: str | None = None
    id_card_no: str | None = None
    address: str | None = None
    emergency_contact_name: str | None = None
    emergency_contact_relation: str | None = None
    emergency_contact_phone: str | None = None
    bank_name: str | None = None
    bank_account_no: str | None = None
    bank_account_name: str | None = None
    notes: str | None = None
    create_login: bool = True
    send_welcome_email: bool = True
    password: str | None = None
    system_role: str | None = None   # PM | DC | HR | BOSS | EMPLOYEE — overrides auto-derived role


class AssignmentIn(BaseModel):
    employee_code: str
    role_in_project: str = "DTE"
    clock_type: str | None = None
    job_level: str | None = None
    start_date: str | None = None
    end_date: str | None = None
    allocation_pct: int = 100
    is_active: bool = True
    notes: str | None = None


class AssignmentUpdateIn(BaseModel):
    role_in_project: str | None = None
    clock_type: str | None = None
    job_level: str | None = None
    allocation_pct: int | None = None
    start_date: str | None = None
    end_date: str | None = None
    is_active: bool | None = None
    notes: str | None = None


class TransferIn(BaseModel):
    from_assignment_id: int
    to_project_code: str
    role_in_project: str = "DTE"
    clock_type: str | None = None
    job_level: str | None = None
    start_date: str | None = None
    end_date: str | None = None
    allocation_pct: int = 100


class SiteIn(BaseModel):
    site_code: str
    site_name: str | None = None
    customer: str | None = None
    lat: float | None = None
    lng: float | None = None
    gps_radius_m: int = 500
    province: str | None = None
    district: str | None = None
    is_active: bool = True


class POIn(BaseModel):
    po_target: str = "RF"
    project_code: str | None = None
    po_number: str
    po_line: str | None = None
    du_id: str | None = None
    item_dis: str | None = None
    cluster_site: str | None = None
    owner: str | None = None
    lat_long: str | None = None
    on_air: str | None = None
    cluster_type: str | None = None
    work_type: str | None = None  # SSV | PAC
    workflow_status: str | None = None
    mapping_confidence: int | None = None
    mapping_rule: str | None = None
    need_mapping_review: bool | None = None
    hold_reason: str | None = None
    expected_release_date: str | None = None


class POWorkflowIn(BaseModel):
    action: str
    note: str | None = None
    hold_reason: str | None = None
    expected_release_date: str | None = None
    current_owner_user: str | None = None
    site_code: str | None = None
    project_code: str | None = None
    work_type: str | None = None
    # Plan DTE
    planned_dte_codes: str | None = None
    planned_dte_names: str | None = None
    planned_start_date: str | None = None
    planned_end_date: str | None = None
    # Leader Review
    leader_code: str | None = None
    leader_note: str | None = None
    hw_evidence_url: str | None = None
    # Finance close-out
    hw_invoice_no: str | None = None


def _date_or_none(value: str | None):
    if not value:
        return None
    from datetime import date
    try:
        return date.fromisoformat(str(value)[:10])
    except (ValueError, TypeError):
        return None


def _now_utc() -> datetime:
    return datetime.now(timezone.utc)


def _po_aging_days(row: ProjectPO) -> int:
    basis = row.last_action_at or row.created_at
    if not basis:
        return 0
    if basis.tzinfo is None:
        basis = basis.replace(tzinfo=timezone.utc)
    return max(0, (_now_utc() - basis).days)


# --- PO Collection Tracking helpers ----------------------------------------
# Two parallel money legs are tracked per PO:
#   • Collection IN  — billing Huawei (hw_billed_at / AC1 / AC2 invoices)
#   • Payment OUT    — paying the DTE field engineer (dte_paid_at)
# These derive a coarse status used by the Collection Tracking dashboard so
# Finance/Project/Executives can answer "what's collected, what's pending,
# what's rejected" at a glance.

# Workflow statuses grouped into phases for the pipeline funnel.
PO_PHASE_MAP = {
    "FINANCE_REVIEW": {"NEW", "AUTO_MAPPED", "NEED_REVIEW", "NEED_MAPPING_REVIEW",
                       "FINANCE_CHECKING", "FINANCE_HOLD", "FINANCE_RECHECK", "RETURNED_TO_FINANCE"},
    "PROJECT_PLAN":   {"PENDING_SITE_MAP", "SITE_MAPPED", "READY_TO_SEND_PROJECT",
                       "WAITING_PROJECT_CHECK", "PROJECT_ADJUSTING", "PROJECT_ACCEPTED",
                       "RETURNED_TO_PROJECT"},
    "APPROVAL":       {"PENDING_APPROVAL", "APPROVED", "LOCKED_FOR_WORK"},
    "EXECUTION":      {"PLANNED", "IN_PROGRESS", "WORK_DONE", "LEADER_CHECKING", "LEADER_APPROVED"},
    "CLOSE_OUT":      {"PENDING_PAYMENT", "PENDING_BILLING", "DTE_PAID", "HW_BILLED", "CLOSED"},
    "HOLD":           {"ON_HOLD"},
    "REJECTED":       {"REJECTED"},
}


def _po_phase(status: str | None) -> str:
    s = (status or "").upper()
    for phase, members in PO_PHASE_MAP.items():
        if s in members:
            return phase
    return "FINANCE_REVIEW"


def _po_billing_state(row: ProjectPO) -> str:
    """Collection-IN state: has Finance billed Huawei for this PO yet?

    REJECTED   — PO was rejected, nothing to collect
    BILLED     — fully invoiced (hw_billed / closed, or both AC milestones done)
    PARTIAL    — first invoice issued (AC1) but not the final one
    NOT_BILLED — nothing invoiced yet
    """
    status = (row.workflow_status or "").upper()
    if status == "REJECTED":
        return "REJECTED"
    fully_billed = bool(row.hw_billed_at) or status in {"HW_BILLED", "CLOSED"} or (
        bool(row.ac1_billed_at) and bool(row.ac2_billed_at)
    )
    if fully_billed:
        return "BILLED"
    if row.ac1_billed_at or row.ac2_billed_at:
        return "PARTIAL"
    return "NOT_BILLED"


def _po_dte_pay_state(row: ProjectPO) -> str:
    """Payment-OUT state: has Finance paid the DTE for this PO yet?"""
    status = (row.workflow_status or "").upper()
    if status == "REJECTED":
        return "N/A"
    if row.dte_paid_at or status in {"DTE_PAID", "CLOSED"}:
        return "PAID"
    return "UNPAID"


def _po_months(row: ProjectPO) -> set[str]:
    """All "YYYY-MM" months this PO touches — billing plan, actual billed, paid,
    or on-air. Used by the month-range filter so both planned and collected POs
    surface within the selected window."""
    months: set[str] = set()
    for v in (row.ac1_plan_month, row.ac2_plan_month):
        if v:
            months.add(v.strip())
    for dt in (row.ac1_billed_at, row.ac2_billed_at, row.hw_billed_at, row.dte_paid_at):
        if dt:
            months.add(dt.strftime("%Y-%m"))
    if row.on_air:
        months.add(row.on_air.strftime("%Y-%m"))
    return months


def _infer_po_workflow_defaults(body: POIn) -> dict:
    if body.workflow_status:
        status = body.workflow_status.upper()
    elif body.project_code and body.work_type:
        # ระบุ project_code + work_type ครบ → Auto-map สำเร็จ → Finance รีวิว
        status = "AUTO_MAPPED"
    elif body.project_code:
        # มี project_code แต่ไม่มี work_type → Finance กรอกเพิ่ม
        status = "NEED_REVIEW"
    else:
        # ไม่มี project_code → Finance ต้องระบุ
        status = "NEED_REVIEW"

    if status not in PO_WORKFLOW_STATUSES:
        raise HTTPException(400, f"Invalid PO workflow_status '{status}'.")

    confidence = body.mapping_confidence
    if confidence is None:
        confidence = 90 if status == "AUTO_MAPPED" else 20

    owner = "FINANCE"  # Finance reviews first before sending to Project

    return {
        "workflow_status": status,
        "mapping_confidence": max(0, min(100, int(confidence or 0))),
        "mapping_rule": body.mapping_rule or (
            "Auto-mapped by project_code + work_type" if status == "AUTO_MAPPED"
            else "Needs Finance review"
        ),
        "need_mapping_review": status == "NEED_REVIEW",
        "current_owner_role": owner,
        "last_action_at": _now_utc(),
        "expected_release_date": _date_or_none(body.expected_release_date),
    }


def employee_to_dict(row: Employee) -> dict:
    return {
        "id": row.id,
        "employee_code": row.employee_code,
        "email": row.email,
        "full_name": row.full_name,
        "first_name": row.first_name,
        "last_name": row.last_name,
        "preferred_name": row.preferred_name,
        "personal_email": row.personal_email,
        "phone": row.phone,
        "work_phone": row.work_phone,
        "department": row.department,
        "job_title": row.job_title,
        "job_level": row.job_level,
        "manager_name": row.manager_name,
        "cost_center": row.cost_center,
        "work_location": row.work_location,
        "project_team": row.project_team,
        "section_name": row.section_name,
        "project_role": row.project_role,
        "project_code": row.project_code,
        "project_name": row.project_name,
        "position": row.position,
        "status": row.status,
        "employment_type": row.employment_type,
        "contract_type": row.contract_type,
        "contract_start_date": row.contract_start_date.isoformat() if row.contract_start_date else None,
        "contract_end_date": row.contract_end_date.isoformat() if row.contract_end_date else None,
        "contract_duration_months": row.contract_duration_months,
        "hire_date": row.hire_date.isoformat() if row.hire_date else None,
        "probation_end_date": row.probation_end_date.isoformat() if row.probation_end_date else None,
        "termination_date": row.termination_date.isoformat() if row.termination_date else None,
        "date_of_birth": row.date_of_birth.isoformat() if row.date_of_birth else None,
        "gender": row.gender,
        "nationality": row.nationality,
        "id_card_no": row.id_card_no,
        "address": row.address,
        "base_lat": getattr(row, "base_lat", None),
        "base_lng": getattr(row, "base_lng", None),
        "emergency_contact_name": row.emergency_contact_name,
        "emergency_contact_relation": row.emergency_contact_relation,
        "emergency_contact_phone": row.emergency_contact_phone,
        "bank_name": row.bank_name,
        "bank_account_no": row.bank_account_no,
        "bank_account_name": row.bank_account_name,
        "photo_url": row.photo_url,
        "notes": row.notes,
        "source": row.source,
        "created_at": row.created_at.isoformat() if row.created_at else None,
        "updated_at": row.updated_at.isoformat() if row.updated_at else None,
    }


def project_to_dict(row: ProjectCatalog) -> dict:
    project_type = row.team if row.team in {"RF", "TE", "DTE", "DTA"} else "OTHER"
    customer = _customer_from_project(row.project_code)
    return {
        "id": row.id,
        "project_code": row.project_code,
        "project_name": row.project_name,
        "project_type": project_type,
        "status": "ACTIVE",
        "start_date": "2026-01-01",
        "end_date": None,
        "customer": {"id": 0, "code": customer, "name": customer},
        "team": {"code": row.team, "name": f"{row.team} Team"},
        "headcount": row.headcount,
        "site_count": 0,
    }


def project_list_to_dict(row: ProjectCatalog, headcount: int = 0, site_count: int = 0) -> dict:
    project_type = row.team if row.team in {"RF", "TE", "DTE", "DTA"} else "OTHER"
    customer = _customer_from_project(row.project_code)
    return {
        "id": row.id,
        "project_code": row.project_code,
        "project_name": row.project_name,
        "project_type": project_type,
        "status": "ACTIVE",
        "start_date": "2026-01-01",
        "end_date": None,
        "customer": {"id": 0, "code": customer, "name": customer},
        "team": {"code": row.team, "name": f"{row.team} Team"},
        "headcount": headcount or row.headcount,
        "site_count": site_count,
    }


def assignment_to_dict(row: ProjectAssignment, employee: Employee | None = None) -> dict:
    clock_type = _normalize_assignment_clock_type(row.role_in_project, getattr(row, "clock_type", None))
    return {
        "id": row.id,
        "project_code": row.project_code,
        "employee_code": row.employee_code,
        "full_name": employee.full_name if employee else row.employee_code,
        "email": employee.email if employee else None,
        "position_name": employee.position or employee.job_title or employee.project_team if employee else "",
        "project_team": employee.project_team if employee else "",
        "job_level": row.job_level,
        "role_in_project": row.role_in_project,
        "clock_type": clock_type,
        "start_date": row.start_date.isoformat() if row.start_date else None,
        "end_date": row.end_date.isoformat() if row.end_date else None,
        "allocation_pct": row.allocation_pct,
        "is_active": row.is_active,
        "notes": row.notes,
    }


def site_to_dict(row: ProjectSite | ClockSite) -> dict:
    return {
        "id": row.id,
        "project_code": row.project_code,
        "site_code": row.site_code,
        "site_name": row.site_name,
        "customer": row.customer,
        "latitude": row.lat,
        "longitude": row.lng,
        "lat": row.lat,
        "lng": row.lng,
        "gps_radius_m": row.gps_radius_m,
        "gpsRadiusM": row.gps_radius_m,
        "province": getattr(row, "province", None),
        "district": getattr(row, "district", None),
        "is_active": row.is_active,
        "full_on_air":     row.full_on_air.isoformat() if getattr(row, "full_on_air", None) else None,
        "cluster_ready":   row.cluster_ready.isoformat() if getattr(row, "cluster_ready", None) else None,
        "rf_cluster_name": getattr(row, "rf_cluster_name", None),
        "rf_cluster_owner": getattr(row, "rf_cluster_owner", None),
    }


def po_to_dict(row: ProjectPO) -> dict:
    return {
        "id": row.id,
        "po_target": row.po_target,
        "project_code": row.project_code,
        "po_number": row.po_number,
        "po_line": row.po_line,
        "du_id": row.du_id,
        "item_dis": row.item_dis,
        "cluster_site": row.cluster_site,
        "owner": row.owner,
        "lat_long": row.lat_long,
        "on_air": row.on_air.isoformat() if row.on_air else None,
        "cluster_type": row.cluster_type,
        "work_type": getattr(row, "work_type", None),
        "site_code": getattr(row, "site_code", None),
        "project": row.project_code,
        "workflow_status": getattr(row, "workflow_status", None) or "NEW",
        "mapping_confidence": getattr(row, "mapping_confidence", 0) or 0,
        "mapping_rule": getattr(row, "mapping_rule", None),
        "need_mapping_review": bool(getattr(row, "need_mapping_review", True)),
        "current_owner_role": getattr(row, "current_owner_role", None) or "FINANCE",
        "current_owner_user": getattr(row, "current_owner_user", None),
        "hold_reason": getattr(row, "hold_reason", None),
        "expected_release_date": row.expected_release_date.isoformat() if getattr(row, "expected_release_date", None) else None,
        "finance_checked_at": row.finance_checked_at.isoformat() if getattr(row, "finance_checked_at", None) else None,
        "sent_to_project_at": row.sent_to_project_at.isoformat() if getattr(row, "sent_to_project_at", None) else None,
        "project_accepted_at": row.project_accepted_at.isoformat() if getattr(row, "project_accepted_at", None) else None,
        "approved_at": row.approved_at.isoformat() if getattr(row, "approved_at", None) else None,
        "revision": getattr(row, "revision", 1) or 1,
        "locked": bool(getattr(row, "locked", False)),
        "last_action_at": row.last_action_at.isoformat() if getattr(row, "last_action_at", None) else None,
        "aging_days": _po_aging_days(row),
        # Plan DTE
        "planned_dte_codes": getattr(row, "planned_dte_codes", None),
        "planned_dte_names": getattr(row, "planned_dte_names", None),
        "planned_start_date": row.planned_start_date.isoformat() if getattr(row, "planned_start_date", None) else None,
        "planned_end_date": row.planned_end_date.isoformat() if getattr(row, "planned_end_date", None) else None,
        "planned_duration_days": float(row.planned_duration_days) if getattr(row, "planned_duration_days", None) is not None else None,
        "planned_at": row.planned_at.isoformat() if getattr(row, "planned_at", None) else None,
        "work_started_at": row.work_started_at.isoformat() if getattr(row, "work_started_at", None) else None,
        "work_done_at": row.work_done_at.isoformat() if getattr(row, "work_done_at", None) else None,
        # Leader Review
        "leader_code": getattr(row, "leader_code", None),
        "leader_checked_at": row.leader_checked_at.isoformat() if getattr(row, "leader_checked_at", None) else None,
        "leader_note": getattr(row, "leader_note", None),
        # HW Evidence
        "hw_evidence_url": getattr(row, "hw_evidence_url", None),
        "hw_evidence_confirmed_at": row.hw_evidence_confirmed_at.isoformat() if getattr(row, "hw_evidence_confirmed_at", None) else None,
        # Finance Pay DTE
        "dte_paid_at": row.dte_paid_at.isoformat() if getattr(row, "dte_paid_at", None) else None,
        "dte_paid_by": getattr(row, "dte_paid_by", None),
        # Finance Bill HW
        "hw_billed_at": row.hw_billed_at.isoformat() if getattr(row, "hw_billed_at", None) else None,
        "hw_billed_by": getattr(row, "hw_billed_by", None),
        "hw_invoice_no": getattr(row, "hw_invoice_no", None),
        "line_amount": float(row.line_amount) if getattr(row, "line_amount", None) is not None else None,
        "payment_terms": getattr(row, "payment_terms", None),
        "hw_po_status": getattr(row, "hw_po_status", None),
        "vendor": getattr(row, "vendor", None) or "HW",
        "hw_data": getattr(row, "hw_data", None),
        "ace_project_code": getattr(row, "ace_project_code", None),
        "hw_id": getattr(row, "hw_id", None),
        "hw_status_changed_at": row.hw_status_changed_at.isoformat() if getattr(row, "hw_status_changed_at", None) else None,
        "hw_prev_status": getattr(row, "hw_prev_status", None),
        "hw_first_seen_at": row.hw_first_seen_at.isoformat() if getattr(row, "hw_first_seen_at", None) else None,
        "ac1_plan_month": getattr(row, "ac1_plan_month", None),
        "ac1_billed_at": row.ac1_billed_at.isoformat() if getattr(row, "ac1_billed_at", None) else None,
        "ac1_invoice_no": getattr(row, "ac1_invoice_no", None),
        "ac2_plan_month": getattr(row, "ac2_plan_month", None),
        "ac2_billed_at": row.ac2_billed_at.isoformat() if getattr(row, "ac2_billed_at", None) else None,
        "ac2_invoice_no": getattr(row, "ac2_invoice_no", None),
    }


def _customer_from_project(project_code: str | None) -> str:
    text = (project_code or "").upper()
    for prefix in ("AIS", "TRUE", "HWT", "ZTE", "NBTC", "NT", "WW", "DTAC"):
        if text.startswith(prefix):
            return prefix
    return "ACE"


def _role_from_employee(row: Employee) -> str:
    text = f"{row.project_team or ''} {row.project_role or ''} {row.position or ''}".upper()
    if "REPORT" in text:
        return "REPORT_PREP"
    if "DTA" in text or "ANALYSIS" in text or "ANALYST" in text:
        return "DTA"
    if "DTE" in text or "DRIVE TEST ENGINEER" in text or row.project_team == "RF":
        return "DTE"
    if row.project_team == "TE" or "SITE" in text or "SYSTEM" in text:
        return "TE"
    return "OTHER"


def _generate_password(length: int = 12) -> str:
    alphabet = string.ascii_letters + string.digits
    while True:
        password = "".join(secrets.choice(alphabet) for _ in range(length))
        if any(c.islower() for c in password) and any(c.isupper() for c in password) and any(c.isdigit() for c in password):
            return password


def _validate_initial_password(password: str, *, email: str | None = None, employee_code: str | None = None) -> None:
    try:
        validate_password_policy(password, email=email, employee_code=employee_code)
    except ValueError as exc:
        raise HTTPException(400, str(exc))


_VALID_SYSTEM_ROLES = {"EMPLOYEE", "PM", "DC", "HR", "BOSS", "ADMIN", "HR_ADMIN", "HR_VIEWER", "PROJECT_ADMIN", "DIRECTOR", "ACCOUNTING", "SYSTEM_ADMIN"}

def _auth_role_for_employee(row: Employee, system_role: str | None = None) -> str:
    if system_role and system_role.upper() in _VALID_SYSTEM_ROLES:
        requested = system_role.upper()
        legacy = {"HR": "HR_ADMIN", "BOSS": "DIRECTOR", "ADMIN": "SYSTEM_ADMIN", "DC": "DIRECTOR"}
        return legacy.get(requested, requested)
    if row.department == "Human Resources":
        return "HR"
    return "EMPLOYEE"


# DTA = office analyst (DAILY) — not field, does NOT go to sites
_PER_SITE_ROLES = {"DTE"}
_PHOTO_ROLES    = _PER_SITE_ROLES  # all field roles require selfie

def _normalize_assignment_clock_type(role: str | None, clock_type: str | None) -> str:
    if role != "DTE":
        return "DAILY"
    normalized = (clock_type or "PER_SITE").upper()
    if normalized not in {"DAILY", "PER_SITE"}:
        raise HTTPException(400, "clock_type must be DAILY or PER_SITE for DTE assignments.")
    return normalized


def _clock_config_for_employee(row: Employee) -> tuple[str, bool, bool]:
    role = _role_from_employee(row)
    # Auto-derive PER_SITE from contract_type (e.g., "Subcontractor Per-Site")
    contract = (row.contract_type or "").lower()
    if "per-site" in contract or "per site" in contract:
        return "PER_SITE", True, True
    # GPS required for everyone (DAILY office checks work_location radius;
    # PER_SITE checks selected site radius)
    if role in _PER_SITE_ROLES:
        return "DAILY", True, True
    return "DAILY", True, False


async def _clock_config_from_project_assignments(db: AsyncSession, row: Employee) -> tuple[str, bool, bool]:
    assignments = (await db.execute(
        select(ProjectAssignment).where(
            ProjectAssignment.employee_code == row.employee_code,
            ProjectAssignment.is_active.is_(True),
        )
    )).scalars().all()
    dte_assignments = [a for a in assignments if a.role_in_project == "DTE"]
    if not dte_assignments:
        return "DAILY", True, False
    is_per_site = any(_normalize_assignment_clock_type(a.role_in_project, a.clock_type) == "PER_SITE" for a in dte_assignments)
    # GPS required regardless (PER_SITE checks site radius, DAILY checks work location)
    return ("PER_SITE", True, True) if is_per_site else ("DAILY", True, True)


async def _sync_auth_user_from_employee(db: AsyncSession, row: Employee) -> None:
    auth_user = (
        await db.execute(select(AuthUser).where(AuthUser.employee_code == row.employee_code))
    ).scalar_one_or_none()
    if not auth_user:
        return
    role = _role_from_employee(row)
    clock_type, gps_required, photo_required = await _clock_config_from_project_assignments(db, row)
    auth_user.department = row.department or auth_user.department
    auth_user.team = row.project_team or auth_user.team
    auth_user.position_code = role
    auth_user.position_name = row.position or _project_position_from_role(role)
    auth_user.clock_type = clock_type
    auth_user.gps_required = gps_required
    auth_user.photo_required = photo_required


@router.get("/employees/next-code")
async def next_employee_code(
    payload: dict = Depends(require_hr_user),
    db: AsyncSession = Depends(get_db),
):
    """Return a suggested next employee code based on existing ACE### codes."""
    rows = (await db.execute(
        select(Employee.employee_code).where(Employee.employee_code.regexp_match(r'^ACE\d+$'))
    )).scalars().all()
    nums = []
    for code in rows:
        try:
            nums.append(int(code[3:]))
        except ValueError:
            pass
    next_num = (max(nums) + 1) if nums else 700
    return {"next_code": f"ACE{next_num:03d}"}


@router.get("/employees")
async def list_employees(
    search: str = Query(""),
    department: str = Query(""),
    project_team: str = Query(""),
    status: str = Query(""),
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(Employee)
    if search:
        q = f"%{search}%"
        stmt = stmt.where(or_(Employee.full_name.ilike(q), Employee.email.ilike(q), Employee.employee_code.ilike(q)))
    if department:
        stmt = stmt.where(Employee.department == department)
    if project_team:
        stmt = stmt.where(Employee.project_team == project_team)
    if status:
        stmt = stmt.where(Employee.status == status)
    rows = (await db.execute(stmt.order_by(Employee.full_name))).scalars().all()
    return {"data": [employee_to_dict(row) for row in rows], "total": len(rows)}


class SystemAccessPatchIn(BaseModel):
    is_active: bool | None = None
    role: str | None = None
    must_change_password: bool | None = None
    locked: bool | None = None


class ResetPasswordIn(BaseModel):
    new_password: str | None = None
    must_change_password: bool = True


@router.get("/employees/{employee_id}")
async def get_employee(
    employee_id: int,
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    row = (await db.execute(select(Employee).where(Employee.id == employee_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Employee not found")
    return employee_to_dict(row)


@router.get("/employees/{employee_id}/system-access")
async def get_employee_system_access(
    employee_id: int,
    payload: dict = Depends(require_hr_read_user),
    db: AsyncSession = Depends(get_db),
):
    row = (await db.execute(select(Employee).where(Employee.id == employee_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Employee not found")
    return await _system_access_dict(db, row)


@router.post("/employees/{employee_id}/create-login")
async def create_employee_login(
    employee_id: int,
    request: Request,
    payload: dict = Depends(require_hr_user),
    db: AsyncSession = Depends(get_db),
):
    row = (await db.execute(select(Employee).where(Employee.id == employee_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Employee not found")
    if row.status in {"TERMINATED", "RESIGNED", "ARCHIVED"}:
        raise HTTPException(400, "Terminated or archived employees cannot have active login.")
    if not row.email:
        raise HTTPException(400, "Company email is required before login creation.")
    exists = (await db.execute(select(AuthUser).where(AuthUser.employee_code == row.employee_code))).scalar_one_or_none()
    if exists:
        raise HTTPException(400, "Login already exists.")
    password = _generate_password()
    _validate_initial_password(password, email=row.email, employee_code=row.employee_code)
    clock_type, gps_required, photo_required = _clock_config_for_employee(row)
    names = (row.full_name or row.employee_code).split()
    user = AuthUser(
        employee_code=row.employee_code,
        password_hash=hash_password(password),
        first_name=row.first_name or names[0],
        last_name=row.last_name or " ".join(names[1:]),
        email=row.email,
        department=row.department,
        team=row.project_team,
        position_code=_role_from_employee(row),
        position_name=row.position or row.job_title or "Employee",
        role=_auth_role_for_employee(row, None),
        clock_type=clock_type,
        gps_required=gps_required,
        photo_required=photo_required,
        work_location_name=row.work_location or row.project_code or "ACE Head Office",
        allowed_radius_m=300,
        must_change_password=True,
        created_by=payload.get("id"),
    )
    db.add(user)
    await write_audit_log(db, action="login_created", entity_type="auth_user", entity_id=row.employee_code, employee_id=row.id, payload=payload, new_value={"employee_code": row.employee_code, "email": row.email, "role": user.role}, changed_fields=["employee_code", "email", "role"], request=request, source="Employee API")
    await db.commit()
    access = await _system_access_dict(db, row)
    return access | {"initial_password": password}


class BaseLocationPatchIn(BaseModel):
    base_lat: float | None = None
    base_lng: float | None = None
    address: str | None = None


@router.patch("/employees/by-code/{employee_code}/base-location")
async def patch_employee_base_location(
    employee_code: str,
    body: BaseLocationPatchIn,
    request: Request,
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    """Update DTE Home Base (lat/lng + optional address) used by Pre-Site Plan view.

    Allowed roles: PROJECT_ADMIN, PM, HR_ADMIN, SUPER_ADMIN, SYSTEM_ADMIN
    (require_project_user covers PROJECT + admin tiers).
    """
    row = (await db.execute(
        select(Employee).where(Employee.employee_code == employee_code)
    )).scalar_one_or_none()
    if not row:
        raise HTTPException(404, f"Employee {employee_code} not found")

    # Light validation: lat in [-90,90], lng in [-180,180]; allow None to clear
    if body.base_lat is not None and not (-90 <= body.base_lat <= 90):
        raise HTTPException(400, "base_lat must be in [-90, 90]")
    if body.base_lng is not None and not (-180 <= body.base_lng <= 180):
        raise HTTPException(400, "base_lng must be in [-180, 180]")

    old = {"base_lat": row.base_lat, "base_lng": row.base_lng, "address": row.address}
    changed = []
    patch = body.model_dump(exclude_unset=True)
    for field, value in patch.items():
        if getattr(row, field) != value:
            setattr(row, field, value)
            changed.append(field)
    if changed:
        new = {f: getattr(row, f) for f in changed}
        await write_audit_log(
            db,
            action="employee_base_location_updated",
            entity_type="employee",
            entity_id=row.id,
            employee_id=row.id,
            payload=payload,
            old_value={f: old.get(f) for f in changed},
            new_value=new,
            changed_fields=changed,
            request=request,
            source="Pre-Site Plan",
        )
    await db.commit()
    await db.refresh(row)
    return {
        "employee_code": row.employee_code,
        "base_lat": row.base_lat,
        "base_lng": row.base_lng,
        "address": row.address,
        "changed_fields": changed,
    }


@router.patch("/employees/{employee_id}/system-access")
async def patch_employee_system_access(
    employee_id: int,
    body: SystemAccessPatchIn,
    request: Request,
    payload: dict = Depends(require_hr_user),
    db: AsyncSession = Depends(get_db),
):
    row = (await db.execute(select(Employee).where(Employee.id == employee_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Employee not found")
    if row.status == "ARCHIVED":
        raise HTTPException(400, "Archived employees are read-only.")
    user = (await db.execute(select(AuthUser).where(AuthUser.employee_code == row.employee_code))).scalar_one_or_none()
    if not user:
        raise HTTPException(404, "Login not found")
    if body.is_active is True and row.status in {"TERMINATED", "RESIGNED", "ARCHIVED"}:
        raise HTTPException(400, "Terminated or archived employees must not have active login.")
    old = {"is_active": user.is_active, "role": user.role, "must_change_password": user.must_change_password, "locked_until": user.locked_until}
    changed = []
    if body.is_active is not None and user.is_active != body.is_active:
        user.is_active = body.is_active; changed.append("is_active")
    if body.role is not None and body.role.upper() in ROLE_SCOPES and user.role != body.role.upper():
        user.role = body.role.upper(); changed.append("role")
    if body.must_change_password is not None and user.must_change_password != body.must_change_password:
        user.must_change_password = body.must_change_password; changed.append("must_change_password")
    if body.locked is not None:
        user.locked_until = datetime.now(timezone.utc) + timedelta(days=365) if body.locked else None
        changed.append("locked_until")
    if changed:
        action = "login_deactivated" if "is_active" in changed and not user.is_active else "login_updated"
        if "locked_until" in changed:
            action = "login_locked" if body.locked else "login_unlocked"
        await write_audit_log(db, action=action, entity_type="auth_user", entity_id=user.id, employee_id=row.id, payload=payload, old_value=old, new_value={"is_active": user.is_active, "role": user.role, "must_change_password": user.must_change_password, "locked_until": user.locked_until}, changed_fields=changed, request=request, source="Employee API")
    await db.commit()
    return await _system_access_dict(db, row)


@router.post("/employees/{employee_id}/reset-password")
async def reset_employee_password(
    employee_id: int,
    body: ResetPasswordIn,
    request: Request,
    payload: dict = Depends(require_hr_user),
    db: AsyncSession = Depends(get_db),
):
    row = (await db.execute(select(Employee).where(Employee.id == employee_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Employee not found")
    user = (await db.execute(select(AuthUser).where(AuthUser.employee_code == row.employee_code))).scalar_one_or_none()
    if not user:
        raise HTTPException(404, "Login not found")
    password = body.new_password or _generate_password()
    _validate_initial_password(password, email=user.email, employee_code=user.employee_code)
    user.password_hash = hash_password(password)
    user.must_change_password = body.must_change_password
    user.failed_login_count = 0
    user.locked_until = None
    user.password_changed_at = datetime.now(timezone.utc)
    user.token_version = (user.token_version or 1) + 1
    await write_audit_log(db, action="password_reset_requested", entity_type="auth_user", entity_id=user.id, employee_id=row.id, payload=payload, new_value={"employee_code": user.employee_code, "must_change_password": user.must_change_password}, changed_fields=["password_hash", "must_change_password", "token_version"], request=request, source="Employee API")
    await db.commit()
    access = await _system_access_dict(db, row)
    return {"success": True, "initial_password": password, "system_access": access}


DOCUMENTS_DIR = Path("/app/photos/documents")
PHOTOS_DIR = Path("/app/photos/employee_avatars")
DOCUMENTS_DIR.mkdir(parents=True, exist_ok=True)
PHOTOS_DIR.mkdir(parents=True, exist_ok=True)

ALLOWED_DOC_MIMES = {
    "image/jpeg", "image/png", "image/webp",
    "application/pdf",
}
MAX_DOC_BYTES = 10 * 1024 * 1024  # 10 MB


@router.get("/employees/{employee_id}/documents")
async def list_employee_documents(
    employee_id: int,
    payload: dict = Depends(require_hr_read_user),
    db: AsyncSession = Depends(get_db),
):
    emp = (await db.execute(select(Employee).where(Employee.id == employee_id))).scalar_one_or_none()
    if not emp:
        raise HTTPException(404, "Employee not found")
    rows = (await db.execute(text(
        "SELECT id, doc_type, file_name, file_url, mime_type, size_bytes, notes, uploaded_at, uploaded_by "
        "FROM employee_documents WHERE employee_code=:c ORDER BY uploaded_at DESC"
    ), {"c": emp.employee_code})).mappings().all()
    return {"data": [dict(r) | {"uploaded_at": r["uploaded_at"].isoformat() if r["uploaded_at"] else None} for r in rows], "total": len(rows)}


@router.post("/employees/{employee_id}/documents", status_code=201)
async def upload_employee_document(
    employee_id: int,
    doc_type: str = Form(...),
    notes: str | None = Form(None),
    file: UploadFile = File(...),
    request: Request = None,
    payload: dict = Depends(require_hr_user),
    db: AsyncSession = Depends(get_db),
):
    emp = (await db.execute(select(Employee).where(Employee.id == employee_id))).scalar_one_or_none()
    if not emp:
        raise HTTPException(404, "Employee not found")
    if file.content_type and file.content_type not in ALLOWED_DOC_MIMES:
        raise HTTPException(400, f"Unsupported file type: {file.content_type}")

    content = await file.read()
    if len(content) > MAX_DOC_BYTES:
        raise HTTPException(400, f"File too large (max {MAX_DOC_BYTES // (1024*1024)} MB)")

    ext = (file.filename.rsplit(".", 1)[-1] if file.filename and "." in file.filename else "bin").lower()
    safe_name = f"{emp.employee_code}_{doc_type}_{uuid.uuid4().hex[:8]}.{ext}"
    full_path = DOCUMENTS_DIR / safe_name
    full_path.write_bytes(content)

    file_url = f"/photos/documents/{safe_name}"
    await db.execute(text(
        "INSERT INTO employee_documents (employee_code, doc_type, file_name, file_url, mime_type, size_bytes, notes, uploaded_by) "
        "VALUES (:c, :t, :fn, :u, :m, :s, :n, :ub)"
    ), {
        "c": emp.employee_code, "t": doc_type, "fn": file.filename or safe_name, "u": file_url,
        "m": file.content_type, "s": len(content), "n": notes, "ub": payload.get("employeeCode"),
    })
    await write_audit_log(
        db, action="employee_document_uploaded", entity_type="employee", entity_id=emp.id,
        employee_id=emp.id, payload=payload,
        new_value={"doc_type": doc_type, "file_name": file.filename, "size_bytes": len(content)},
        changed_fields=["doc_type", "file_url"], request=request, source="Employee API",
    )
    await db.commit()
    return {"status": "ok", "file_url": file_url, "doc_type": doc_type}


@router.delete("/employees/{employee_id}/documents/{doc_id}")
async def delete_employee_document(
    employee_id: int,
    doc_id: int,
    request: Request = None,
    payload: dict = Depends(require_hr_user),
    db: AsyncSession = Depends(get_db),
):
    row = (await db.execute(text(
        "SELECT id, employee_code, doc_type, file_url FROM employee_documents WHERE id=:i AND employee_code=(SELECT employee_code FROM employees WHERE id=:eid)"
    ), {"i": doc_id, "eid": employee_id})).mappings().first()
    if not row:
        raise HTTPException(404, "Document not found")
    # remove file (best-effort)
    try:
        Path(f"/app{row['file_url']}").unlink(missing_ok=True)
    except Exception:
        pass
    await db.execute(text("DELETE FROM employee_documents WHERE id=:i"), {"i": doc_id})
    await write_audit_log(
        db, action="employee_document_deleted", entity_type="employee", entity_id=employee_id,
        employee_id=employee_id, payload=payload,
        old_value={"doc_type": row["doc_type"], "file_url": row["file_url"]},
        changed_fields=["file_url"], request=request, source="Employee API",
    )
    await db.commit()
    return {"status": "ok"}


@router.post("/employees/{employee_id}/photo", status_code=201)
async def upload_employee_photo(
    employee_id: int,
    file: UploadFile = File(...),
    request: Request = None,
    payload: dict = Depends(require_hr_user),
    db: AsyncSession = Depends(get_db),
):
    emp = (await db.execute(select(Employee).where(Employee.id == employee_id))).scalar_one_or_none()
    if not emp:
        raise HTTPException(404, "Employee not found")
    if file.content_type not in {"image/jpeg", "image/png", "image/webp"}:
        raise HTTPException(400, f"Unsupported image type: {file.content_type}")
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(400, "Image too large (max 5 MB)")
    ext = (file.filename.rsplit(".", 1)[-1] if file.filename and "." in file.filename else "jpg").lower()
    safe_name = f"{emp.employee_code}_{uuid.uuid4().hex[:6]}.{ext}"
    full_path = PHOTOS_DIR / safe_name
    full_path.write_bytes(content)
    file_url = f"/photos/employee_avatars/{safe_name}"
    emp.photo_url = file_url
    await write_audit_log(
        db, action="employee_photo_updated", entity_type="employee", entity_id=emp.id,
        employee_id=emp.id, payload=payload,
        new_value={"photo_url": file_url, "size_bytes": len(content)},
        changed_fields=["photo_url"], request=request, source="Employee API",
    )
    await db.commit()
    return {"status": "ok", "photo_url": file_url}


@router.get("/employees/{employee_id}/contracts")
async def list_employee_contracts(
    employee_id: int,
    payload: dict = Depends(require_hr_read_user),
    db: AsyncSession = Depends(get_db),
):
    emp = (await db.execute(select(Employee).where(Employee.id == employee_id))).scalar_one_or_none()
    if not emp:
        raise HTTPException(404, "Employee not found")
    result = await db.execute(text(
        "SELECT id, employment_type, contract_type, duration_months, start_date, end_date, status, notes, created_at "
        "FROM employee_contracts WHERE employee_code=:c ORDER BY start_date DESC NULLS LAST, id DESC"
    ), {"c": emp.employee_code})
    items = []
    for r in result.mappings():
        items.append({
            "id": r["id"],
            "employment_type": r["employment_type"],
            "contract_type": r["contract_type"],
            "duration_months": r["duration_months"],
            "start_date": r["start_date"].isoformat() if r["start_date"] else None,
            "end_date": r["end_date"].isoformat() if r["end_date"] else None,
            "status": r["status"],
            "notes": r["notes"],
            "created_at": r["created_at"].isoformat() if r["created_at"] else None,
        })
    return {"data": items, "total": len(items)}


class ContractExtendIn(BaseModel):
    contract_type: str
    duration_months: int | None = None
    start_date: str | None = None
    end_date: str | None = None
    notes: str | None = None


@router.post("/employees/{employee_id}/contracts/extend", status_code=201)
async def extend_employee_contract(
    employee_id: int,
    body: ContractExtendIn,
    request: Request,
    payload: dict = Depends(require_hr_user),
    db: AsyncSession = Depends(get_db),
):
    emp = (await db.execute(select(Employee).where(Employee.id == employee_id))).scalar_one_or_none()
    if not emp:
        raise HTTPException(404, "Employee not found")

    new_start = _date_or_none(body.start_date) or emp.contract_end_date or date.today()
    months = body.duration_months
    if months is None and body.contract_type:
        m = re.search(r"(\d+)\s*Months?", body.contract_type, re.IGNORECASE)
        months = int(m.group(1)) if m else None
    new_end = _date_or_none(body.end_date)
    if not new_end and months and new_start:
        y, m = new_start.year, new_start.month + months
        while m > 12:
            y += 1
            m -= 12
        # last day of month if day overflows
        try:
            new_end = date(y, m, new_start.day) - timedelta(days=1)
        except ValueError:
            new_end = date(y, m, 1) - timedelta(days=1) + timedelta(days=28)

    # Close current active contract row(s)
    await db.execute(text(
        "UPDATE employee_contracts SET status='ENDED' WHERE employee_code=:c AND status='ACTIVE'"
    ), {"c": emp.employee_code})

    # Insert new contract history
    await db.execute(text(
        "INSERT INTO employee_contracts (employee_code, employment_type, contract_type, duration_months, start_date, end_date, status, notes) "
        "VALUES (:c, :et, :ct, :m, :s, :e, 'ACTIVE', :n)"
    ), {
        "c": emp.employee_code, "et": emp.employment_type, "ct": body.contract_type,
        "m": months, "s": new_start, "e": new_end, "n": body.notes or "Extended/Renewed via HR UI",
    })

    # Update employees row
    emp.contract_type = body.contract_type
    emp.contract_start_date = new_start
    emp.contract_end_date = new_end
    emp.contract_duration_months = months
    if emp.status == "TERMINATED":
        emp.status = "ACTIVE"
        emp.termination_date = None

    await write_audit_log(
        db, action="employee_contract_extended", entity_type="employee",
        entity_id=emp.id, employee_id=emp.id, payload=payload,
        new_value={"contract_type": body.contract_type, "start_date": str(new_start), "end_date": str(new_end) if new_end else None, "duration_months": months},
        changed_fields=["contract_type", "contract_start_date", "contract_end_date", "contract_duration_months"],
        request=request, source="Employee API",
    )
    await db.commit()
    return {"status": "ok", "employee_id": emp.id, "contract_type": body.contract_type, "start_date": str(new_start), "end_date": str(new_end) if new_end else None}


@router.get("/employees/{employee_id}/project-readiness")
async def get_employee_project_readiness(employee_id: int, payload: dict = Depends(require_hr_read_user), db: AsyncSession = Depends(get_db)):
    row = (await db.execute(select(Employee).where(Employee.id == employee_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Employee not found")
    return await _readiness_dict(db, row) | {"clock": await _clock_eligibility_dict(db, row), "kpi": await _kpi_eligibility_dict(db, row)}


class EmployeePatchIn(BaseModel):
    phone: str | None = None
    work_phone: str | None = None
    work_location: str | None = None
    notes: str | None = None
    employment_type: str | None = None
    contract_type: str | None = None
    contract_start_date: str | None = None
    contract_end_date: str | None = None
    contract_duration_months: int | None = None
    hire_date: str | None = None
    probation_end_date: str | None = None
    termination_date: str | None = None
    gender: str | None = None
    date_of_birth: str | None = None
    nationality: str | None = None
    id_card_no: str | None = None
    address: str | None = None
    emergency_contact_name: str | None = None
    emergency_contact_relation: str | None = None
    emergency_contact_phone: str | None = None
    bank_name: str | None = None
    bank_account_no: str | None = None
    bank_account_name: str | None = None
    photo_url: str | None = None
    manager_name: str | None = None
    manager_code: str | None = None
    cost_center: str | None = None
    job_level: str | None = None


def _employee_required_missing(row: Employee) -> list[str]:
    checks = [
        ("Company email", row.email),
        ("Phone", row.phone or row.work_phone),
        ("Department", row.department),
        ("Position", row.position or row.job_title or row.project_role),
        ("Manager", row.manager_name or row.manager_code),
        ("Cost center", row.cost_center),
        ("Work location", row.work_location or row.project_code),
        ("Job level", row.job_level),
    ]
    return [label for label, value in checks if not value or str(value).strip() in {"", "-", "—"}]


def _assignment_required(row: Employee) -> bool:
    return (row.department or "").lower() in {"project", "project management"} or bool(row.project_team in {"RF", "TE"})


async def _active_assignments(db: AsyncSession, employee_code: str) -> list[ProjectAssignment]:
    return (
        await db.execute(
            select(ProjectAssignment).where(
                ProjectAssignment.employee_code == employee_code,
                ProjectAssignment.is_active.is_(True),
            )
        )
    ).scalars().all()


def _account_locked(user: AuthUser | None) -> bool:
    return bool(user and user.locked_until and user.locked_until > datetime.now(timezone.utc))


async def _latest_welcome(db: AsyncSession, email: str | None):
    from app.models.email import EmailOutbox
    if not email:
        return None
    return (
        await db.execute(
            select(EmailOutbox)
            .where(EmailOutbox.recipient == email, EmailOutbox.subject.ilike("%welcome%"))
            .order_by(EmailOutbox.created_at.desc())
            .limit(1)
        )
    ).scalar_one_or_none()


async def _system_access_dict(db: AsyncSession, row: Employee) -> dict:
    user = (await db.execute(select(AuthUser).where(AuthUser.employee_code == row.employee_code))).scalar_one_or_none()
    latest_welcome = await _latest_welcome(db, row.email)
    account_locked = _account_locked(user)
    status = "NOT_CREATED"
    if user:
        if account_locked:
            status = "LOCKED"
        elif user.is_active:
            status = "ACTIVE"
        else:
            status = "DISABLED"
    return {
        "employee_id": row.id,
        "employee_code": row.employee_code,
        "login_created": bool(user),
        "login_active": bool(user and user.is_active and not account_locked),
        "username": user.email or user.employee_code if user else None,
        "company_email": row.email,
        "role": user.role if user else None,
        "role_name": ROLE_LABELS.get(user.role, user.role.replace("_", " ").title()) if user else None,
        "permission_scope": sorted(ROLE_SCOPES.get(user.role, set())) if user else [],
        "last_login_at": user.last_login_at.isoformat() if user and user.last_login_at else None,
        "welcome_email_sent": bool(latest_welcome and latest_welcome.status == "SENT"),
        "welcome_email_sent_at": latest_welcome.sent_at.isoformat() if latest_welcome and latest_welcome.sent_at else None,
        "welcome_email_status": latest_welcome.status if latest_welcome else "PENDING",
        "password_reset_required": bool(user and user.must_change_password),
        "account_locked": account_locked,
        "account_status": status,
        "read_only": row.status in {"ARCHIVED"},
    }


async def _readiness_dict(db: AsyncSession, row: Employee) -> dict:
    user = (await db.execute(select(AuthUser).where(AuthUser.employee_code == row.employee_code))).scalar_one_or_none()
    assignments = await _active_assignments(db, row.employee_code)
    missing = _employee_required_missing(row)
    if not user:
        missing.append("Login account")
    elif not user.is_active:
        missing.append("Active login")
    if row.status not in {"ACTIVE", "PROBATION"}:
        missing.append("Active or probation employee status")
    assignment_needed = _assignment_required(row)
    assigned = bool(assignments or row.project_code)
    if assignment_needed and not assigned:
        missing.append("Project assignment")
    total = 6 + (1 if assignment_needed else 0)
    score = max(0, round(((total - len(set(missing))) / total) * 100))
    if row.status in {"TERMINATED", "RESIGNED", "ARCHIVED"}:
        status = "Removed from Project"
    elif assigned and user and user.is_active and not missing:
        status = "Active in Project"
    elif assigned:
        status = "Assigned to Project"
    elif not missing:
        status = "Ready for Project Assignment"
    else:
        status = "Not Ready"
    return {
        "employee_id": row.id,
        "employee_code": row.employee_code,
        "readiness_status": status,
        "readiness_score": score,
        "ready": not missing,
        "missing_requirements": sorted(set(missing)),
        "active_assignment_count": len(assignments),
        "project_assignment_required": assignment_needed,
    }


async def _clock_eligibility_dict(db: AsyncSession, row: Employee) -> dict:
    user = (await db.execute(select(AuthUser).where(AuthUser.employee_code == row.employee_code))).scalar_one_or_none()
    assignments = await _active_assignments(db, row.employee_code)
    reasons = []
    if row.status not in {"ACTIVE", "PROBATION"}:
        reasons.append("Employee is not Active or Probation")
    if not user or not user.is_active:
        reasons.append("Login is not active")
    if not (row.work_location or row.project_code or (user and user.work_location_name)):
        reasons.append("Work location is missing")
    required_missing = _employee_required_missing(row)
    if required_missing:
        reasons.append("Required HR data is incomplete")
    if _assignment_required(row) and not (assignments or row.project_code):
        reasons.append("Project assignment is required")
    if row.status in {"TERMINATED", "RESIGNED", "ARCHIVED"} or row.termination_date:
        reasons.append("Employee is terminated or archived")
    return {"employee_id": row.id, "employee_code": row.employee_code, "eligible": not reasons, "status": "Clock Eligible" if not reasons else "Clock Blocked", "blocking_reasons": reasons, "missing_requirements": required_missing}


async def _kpi_eligibility_dict(db: AsyncSession, row: Employee) -> dict:
    assignments = await _active_assignments(db, row.employee_code)
    reasons = []
    if row.status not in {"ACTIVE", "PROBATION"}:
        reasons.append("Employee is not Active or Probation")
    for label, value in [("Department", row.department), ("Position", row.position or row.job_title or row.project_role), ("Job level", row.job_level), ("Manager", row.manager_name or row.manager_code)]:
        if not value:
            reasons.append(f"{label} is missing")
    if _assignment_required(row) and not (assignments or row.project_code):
        reasons.append("Project assignment is required for project-based KPI")
    return {"employee_id": row.id, "employee_code": row.employee_code, "eligible": not reasons, "status": "KPI Eligible" if not reasons else "KPI Blocked", "blocking_reasons": reasons}


@router.patch("/employees/{employee_id}")
async def patch_employee(
    employee_id: int,
    body: EmployeePatchIn,
    request: Request,
    payload: dict = Depends(require_hr_user),
    db: AsyncSession = Depends(get_db),
):
    row = (await db.execute(select(Employee).where(Employee.id == employee_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Employee not found")
    patch_values = body.model_dump(exclude_none=True)
    old_values = {}
    new_values = {}
    changed_fields = []
    for field, value in patch_values.items():
        old_values[field] = getattr(row, field)
        if field in ("hire_date", "probation_end_date", "termination_date", "date_of_birth", "contract_start_date", "contract_end_date"):
            setattr(row, field, _date_or_none(value))
        else:
            setattr(row, field, value)
        if old_values[field] != getattr(row, field):
            changed_fields.append(field)
            new_values[field] = getattr(row, field)
    if changed_fields:
        action = "employee_profile_updated"
        if any(field in changed_fields for field in ("employment_type", "contract_type", "hire_date", "probation_end_date", "termination_date", "job_level")):
            action = "employee_contract_updated"
        if "termination_date" in changed_fields and row.termination_date:
            action = "employee_terminated"
        await write_audit_log(
            db,
            action=action,
            entity_type="employee",
            entity_id=row.id,
            employee_id=row.id,
            payload=payload,
            old_value=old_values,
            new_value=new_values,
            changed_fields=changed_fields,
            request=request,
            source="Employee API",
        )
        if "status" in changed_fields:
            await write_audit_log(
                db,
                action="employee_status_changed",
                entity_type="employee",
                entity_id=row.id,
                employee_id=row.id,
                payload=payload,
                old_value={"status": old_values.get("status")},
                new_value={"status": row.status},
                changed_fields=["status"],
                request=request,
                source="Employee API",
            )
    await db.commit()
    await db.refresh(row)
    return employee_to_dict(row)


class RelocateIn(BaseModel):
    employee_code: str
    to_project_code: str | None = None
    effective_date: str
    reason: str | None = None
    notes: str | None = None


@router.post("/relocate", status_code=201)
async def relocate_employee(
    body: RelocateIn,
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    emp = (await db.execute(select(Employee).where(Employee.employee_code == body.employee_code))).scalar_one_or_none()
    if not emp:
        raise HTTPException(404, "Employee not found")

    eff_date = _date_or_none(body.effective_date)
    if not eff_date:
        from datetime import date as _date
        eff_date = _date.today()

    new_project = None
    if body.to_project_code:
        new_project = (await db.execute(select(ProjectCatalog).where(ProjectCatalog.project_code == body.to_project_code))).scalar_one_or_none()

    relocation = EmployeeRelocation(
        employee_code=emp.employee_code,
        full_name=emp.full_name,
        from_project_code=emp.project_code,
        from_project_name=emp.project_name,
        to_project_code=body.to_project_code,
        to_project_name=new_project.project_name if new_project else None,
        effective_date=eff_date,
        reason=body.reason,
        notes=body.notes,
    )
    db.add(relocation)

    if emp.project_code:
        curr_assigns = (await db.execute(
            select(ProjectAssignment).where(
                ProjectAssignment.employee_code == emp.employee_code,
                ProjectAssignment.project_code == emp.project_code,
                ProjectAssignment.is_active.is_(True),
            )
        )).scalars().all()
        for a in curr_assigns:
            a.is_active = False
            a.end_date = eff_date

    if body.to_project_code:
        new_assign = ProjectAssignment(
            project_code=body.to_project_code,
            employee_code=emp.employee_code,
            role_in_project=emp.project_role or "OTHER",
            clock_type=_normalize_assignment_clock_type(emp.project_role or "OTHER", None),
            job_level=emp.job_level,
            start_date=eff_date,
            is_active=True,
        )
        db.add(new_assign)

    emp.project_code = body.to_project_code
    emp.project_name = new_project.project_name if new_project else None

    await db.commit()
    await db.refresh(relocation)
    return {
        "id": relocation.id,
        "employee_code": relocation.employee_code,
        "from_project_code": relocation.from_project_code,
        "to_project_code": relocation.to_project_code,
        "effective_date": relocation.effective_date.isoformat(),
    }


@router.get("/employees/{employee_code}/relocations")
async def get_employee_relocations(
    employee_code: str,
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    rows = (await db.execute(
        select(EmployeeRelocation)
        .where(EmployeeRelocation.employee_code == employee_code)
        .order_by(EmployeeRelocation.effective_date.desc())
    )).scalars().all()
    return {
        "data": [
            {
                "id": r.id,
                "from_project_code": r.from_project_code,
                "from_project_name": r.from_project_name,
                "to_project_code": r.to_project_code,
                "to_project_name": r.to_project_name,
                "effective_date": r.effective_date.isoformat() if r.effective_date else None,
                "reason": r.reason,
                "notes": r.notes,
                "created_at": r.created_at.isoformat() if r.created_at else None,
            }
            for r in rows
        ]
    }


@router.get("/project-employees")
async def list_project_employees(
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    rows = (
        await db.execute(
            select(Employee)
            .where(Employee.department.in_(["Project", "Project Management"]))
            .order_by(Employee.project_team, Employee.full_name)
        )
    ).scalars().all()
    return {"data": [employee_to_dict(row) for row in rows], "total": len(rows)}


@router.post("/employees", status_code=201)
async def create_employee(
    body: EmployeeIn,
    request: Request,
    payload: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if payload.get("role") not in {"SUPER_ADMIN", "HR_ADMIN"}:
        raise HTTPException(status_code=403, detail="Not allowed")
    next_no = (await db.execute(select(func.count()).select_from(Employee))).scalar_one() + 1
    employee_code = (body.employee_code or f"ACE-{next_no:04d}").strip().upper()
    exists = (await db.execute(select(Employee).where(Employee.employee_code == employee_code))).scalar_one_or_none()
    if exists:
        raise HTTPException(400, "Employee code already exists.")
    if body.email:
        email_exists = (await db.execute(select(Employee).where(Employee.email == str(body.email)))).scalar_one_or_none()
        if email_exists:
            raise HTTPException(400, "Employee email already exists.")
    first_name = body.first_name
    last_name = body.last_name
    if not first_name and body.full_name:
        parts = body.full_name.split()
        first_name = parts[0] if parts else None
        last_name = " ".join(parts[1:]) if len(parts) > 1 else None
    row = Employee(
        employee_code=employee_code,
        email=body.email,
        full_name=body.full_name,
        first_name=first_name,
        last_name=last_name,
        preferred_name=body.preferred_name,
        personal_email=body.personal_email,
        phone=body.phone,
        work_phone=body.work_phone,
        department=body.department,
        section_name=body.section_name,
        job_title=body.job_title or body.position,
        job_level=body.job_level,
        manager_name=body.manager_name,
        cost_center=body.cost_center,
        work_location=body.work_location,
        project_team=body.project_team,
        project_role=body.project_role,
        project_code=body.project_code,
        project_name=body.project_name,
        position=body.position or body.project_team,
        status=body.status,
        employment_type=body.employment_type or "Permanent",
        contract_type=body.contract_type or "Permanent",
        contract_start_date=_date_or_none(body.contract_start_date),
        contract_end_date=_date_or_none(body.contract_end_date),
        contract_duration_months=body.contract_duration_months,
        hire_date=_date_or_none(body.hire_date),
        probation_end_date=_date_or_none(body.probation_end_date),
        termination_date=_date_or_none(body.termination_date),
        date_of_birth=_date_or_none(body.date_of_birth),
        gender=body.gender,
        nationality=body.nationality or "Thai",
        id_card_no=body.id_card_no,
        address=body.address,
        emergency_contact_name=body.emergency_contact_name,
        emergency_contact_relation=body.emergency_contact_relation,
        emergency_contact_phone=body.emergency_contact_phone,
        bank_name=body.bank_name,
        bank_account_no=body.bank_account_no,
        bank_account_name=body.bank_account_name,
        notes=body.notes,
        source="HR",
    )
    db.add(row)
    if body.department == "Project" and body.project_code:
        role = _role_from_employee(row)
        db.add(ProjectAssignment(
            project_code=body.project_code,
            employee_code=employee_code,
            role_in_project=role,
            clock_type=_normalize_assignment_clock_type(role, None),
            start_date=_date_or_none(body.hire_date),
            allocation_pct=100,
            is_active=True,
        ))
    notification = None
    login_created = False
    if body.create_login:
        if not body.email:
            raise HTTPException(400, "Email is required when create_login is true.")
        login_exists = (await db.execute(select(AuthUser).where(AuthUser.employee_code == employee_code))).scalar_one_or_none()
        if login_exists:
            raise HTTPException(400, "Employee code already has login account.")
        password = body.password or _generate_password()
        _validate_initial_password(password, email=str(body.email), employee_code=employee_code)
        clock_type, gps_required, photo_required = _clock_config_for_employee(row)
        db.add(AuthUser(
            employee_code=employee_code,
            password_hash=hash_password(password),
            first_name=first_name or body.full_name,
            last_name=last_name or "",
            email=str(body.email),
            department=body.department,
            team=body.project_team,
            position_code=_role_from_employee(row),
            position_name=body.position or body.job_title or "Employee",
            role=_auth_role_for_employee(row, body.system_role),
            clock_type=clock_type,
            gps_required=gps_required,
            photo_required=photo_required,
            work_location_name=body.work_location or "ACE Head Office",
            allowed_radius_m=300,
            must_change_password=True,
            created_by=payload.get("id"),
        ))
        login_created = True
        if body.send_welcome_email:
            subject, body_text, body_html = welcome_email(employee_code, body.full_name, password)
            result = await queue_and_send_email(db, str(body.email), subject, body_text, body_html)
            notification = {
                "channel": "email",
                "recipient": str(body.email),
                "status": result.status,
                "outbox_id": result.outbox_id,
                "error_code": result.error_code,
                "error_message": result.error_message,
            }
        else:
            notification = {
                "channel": "email",
                "recipient": str(body.email),
                "status": "SKIPPED",
                "outbox_id": None,
                "error_code": None,
                "error_message": None,
            }
    await db.flush()
    await write_audit_log(
        db,
        action="employee_created",
        entity_type="employee",
        entity_id=row.id,
        employee_id=row.id,
        payload=payload,
        new_value=employee_to_dict(row),
        changed_fields=list(employee_to_dict(row).keys()),
        request=request,
        source="Employee API",
    )
    if login_created:
        await write_audit_log(
            db,
            action="login_created",
            entity_type="auth_user",
            entity_id=employee_code,
            employee_id=row.id,
            payload=payload,
            new_value={"employee_code": employee_code, "email": str(body.email), "role": _auth_role_for_employee(row, body.system_role)},
            changed_fields=["employee_code", "email", "role"],
            request=request,
            source="Employee API",
        )
    if notification:
        await write_audit_log(
            db,
            action="welcome_email_sent" if notification.get("status") == "SENT" else "welcome_email_failed",
            entity_type="email",
            entity_id=notification.get("outbox_id"),
            employee_id=row.id,
            payload=payload,
            new_value=notification,
            changed_fields=["status", "recipient"],
            request=request,
            source="Employee API",
        )
    await db.commit()
    await db.refresh(row)
    return employee_to_dict(row) | {"login_created": body.create_login, "notification": notification}


class ProjectCatalogIn(BaseModel):
    project_code: str
    project_name: str
    team: str = "RF"
    notes: str | None = None


@router.post("/projects", status_code=201)
async def create_project(
    body: ProjectCatalogIn,
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    exists = (await db.execute(select(ProjectCatalog).where(ProjectCatalog.project_code == body.project_code))).scalar_one_or_none()
    if exists:
        raise HTTPException(400, f"Project code '{body.project_code}' already exists.")
    row = ProjectCatalog(
        project_code=body.project_code.strip().upper(),
        project_name=body.project_name.strip(),
        team=body.team,
        headcount=0,
        notes=body.notes,
    )
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return project_to_dict(row)


@router.patch("/projects/{project_code}")
async def patch_project(
    project_code: str,
    body: ProjectCatalogIn,
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    row = (await db.execute(select(ProjectCatalog).where(ProjectCatalog.project_code == project_code))).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Project not found")
    row.project_name = body.project_name.strip()
    row.team = body.team
    if body.notes is not None:
        row.notes = body.notes
    await db.commit()
    await db.refresh(row)
    return project_to_dict(row)


@router.get("/projects")
async def list_projects(
    team: str = Query(""),
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(ProjectCatalog)
    if team:
        stmt = stmt.where(ProjectCatalog.team == team)
    rows = (await db.execute(stmt.order_by(ProjectCatalog.project_code))).scalars().all()
    # Count active employees per project from employees table (source of truth after Excel sync)
    counts = dict((await db.execute(
        select(Employee.project_code, func.count())
        .where(Employee.status == "ACTIVE", Employee.project_code.isnot(None))
        .group_by(Employee.project_code)
    )).all())
    site_counts = dict((await db.execute(
        select(ProjectSite.project_code, func.count())
        .where(ProjectSite.is_active == True)
        .group_by(ProjectSite.project_code)
    )).all())
    return {"data": [project_list_to_dict(row, counts.get(row.project_code, 0), site_counts.get(row.project_code, 0)) for row in rows], "total": len(rows)}


@router.get("/projects/{project_code}/assignments")
async def list_project_assignments(
    project_code: str,
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    rows = (await db.execute(
        select(ProjectAssignment, Employee)
        .join(Employee, Employee.employee_code == ProjectAssignment.employee_code, isouter=True)
        .where(ProjectAssignment.project_code == project_code)
        .order_by(ProjectAssignment.is_active.desc(), Employee.full_name)
    )).all()
    return {"data": [assignment_to_dict(a, e) for a, e in rows], "total": len(rows)}


@router.post("/projects/{project_code}/assignments", status_code=201)
async def create_project_assignment(
    project_code: str,
    body: AssignmentIn,
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    employee = (await db.execute(select(Employee).where(Employee.employee_code == body.employee_code))).scalar_one_or_none()
    if not employee:
        raise HTTPException(404, "Employee not found")
    row = ProjectAssignment(
        project_code=project_code,
        employee_code=body.employee_code,
        role_in_project=body.role_in_project,
        clock_type=_normalize_assignment_clock_type(body.role_in_project, body.clock_type),
        job_level=body.job_level,
        start_date=_date_or_none(body.start_date),
        end_date=_date_or_none(body.end_date),
        allocation_pct=body.allocation_pct,
        is_active=body.is_active,
        notes=body.notes,
    )
    employee.department = "Project Management"
    employee.project_code = project_code
    employee.project_name = employee.project_name or project_code
    employee.project_team = _project_team_from_role(body.role_in_project)
    employee.project_role = body.role_in_project
    employee.job_level = body.job_level
    employee.position = _project_position_from_role(body.role_in_project)
    employee.job_title = employee.position
    db.add(row)
    await db.flush()
    await _sync_auth_user_from_employee(db, employee)
    await db.commit()
    await db.refresh(row)
    return assignment_to_dict(row, employee)


@router.patch("/projects/{project_code}/assignments/{assignment_id}")
async def update_project_assignment(
    project_code: str,
    assignment_id: int,
    body: AssignmentUpdateIn,
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    row = (await db.execute(
        select(ProjectAssignment).where(
            ProjectAssignment.id == assignment_id,
            ProjectAssignment.project_code == project_code,
        )
    )).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Assignment not found")
    if body.role_in_project is not None:
        row.role_in_project = body.role_in_project
    if body.role_in_project is not None or body.clock_type is not None:
        row.clock_type = _normalize_assignment_clock_type(row.role_in_project, body.clock_type or row.clock_type)
    if body.job_level is not None:
        row.job_level = body.job_level
    if body.allocation_pct is not None:
        row.allocation_pct = body.allocation_pct
    if body.start_date is not None:
        row.start_date = _date_or_none(body.start_date)
    if body.end_date is not None:
        row.end_date = _date_or_none(body.end_date)
    if body.is_active is not None:
        row.is_active = body.is_active
    if body.notes is not None:
        row.notes = body.notes
    await db.commit()
    await db.refresh(row)
    employee = (await db.execute(
        select(Employee).where(Employee.employee_code == row.employee_code)
    )).scalar_one_or_none()
    if employee:
        if body.role_in_project is not None:
            employee.project_role = body.role_in_project
            employee.project_team = _project_team_from_role(body.role_in_project)
            employee.position = _project_position_from_role(body.role_in_project)
            employee.job_title = employee.position
        if body.job_level is not None:
            employee.job_level = body.job_level
        employee.department = "Project Management"
        await _sync_auth_user_from_employee(db, employee)
        await db.commit()
    return assignment_to_dict(row, employee)


@router.post("/employees/{employee_code}/transfer", status_code=201)
async def transfer_employee(
    employee_code: str,
    body: TransferIn,
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    from datetime import date as date_cls
    today = date_cls.today().isoformat()

    old_row = (await db.execute(
        select(ProjectAssignment).where(
            ProjectAssignment.id == body.from_assignment_id,
            ProjectAssignment.employee_code == employee_code,
        )
    )).scalar_one_or_none()
    if not old_row:
        raise HTTPException(404, "Assignment not found")

    employee = (await db.execute(
        select(Employee).where(Employee.employee_code == employee_code)
    )).scalar_one_or_none()
    if not employee:
        raise HTTPException(404, "Employee not found")

    old_row.end_date = _date_or_none(body.end_date or today)
    old_row.is_active = False

    new_row = ProjectAssignment(
        project_code=body.to_project_code,
        employee_code=employee_code,
        role_in_project=body.role_in_project,
        clock_type=_normalize_assignment_clock_type(body.role_in_project, body.clock_type),
        job_level=body.job_level,
        start_date=_date_or_none(body.start_date or today),
        allocation_pct=body.allocation_pct,
        is_active=True,
    )
    employee.project_code = body.to_project_code
    employee.project_role = body.role_in_project
    employee.project_team = _project_team_from_role(body.role_in_project)
    employee.job_level = body.job_level
    employee.position = _project_position_from_role(body.role_in_project)
    employee.job_title = employee.position
    employee.department = "Project Management"

    db.add(new_row)
    await db.flush()
    await _sync_auth_user_from_employee(db, employee)
    await db.commit()
    await db.refresh(old_row)
    await db.refresh(new_row)
    return {
        "ended": assignment_to_dict(old_row, employee),
        "new": assignment_to_dict(new_row, employee),
    }


@router.get("/employees/{employee_code}/assignment-history")
async def get_employee_assignment_history(
    employee_code: str,
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    rows = (await db.execute(
        select(ProjectAssignment, ProjectCatalog)
        .join(ProjectCatalog, ProjectCatalog.project_code == ProjectAssignment.project_code, isouter=True)
        .where(ProjectAssignment.employee_code == employee_code)
        .order_by(ProjectAssignment.is_active.desc(), ProjectAssignment.start_date.desc().nullslast())
    )).all()
    return {
        "data": [
            {
                **assignment_to_dict(a),
                "project_name": p.project_name if p else a.project_code,
            }
            for a, p in rows
        ]
    }


@router.get("/sites")
async def list_all_sites(
    project_code: str = Query(""),
    referenced_only: bool = Query(False),
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    """List project_sites. After ISDP import the table can hold 100k+ rows, so
    callers that only need to resolve PO→cluster mappings should pass
    `project_code` (sites referenced by that project's POs) or
    `referenced_only=true` (sites referenced by ANY PO). This trims a 136k-row
    payload down to a few hundred and keeps the Pre-Site monitor fast."""
    stmt = select(ProjectSite)
    if project_code or referenced_only:
        # Sites whose code is referenced by a PO via cluster_site / site_code / du_id
        po_q = select(
            ProjectPO.cluster_site, ProjectPO.site_code, ProjectPO.du_id
        )
        if project_code:
            po_q = po_q.where(ProjectPO.ace_project_code == project_code)
        po_rows = (await db.execute(po_q)).all()
        codes: set[str] = set()
        for cs, sc, du in po_rows:
            for v in (cs, sc, du):
                if v:
                    codes.add(v)
                    # Also include the short/root code (e.g. "RYG7235" from
                    # "RYG7235_Flash_RAN_EAS R3") — MasterDB keys lat/lng by the
                    # short Site ID, so the frontend's short-key fallback needs it.
                    root = str(v).split("_", 1)[0].strip()
                    if root and root != v:
                        codes.add(root)
        if not codes:
            return {"data": [], "total": 0}
        stmt = stmt.where(ProjectSite.site_code.in_(codes))
    rows = (await db.execute(stmt.order_by(ProjectSite.site_code))).scalars().all()
    return {"data": [site_to_dict(row) for row in rows], "total": len(rows)}


@router.post("/sites/import-masterdb")
async def import_masterdb_sites(
    request: Request,
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    """Import Site ID sheet from MasterDB xlsx → upsert project_sites (site_code, lat, lng, province)"""
    import io
    import openpyxl

    form   = await request.form()
    upload = form.get("file")
    if not upload or not hasattr(upload, "read"):
        raise HTTPException(400, "ต้องแนบไฟล์ Excel (.xlsx)")

    content = await upload.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "ไม่สามารถอ่านไฟล์ได้ — ต้องเป็น .xlsx")

    if "Site ID" not in wb.sheetnames:
        raise HTTPException(400, "ไม่พบ Sheet 'Site ID' ในไฟล์")

    ws = wb["Site ID"]
    # Row 1 = headers, Row 2 = count row, Row 3+ = data
    headers = [str(v).strip() if v is not None else "" for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    try:
        idx_site = headers.index("Site ID")
        idx_lat  = headers.index("Latitude")
        idx_lng  = headers.index("Longitude")
    except ValueError as e:
        raise HTTPException(400, f"ไม่พบ column: {e}")

    idx_province = headers.index("Province") if "Province" in headers else None

    # Load existing site_codes to decide insert vs update
    existing: dict[str, int] = {
        r[0]: r[1]
        for r in (await db.execute(select(ProjectSite.site_code, ProjectSite.id))).all()
    }

    inserted = updated = skipped = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        site_id = row[idx_site]
        if not site_id:
            continue
        site_id = str(site_id).strip()
        try:
            lat = float(row[idx_lat]) if row[idx_lat] is not None else None
            lng = float(row[idx_lng]) if row[idx_lng] is not None else None
        except (ValueError, TypeError):
            skipped += 1
            continue

        province = str(row[idx_province]).strip() if idx_province is not None and row[idx_province] else None

        if site_id in existing:
            await db.execute(
                update(ProjectSite)
                .where(ProjectSite.id == existing[site_id])
                .values(lat=lat, lng=lng, province=province, updated_at=datetime.now(timezone.utc))
            )
            updated += 1
        else:
            db.add(ProjectSite(
                site_code=site_id,
                lat=lat,
                lng=lng,
                province=province,
                is_active=True,
            ))
            existing[site_id] = 0
            inserted += 1

    # Audit log → Master Data tab will surface "Last upload" timestamp
    from app.routers.data_imports import record_data_import
    await record_data_import(
        db,
        file_type="MASTERDB",
        file_name=getattr(upload, "filename", None),
        row_count=inserted + updated + skipped,
        inserted=inserted,
        updated=updated,
        skipped=skipped,
        uploaded_by_code=(payload.get("employee_code") or payload.get("sub")),
        uploaded_by_name=payload.get("name"),
    )

    await db.commit()
    return {"inserted": inserted, "updated": updated, "skipped": skipped, "total": inserted + updated}


@router.post("/sites/import-rollout-plan")
async def import_rollout_plan(
    request: Request,
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    """Import ISDP rollout plan (.xlsm) → update project_sites from Sheet 'Site Rollout Plan'
    Col B=site_code, Col C=DU ID, Col I=rf_cluster_name, Col K=cluster_ready, Col O=full_on_air

    Col B and Col C are BOTH stored as keys into project_sites (sharing the same
    rf_cluster_name from Col I). This is so PAC POs whose `cluster_site` carries a
    DU-ID-style value (e.g. "RYG7235_Flash_RAN_EAS R3") can still be looked up to
    the canonical RF Cluster Name (e.g. "EAS-FLASH-0012")."""
    import io
    import openpyxl
    from datetime import date as date_type

    form   = await request.form()
    upload = form.get("file")
    if not upload or not hasattr(upload, "read"):
        raise HTTPException(400, "ต้องแนบไฟล์ (.xlsx / .xlsm)")

    content = await upload.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "ไม่สามารถอ่านไฟล์ได้")

    if "Site Rollout Plan" not in wb.sheetnames:
        raise HTTPException(400, "ไม่พบ Sheet 'Site Rollout Plan'")

    ws = wb["Site Rollout Plan"]
    # Row 1-3 = headers, data from row 4
    # Col B(1)=Site Code, Col C(2)=DU ID, Col I(8)=RF Cluster Name, Col K(10)=Cluster Ready, Col O(14)=Full On-Air

    def _to_date(val) -> date_type | None:
        if val is None:
            return None
        if hasattr(val, "date"):
            return val.date()
        if isinstance(val, date_type):
            return val
        try:
            return datetime.strptime(str(val)[:10], "%Y-%m-%d").date()
        except Exception:
            return None

    # Collect per site: max full_on_air, max cluster_ready, last rf_cluster_name
    # Col B may contain multiple space-separated site codes for cluster rows → apply to each
    # Col C carries the DU ID (e.g. "RYG7235_Flash_RAN_EAS R3"); we ALSO store it
    # as a site_code so PAC POs whose cluster_site holds a DU-ID-style value can be
    # resolved to their RF Cluster Name (Col I).
    site_data: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=4, max_col=15, values_only=True):
        raw_site = row[1]
        raw_du   = row[2]
        if not raw_site and not raw_du:
            continue
        rf_name  = str(row[8]).strip() if row[8] else None
        cl_ready = _to_date(row[10])
        owner    = str(row[11]).strip() if row[11] else None   # Col L = Owner
        on_air   = _to_date(row[14])

        # Col E/F = Common Site Latitude/Longitude — the ONLY coord source for
        # full-name DU-ID sites (e.g. "RYG7343_Relocate_East R3") which don't
        # exist in MasterDB. Used to plot SSV/PAC sites on the map.
        def _to_float(v):
            try:
                f = float(v)
                return f if -90 <= f <= 180 and f != 0 else None
            except (TypeError, ValueError):
                return None
        lat = _to_float(row[4])
        lng = _to_float(row[5])

        tokens: list[str] = []
        if raw_site:
            tokens.extend(str(raw_site).split())
        if raw_du:
            tokens.append(str(raw_du).strip())
        # deduplicate while preserving order; drop empties / over-long values
        seen: set[str] = set()
        unique_codes = [s for s in tokens if s and len(s) <= 50 and not (s in seen or seen.add(s))]

        for site_code in unique_codes:
            entry = site_data.setdefault(site_code, {"full_on_air": None, "cluster_ready": None, "rf_cluster_name": None, "rf_cluster_owner": None, "lat": None, "lng": None})
            if on_air and (entry["full_on_air"] is None or on_air > entry["full_on_air"]):
                entry["full_on_air"] = on_air
            if cl_ready and (entry["cluster_ready"] is None or cl_ready > entry["cluster_ready"]):
                entry["cluster_ready"] = cl_ready
            if rf_name and not entry["rf_cluster_name"]:
                entry["rf_cluster_name"] = rf_name
            if owner and not entry["rf_cluster_owner"]:
                entry["rf_cluster_owner"] = owner
            if lat is not None and entry["lat"] is None:
                entry["lat"] = lat
            if lng is not None and entry["lng"] is None:
                entry["lng"] = lng

    if not site_data:
        return {"updated": 0, "inserted": 0, "total_sites_in_file": 0}

    # Pre-count which site_codes already exist (for accurate insert/update reporting)
    existing_codes: set[str] = {
        r[0]
        for r in (await db.execute(select(ProjectSite.site_code))).all()
    }

    # Bulk UPSERT — single round-trip vs. 25k+ sequential UPDATEs.
    # COALESCE preserves the existing value when the new row's column is NULL,
    # matching the previous loop semantics (which only set fields where v is not None).
    now = datetime.now(timezone.utc)
    rows = [
        {
            "site_code":        site_code,
            "is_active":        True,
            "full_on_air":      info.get("full_on_air"),
            "cluster_ready":    info.get("cluster_ready"),
            "rf_cluster_name":  info.get("rf_cluster_name"),
            "rf_cluster_owner": info.get("rf_cluster_owner"),
            "lat":              info.get("lat"),
            "lng":              info.get("lng"),
            "updated_at":       now,
        }
        for site_code, info in site_data.items()
    ]

    # Chunk the insert to keep each statement under asyncpg's bind-parameter
    # limit (32,767 = 2^15-1). pg_insert binds ALL of ProjectSite's columns per
    # row (~17), not just the dict keys, so size off the full table column count
    # with headroom. CHUNK=1500 × 17 cols = 25.5k params → safe.
    CHUNK = 1500
    for i in range(0, len(rows), CHUNK):
        batch = rows[i:i + CHUNK]
        stmt  = pg_insert(ProjectSite).values(batch)
        stmt  = stmt.on_conflict_do_update(
            index_elements=["site_code"],
            set_={
                "full_on_air":      func.coalesce(stmt.excluded.full_on_air,      ProjectSite.full_on_air),
                "cluster_ready":    func.coalesce(stmt.excluded.cluster_ready,    ProjectSite.cluster_ready),
                "rf_cluster_name":  func.coalesce(stmt.excluded.rf_cluster_name,  ProjectSite.rf_cluster_name),
                "rf_cluster_owner": func.coalesce(stmt.excluded.rf_cluster_owner, ProjectSite.rf_cluster_owner),
                # COALESCE existing-first so MasterDB coords stay authoritative;
                # ISDP only fills lat/lng for sites that don't have them yet.
                "lat":              func.coalesce(ProjectSite.lat, stmt.excluded.lat),
                "lng":              func.coalesce(ProjectSite.lng, stmt.excluded.lng),
                "updated_at":       stmt.excluded.updated_at,
            },
        )
        await db.execute(stmt)

    updated  = sum(1 for c in site_data if c     in existing_codes)
    inserted = sum(1 for c in site_data if c not in existing_codes)

    # Audit log → Master Data tab will surface "Last upload" timestamp
    from app.routers.data_imports import record_data_import
    await record_data_import(
        db,
        file_type="ISDP",
        file_name=getattr(upload, "filename", None),
        row_count=len(site_data),
        inserted=inserted,
        updated=updated,
        uploaded_by_code=(payload.get("employee_code") or payload.get("sub")),
        uploaded_by_name=payload.get("name"),
    )

    await db.commit()

    return {
        "updated": updated,
        "inserted": inserted,
        "total_sites_in_file": len(site_data),
    }


@router.get("/projects/{project_code}/sites")
async def list_project_sites(
    project_code: str,
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    rows = (await db.execute(
        select(ProjectSite)
        .where(ProjectSite.project_code == project_code)
        .order_by(ProjectSite.site_code)
    )).scalars().all()
    return {"data": [site_to_dict(row) for row in rows], "total": len(rows)}


@router.post("/projects/{project_code}/sites", status_code=201)
async def create_project_site(
    project_code: str,
    body: SiteIn,
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    row = ProjectSite(project_code=project_code, **body.model_dump())
    db.add(row)
    clock_site = (await db.execute(select(ClockSite).where(ClockSite.site_code == body.site_code))).scalar_one_or_none()
    if clock_site:
        clock_site.project_code = project_code
        clock_site.site_name = body.site_name
        clock_site.customer = body.customer
        clock_site.lat = body.lat
        clock_site.lng = body.lng
        clock_site.gps_radius_m = body.gps_radius_m
        clock_site.is_active = body.is_active
    else:
        db.add(ClockSite(
            site_code=body.site_code,
            site_name=body.site_name,
            customer=body.customer,
            project_code=project_code,
            lat=body.lat,
            lng=body.lng,
            gps_radius_m=body.gps_radius_m,
            is_active=body.is_active,
        ))
    await db.commit()
    await db.refresh(row)
    return site_to_dict(row)


@router.get("/project-pos")
async def list_project_pos(
    po_target: str = Query(""),
    workflow_status: str = Query(""),
    project_code: str = Query(""),
    ace_project_code: str = Query(""),
    work_type: str = Query(""),
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(ProjectPO)
    if po_target:
        stmt = stmt.where(ProjectPO.po_target == po_target)
    if workflow_status:
        stmt = stmt.where(ProjectPO.workflow_status == workflow_status.upper())
    if project_code:
        stmt = stmt.where(ProjectPO.project_code == project_code.upper())
    if ace_project_code:
        stmt = stmt.where(ProjectPO.ace_project_code == ace_project_code.upper())
    if work_type:
        stmt = stmt.where(ProjectPO.work_type == work_type.upper())
    rows = (await db.execute(stmt.order_by(ProjectPO.po_number, ProjectPO.po_line))).scalars().all()
    return {"data": [po_to_dict(row) for row in rows], "total": len(rows)}


@router.get("/project-pos/collection-dashboard")
async def project_po_collection_dashboard(
    ace_project_code: str = Query(""),
    work_type: str = Query(""),
    vendor: str = Query(""),
    owner_role: str = Query(""),
    billing_state: str = Query(""),   # NOT_BILLED | PARTIAL | BILLED | REJECTED
    dte_pay_state: str = Query(""),   # PAID | UNPAID
    aging_min: int = Query(0),        # only rows stuck >= N days
    month_from: str = Query(""),      # "YYYY-MM" — filter PO months >= this
    month_to: str = Query(""),        # "YYYY-MM" — filter PO months <= this
    payload: dict = Depends(require_po_tracking_user),
    db: AsyncSession = Depends(get_db),
):
    """Aggregated PO collection-tracking dashboard.

    Answers, across both money legs (collect from HW · pay the DTE):
      • what is collected / partially billed / not billed / rejected
      • where each PO is sitting and who owns it now (Finance vs Project)
      • aging — POs stuck too long in their current stage

    Returns KPI summary, pipeline funnel, breakdowns (owner/vendor/project),
    an aging watchlist, and the filtered row list (one call, ACCOUNTING-safe).
    """
    stmt = select(ProjectPO)
    if ace_project_code:
        stmt = stmt.where(ProjectPO.ace_project_code == ace_project_code.upper())
    if work_type:
        stmt = stmt.where(ProjectPO.work_type == work_type.upper())
    if vendor:
        stmt = stmt.where(ProjectPO.vendor == vendor.upper())
    if owner_role:
        stmt = stmt.where(ProjectPO.current_owner_role == owner_role.upper())
    all_rows = (await db.execute(
        stmt.order_by(ProjectPO.po_number, ProjectPO.po_line)
    )).scalars().all()

    bs_filter = billing_state.upper()
    ps_filter = dte_pay_state.upper()
    mf = month_from.strip()
    mt = month_to.strip()

    def _in_month_range(row: ProjectPO) -> bool:
        if not mf and not mt:
            return True
        months = _po_months(row)
        if not months:
            return False
        if mf and not any(m >= mf for m in months):
            return False
        if mt and not any(m <= mt for m in months):
            return False
        # require at least one month inside [mf, mt]
        return any((not mf or m >= mf) and (not mt or m <= mt) for m in months)

    def _amt(row: ProjectPO) -> float:
        try:
            return float(row.line_amount) if row.line_amount is not None else 0.0
        except (TypeError, ValueError):
            return 0.0

    # Accumulators
    summary = {k: 0.0 for k in (
        "total_value", "billed_value", "partial_value", "not_billed_value",
        "rejected_value", "dte_paid_value", "dte_unpaid_value",
    )}
    counts = {k: 0 for k in (
        "total", "billed", "partial", "not_billed", "rejected",
        "dte_paid", "dte_unpaid", "no_amount",
    )}
    by_status: dict[str, dict] = {}
    by_phase: dict[str, dict] = {}
    by_owner: dict[str, dict] = {}
    by_vendor: dict[str, dict] = {}
    by_project: dict[str, dict] = {}
    monthly: dict[str, dict] = {}      # "YYYY-MM" → {collected, dte_paid}
    rows_out: list[dict] = []
    aging_watch: list[dict] = []

    def _bump(bucket: dict, key: str, value: float):
        slot = bucket.setdefault(key, {"count": 0, "value": 0.0})
        slot["count"] += 1
        slot["value"] += value

    for row in all_rows:
        if not _in_month_range(row):
            continue
        b_state = _po_billing_state(row)
        p_state = _po_dte_pay_state(row)
        # Row-level filters that depend on derived state
        if bs_filter and b_state != bs_filter:
            continue
        if ps_filter and p_state != ps_filter:
            continue
        aging = _po_aging_days(row)
        if aging_min and aging < aging_min:
            continue

        amt = _amt(row)
        phase = _po_phase(row.workflow_status)

        counts["total"] += 1
        summary["total_value"] += amt
        if amt == 0.0:
            counts["no_amount"] += 1

        if b_state == "BILLED":
            counts["billed"] += 1; summary["billed_value"] += amt
        elif b_state == "PARTIAL":
            counts["partial"] += 1; summary["partial_value"] += amt
        elif b_state == "REJECTED":
            counts["rejected"] += 1; summary["rejected_value"] += amt
        else:
            counts["not_billed"] += 1; summary["not_billed_value"] += amt

        if p_state == "PAID":
            counts["dte_paid"] += 1; summary["dte_paid_value"] += amt
        elif p_state == "UNPAID":
            counts["dte_unpaid"] += 1; summary["dte_unpaid_value"] += amt

        # Monthly trend — attribute amount to the month each leg actually completed
        if b_state == "BILLED" and row.hw_billed_at:
            monthly.setdefault(row.hw_billed_at.strftime("%Y-%m"), {"collected": 0.0, "dte_paid": 0.0})["collected"] += amt
        if p_state == "PAID" and row.dte_paid_at:
            monthly.setdefault(row.dte_paid_at.strftime("%Y-%m"), {"collected": 0.0, "dte_paid": 0.0})["dte_paid"] += amt

        _bump(by_status, (row.workflow_status or "NEW").upper(), amt)
        _bump(by_phase, phase, amt)
        _bump(by_owner, (row.current_owner_role or "FINANCE").upper(), amt)
        _bump(by_vendor, (row.vendor or "HW").upper(), amt)
        _bump(by_project, (row.ace_project_code or "—").upper(), amt)

        rec = po_to_dict(row)
        rec["billing_state"] = b_state
        rec["dte_pay_state"] = p_state
        rec["phase"] = phase
        rec["aging_days"] = aging
        rows_out.append(rec)

        # Aging watchlist: not closed, not rejected, still owed collection
        if b_state in {"NOT_BILLED", "PARTIAL"} and (row.workflow_status or "").upper() not in {"CLOSED", "REJECTED"}:
            aging_watch.append({
                "id": row.id,
                "po_number": row.po_number,
                "po_line": row.po_line,
                "cluster_site": row.cluster_site,
                "site_code": row.site_code,
                "vendor": (row.vendor or "HW"),
                "work_type": row.work_type,
                "ace_project_code": row.ace_project_code,
                "workflow_status": row.workflow_status,
                "current_owner_role": row.current_owner_role,
                "current_owner_user": row.current_owner_user,
                "billing_state": b_state,
                "line_amount": amt,
                "aging_days": aging,
            })

    # Collection rate = billed value over collectable (exclude rejected) value
    collectable = summary["total_value"] - summary["rejected_value"]
    collection_rate = round(100.0 * summary["billed_value"] / collectable, 1) if collectable > 0 else 0.0
    # DTE pay rate over rows that have a pay state (paid + unpaid)
    payable_universe = counts["dte_paid"] + counts["dte_unpaid"]
    dte_pay_rate = round(100.0 * counts["dte_paid"] / payable_universe, 1) if payable_universe > 0 else 0.0

    aging_watch.sort(key=lambda r: r["aging_days"], reverse=True)

    # Plan vs Actual — top-down billing targets (BillingPlan) per month, scoped
    # to the same project/vendor filters and the month range when set.
    plan_stmt = select(BillingPlan)
    if ace_project_code:
        plan_stmt = plan_stmt.where(BillingPlan.ace_project_code == ace_project_code.upper())
    if vendor:
        plan_stmt = plan_stmt.where(BillingPlan.vendor == vendor.upper())
    plan_rows = (await db.execute(plan_stmt)).scalars().all()
    plan_by_month: dict[str, float] = {}
    for pr in plan_rows:
        m = (pr.month or "").strip()
        if not m:
            continue
        if mf and m < mf:
            continue
        if mt and m > mt:
            continue
        plan_by_month[m] = plan_by_month.get(m, 0.0) + float(pr.planned_amount or 0)

    all_months = sorted(set(monthly.keys()) | set(plan_by_month.keys()))
    monthly_out = [{
        "month": m,
        "plan": round(plan_by_month.get(m, 0.0), 2),
        "collected": round(monthly.get(m, {}).get("collected", 0.0), 2),
        "dte_paid": round(monthly.get(m, {}).get("dte_paid", 0.0), 2),
    } for m in all_months]
    total_plan = round(sum(plan_by_month.values()), 2)
    plan_collection_rate = round(100.0 * summary["billed_value"] / total_plan, 1) if total_plan > 0 else 0.0

    def _flatten(bucket: dict) -> list[dict]:
        return sorted(
            ({"key": k, "count": v["count"], "value": round(v["value"], 2)} for k, v in bucket.items()),
            key=lambda x: x["value"], reverse=True,
        )

    return {
        "summary": {
            **{k: round(v, 2) for k, v in summary.items()},
            **counts,
            "collection_rate": collection_rate,
            "dte_pay_rate": dte_pay_rate,
            "outstanding_value": round(summary["not_billed_value"] + summary["partial_value"], 2),
            "total_plan": total_plan,
            "plan_collection_rate": plan_collection_rate,
        },
        "by_status": _flatten(by_status),
        "by_phase": _flatten(by_phase),
        "by_owner": _flatten(by_owner),
        "by_vendor": _flatten(by_vendor),
        "by_project": _flatten(by_project),
        "monthly": monthly_out,
        "aging_watch": aging_watch[:60],
        "data": rows_out,
        "total": len(rows_out),
    }


class ReassignAceIn(BaseModel):
    po_ids: list[int]
    ace_project_code: str | None = None  # None / "" → move to NEED_REVIEW (unassigned)


@router.post("/project-pos/reassign-ace-project")
async def reassign_ace_project(
    body: ReassignAceIn,
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    """Bulk move PO lines to a different ACE project_code (manual override).
    Pass ace_project_code=null/"" to send them back to NEED_REVIEW (unassigned)."""
    if not body.po_ids:
        raise HTTPException(400, "po_ids is required")
    new_code = (body.ace_project_code or "").strip().upper() or None

    # Validate against the project catalog when a code is given
    if new_code:
        valid = {
            r[0].upper()
            for r in (await db.execute(select(ProjectCatalog.project_code))).all()
            if r[0]
        }
        if new_code not in valid:
            raise HTTPException(400, f"Unknown ACE project_code '{new_code}'")

    rows = (
        await db.execute(select(ProjectPO).where(ProjectPO.id.in_(body.po_ids)))
    ).scalars().all()
    for row in rows:
        row.ace_project_code = new_code
        row.updated_at = datetime.now(timezone.utc)
    await db.commit()
    return {"updated": len(rows), "ace_project_code": new_code}


class ReassignWorkTypeIn(BaseModel):
    po_ids: list[int]
    work_type: str | None = None  # "PAC" | "SSV" | None (Non-DT)


@router.post("/project-pos/reassign-work-type")
async def reassign_work_type(
    body: ReassignWorkTypeIn,
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    """Phase 2: set a PO's work_type (PAC → DTA, SSV → DTE, null → Non-DT).
    Used by projects/manage to split work for the Pre-Site monitors."""
    if not body.po_ids:
        raise HTTPException(400, "po_ids is required")
    wt = (body.work_type or "").strip().upper() or None
    if wt and wt not in ("PAC", "SSV"):
        raise HTTPException(400, "work_type must be PAC, SSV, or null")
    rows = (
        await db.execute(select(ProjectPO).where(ProjectPO.id.in_(body.po_ids)))
    ).scalars().all()
    for row in rows:
        row.work_type = wt
        row.updated_at = datetime.now(timezone.utc)
    await db.commit()
    return {"updated": len(rows), "work_type": wt}


@router.post("/project-pos/import")
async def import_project_pos_excel(
    request: Request,
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    import io
    import openpyxl
    from fastapi import UploadFile, File
    from starlette.datastructures import UploadFile as StarletteUploadFile

    form = await request.form()
    upload = form.get("file")
    if not upload or not hasattr(upload, "read"):
        raise HTTPException(400, "ต้องแนบไฟล์ Excel (.xlsx)")

    content = await upload.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "ไม่สามารถอ่านไฟล์ได้ — ต้องเป็น .xlsx เท่านั้น")
    ws = wb.active

    headers = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    ALIASES = {
        "po_number": ["po number", "po_number", "ponumber", "po no", "po no.", "เลขที่ po", "หมายเลข po"],
        "po_line": ["po line", "po_line", "poline", "line", "line no", "line no.", "line item"],
        "vendor": ["vendor", "vendor type", "ผู้ขาย", "hw/zte", "type"],
        "project_code": ["project code", "project_code", "project", "โปรเจค", "project id", "proj code"],
        "work_type": ["work type", "work_type", "worktype", "ssv/pac", "type of work", "งาน"],
        "du_id": ["du id", "du_id", "duid", "site id", "du", "site", "สถานี"],
        "item_dis": ["item description", "item_dis", "description", "รายการ", "item dis", "item", "detail"],
        "on_air": ["on air", "on_air", "onair", "air date", "expected date", "วันออนแอร์"],
    }

    col_idx: dict[str, int] = {}
    for field, aliases in ALIASES.items():
        for i, h in enumerate(headers):
            if h in aliases:
                col_idx[field] = i
                break

    if "po_number" not in col_idx:
        raise HTTPException(400, f"ไม่พบคอลัมน์ 'PO Number' ใน row แรกของไฟล์ (พบ: {', '.join(headers[:10])})")

    # load existing (po_number, po_line) combos for duplicate check
    existing_keys: set[tuple[str, str]] = set(
        (r[0], r[1] or "") for r in (
            await db.execute(select(ProjectPO.po_number, ProjectPO.po_line))
        ).all()
    )

    imported: list[dict] = []
    skipped: list[dict] = []
    errors: list[dict] = []

    def _cell(row_values, field):
        idx = col_idx.get(field)
        if idx is None or idx >= len(row_values):
            return None
        v = row_values[idx]
        return str(v).strip() if v is not None else None

    for row_idx, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        po_number = _cell(row_values, "po_number")
        if not po_number or po_number.lower() in ("none", ""):
            continue

        po_line = _cell(row_values, "po_line") or ""
        key = (po_number, po_line)

        if key in existing_keys:
            skipped.append({"po_number": po_number, "po_line": po_line, "reason": "duplicate"})
            continue

        vendor_raw = (_cell(row_values, "vendor") or "").upper()
        if "ZTE" in vendor_raw:
            po_target = "TE"
        else:
            po_target = "RF"

        project_code_raw = (_cell(row_values, "project_code") or "").upper() or None
        work_type_raw = (_cell(row_values, "work_type") or "").upper() or None
        if work_type_raw and work_type_raw not in ("SSV", "PAC"):
            work_type_raw = None

        try:
            body = POIn(
                po_target=po_target,
                po_number=po_number,
                po_line=po_line or None,
                project_code=project_code_raw,
                work_type=work_type_raw,
                du_id=_cell(row_values, "du_id"),
                item_dis=_cell(row_values, "item_dis"),
                on_air=_cell(row_values, "on_air"),
            )
            po_obj = ProjectPO(
                **body.model_dump(exclude={"on_air", "workflow_status", "mapping_confidence", "mapping_rule", "need_mapping_review", "expected_release_date"}),
                on_air=_date_or_none(body.on_air),
                **_infer_po_workflow_defaults(body),
                source="EXCEL_IMPORT",
            )
            db.add(po_obj)
            existing_keys.add(key)
            imported.append({"po_number": po_number, "po_line": po_line, "project_code": project_code_raw})
        except Exception as exc:
            errors.append({"row": row_idx, "po_number": po_number, "error": str(exc)})

    if imported:
        await db.commit()

    return {
        "imported": len(imported),
        "skipped": len(skipped),
        "errors": len(errors),
        "imported_rows": imported,
        "skipped_rows": skipped,
        "error_rows": errors,
    }


def _norm_hdr(h) -> str:
    """Normalize a column header for tolerant matching: lowercase, collapse
    spaces, drop trailing dots. So 'PO NO.', 'po no', 'PO  No.' all match."""
    import re as _re
    s = str(h or "").strip().lower()
    s = _re.sub(r"\s+", " ", s)
    return s.rstrip(". ")


# Normalized header → DB field. Multiple spellings (aliases) map to one field,
# so reordered OR slightly-renamed columns still resolve correctly.
HW_HEADER_ALIASES = {
    "project code": "project_code",
    "site code": "cluster_site",
    "po no": "po_number",
    "po number": "po_number",
    "po line no": "po_line",
    "po line": "po_line",            # alias: "PO Line." → po_line
    "po status": "_hw_po_status",
    "site id": "du_id",
    "item description": "item_dis",
    "line amount": "line_amount",
    "payment terms": "_payment_terms_raw",
}
# Required DB fields — import refuses if any are unresolved
HW_REQUIRED_FIELDS = {"po_number"}


def _build_hw_mapping(headers: list[str]) -> tuple[dict, dict]:
    """Return (col_idx, report). col_idx maps db_field → column index.
    report describes matched / missing / duplicate headers for the preview."""
    col_idx: dict[str, int] = {}
    matched: list[dict] = []
    duplicates: list[str] = []
    seen_fields: set[str] = set()
    for i, h in enumerate(headers):
        field = HW_HEADER_ALIASES.get(_norm_hdr(h))
        if not field:
            continue
        if field in seen_fields:
            duplicates.append(h)
            continue
        col_idx[field] = i
        seen_fields.add(field)
        matched.append({"column": h, "field": field})
    expected = set(HW_HEADER_ALIASES.values())
    missing = sorted(f for f in expected if f not in seen_fields)
    report = {
        "matched": matched,
        "missing_fields": missing,
        "duplicate_headers": duplicates,
        "required_missing": sorted(HW_REQUIRED_FIELDS - seen_fields),
        "ok": not (HW_REQUIRED_FIELDS - seen_fields),
    }
    return col_idx, report


@router.post("/project-pos/import-hw")
async def import_project_pos_hw(
    request: Request,
    dry_run: bool = Query(False),
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    """Import PO จากไฟล์ HW format. Column matching is name-based + alias +
    case-insensitive, so reordered/renamed columns still resolve.
    dry_run=true → return mapping report + would-import counts, no commit."""
    import io
    import openpyxl

    form = await request.form()
    upload = form.get("file")
    if not upload or not hasattr(upload, "read"):
        raise HTTPException(400, "ต้องแนบไฟล์ Excel (.xlsx)")

    # Optional: force every imported PO to a specific ACE project (override the
    # item-prefix rule). Used when importing per-project tracking files.
    force_ace = (form.get("force_ace_project") or "").strip().upper() or None

    content = await upload.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "ไม่สามารถอ่านไฟล์ได้ — ต้องเป็น .xlsx เท่านั้น")

    # Pick the sheet that best matches the HW PO format — score each by how many
    # known HW columns it has, require 'po no'. Robust for multi-sheet workbooks
    # where several sheets contain a 'PO No.' column.
    ws = wb.active
    best_score = -1
    for cand in wb.worksheets:
        try:
            hrow = next(cand.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration:
            continue
        norm = {_norm_hdr(v) for v in hrow if v is not None}
        if "po no" not in norm:
            continue
        score = sum(1 for k in HW_HEADER_ALIASES if k in norm)
        if score > best_score:
            best_score = score
            ws = cand

    headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(v).strip() if v is not None else "" for v in headers_row]

    col_idx, mapping_report = _build_hw_mapping(headers)
    acc_date_idx: int | None = next(
        (i for i, h in enumerate(headers) if _norm_hdr(h) == "acceptance date"), None
    )

    if "po_number" not in col_idx:
        raise HTTPException(400, "ไม่พบคอลัมน์ 'PO NO.' — กรุณาตรวจสอบว่าเป็น HW format")

    # Existing records by hw_id → (db_id, current hw_po_status)
    existing_by_hw_id: dict[str, tuple[int, str | None]] = {
        r[0]: (r[1], r[2])
        for r in (
            await db.execute(
                select(ProjectPO.hw_id, ProjectPO.id, ProjectPO.hw_po_status)
                .where(ProjectPO.hw_id.isnot(None))
            )
        ).all()
    }
    # Fallback: records without hw_id dedup by (po_number, po_line)
    existing_po_keys: set[tuple[str, str]] = set(
        (r[0], r[1] or "")
        for r in (
            await db.execute(
                select(ProjectPO.po_number, ProjectPO.po_line)
                .where(ProjectPO.hw_id.is_(None))
            )
        ).all()
    )

    imported: list[dict] = []
    updated: list[dict] = []
    skipped: list[dict] = []
    errors: list[dict] = []

    def _cell(row_values, field):
        idx = col_idx.get(field)
        if idx is None or idx >= len(row_values):
            return None
        v = row_values[idx]
        return str(v).strip() if v is not None else None

    def _infer_work_type_hw(item_dis: str | None) -> str | None:
        # SSV = Single Site Verification (per-site, → DTE)
        # PAC = Pre DT / SSOA / Cluster (cluster-level, → DTA)
        if not item_dis:
            return None
        lower = item_dis.lower()
        if "single site verification" in lower:
            return "SSV"
        if "pre dt" in lower or "ssoa" in lower or "cluster" in lower:
            return "PAC"
        return None

    def _infer_ace_project_code(item_dis: str | None) -> str | None:
        if not item_dis:
            return None
        if item_dis.startswith("B_"):
            return "HWT2304"
        if item_dis.startswith("A_"):
            return "HWT2604"
        return None

    def _parse_payment_terms(raw: str | None) -> str | None:
        import re
        if not raw:
            return None
        # Extract AC1 (XX%) and AC2 (YY%) from HW Payment Terms text
        hits = re.findall(r"AC\d+\s*\((\d+(?:\.\d+)?)%", raw)
        if len(hits) >= 2:
            p1 = round(float(hits[0]))
            p2 = round(float(hits[1]))
            return f"{p1}/{p2}"
        if len(hits) == 1:
            return "100"
        return None

    for row_idx, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        po_number = _cell(row_values, "po_number")
        if not po_number or po_number.lower() in ("none", ""):
            continue

        # po_line normalised to string
        po_line_raw = _cell(row_values, "po_line")
        try:
            po_line = str(int(float(po_line_raw))) if po_line_raw else ""
        except (ValueError, TypeError):
            po_line = po_line_raw or ""

        # HW system unique ID (21-digit)
        hw_data_val = {}
        for hi, hname in enumerate(headers):
            if hname and hi < len(row_values):
                v = row_values[hi]
                hw_data_val[hname] = str(v) if v is not None else None

        hw_id = (hw_data_val.get("ID") or "").strip() or None

        # Payment Terms + HW PO Status
        payment_terms_val = _parse_payment_terms(_cell(row_values, "_payment_terms_raw"))
        hw_po_status_val = _cell(row_values, "_hw_po_status")

        # ── UPSERT logic ──────────────────────────────────────
        if hw_id and hw_id in existing_by_hw_id:
            db_id, current_status = existing_by_hw_id[hw_id]
            if current_status != hw_po_status_val:
                if not dry_run:
                    await db.execute(
                        update(ProjectPO)
                        .where(ProjectPO.id == db_id)
                        .values(
                            hw_po_status=hw_po_status_val,
                            hw_prev_status=current_status,   # เก็บ status เดิม (เช่น OPEN → Cancelled)
                            hw_status_changed_at=datetime.now(timezone.utc),
                            hw_data=hw_data_val,
                            updated_at=datetime.now(timezone.utc),
                        )
                    )
                existing_by_hw_id[hw_id] = (db_id, hw_po_status_val)
                updated.append({"po_number": po_number, "po_line": po_line,
                                 "old_status": current_status, "new_status": hw_po_status_val})
            else:
                skipped.append({"po_number": po_number, "po_line": po_line, "reason": "no change"})
            continue

        if not hw_id and (po_number, po_line) in existing_po_keys:
            skipped.append({"po_number": po_number, "po_line": po_line, "reason": "duplicate"})
            continue
        # ──────────────────────────────────────────────────────

        project_code = (_cell(row_values, "project_code") or "").upper() or None
        cluster_site = _cell(row_values, "cluster_site")
        du_id = _cell(row_values, "du_id")
        item_dis = _cell(row_values, "item_dis")
        work_type = _infer_work_type_hw(item_dis)
        ace_project_code = force_ace or _infer_ace_project_code(item_dis)

        _la_raw = col_idx.get("line_amount")
        line_amount_val = None
        if _la_raw is not None and _la_raw < len(row_values):
            try:
                v = row_values[_la_raw]
                line_amount_val = float(v) if v is not None else None
            except (ValueError, TypeError):
                pass

        on_air_raw = None
        if acc_date_idx is not None and acc_date_idx < len(row_values):
            v = row_values[acc_date_idx]
            on_air_raw = str(v).strip() if v is not None else None

        try:
            body = POIn(
                po_target="RF",
                po_number=po_number,
                po_line=po_line or None,
                project_code=project_code,
                cluster_site=cluster_site,
                work_type=work_type,
                du_id=du_id,
                item_dis=item_dis,
                on_air=on_air_raw,
            )
            po_obj = ProjectPO(
                **body.model_dump(exclude={"on_air", "workflow_status", "mapping_confidence", "mapping_rule", "need_mapping_review", "expected_release_date"}),
                on_air=_date_or_none(body.on_air),
                **_infer_po_workflow_defaults(body),
                line_amount=line_amount_val,
                payment_terms=payment_terms_val,
                hw_po_status=hw_po_status_val,
                hw_data=hw_data_val,
                ace_project_code=ace_project_code,
                hw_id=hw_id,
                hw_first_seen_at=datetime.now(timezone.utc),   # รอบ import แรกที่เจอ (= NEW)
            )
            if not dry_run:
                db.add(po_obj)
            if hw_id:
                existing_by_hw_id[hw_id] = (0, hw_po_status_val)
            else:
                existing_po_keys.add((po_number, po_line))
            imported.append({
                "po_number": po_number,
                "po_line": po_line,
                "project_code": project_code,
                "ace_project_code": ace_project_code,
                "work_type": work_type,
                "item_dis": (item_dis or "")[:60],
            })
        except Exception as exc:
            errors.append({"row": row_idx, "po_number": po_number, "error": str(exc)})

    if not dry_run:
        if imported or updated:
            await db.commit()
        # Save import log (only on real import)
        log = HWImportLog(
            file_name=getattr(upload, "filename", None),
            file_size_kb=round(len(content) / 1024, 1),
            imported=len(imported),
            updated=len(updated),
            skipped=len(skipped),
            errors=len(errors),
            imported_by=payload.get("sub") or payload.get("employee_code"),
        )
        db.add(log)

        # Cross-cutting audit log for the Master Data tab
        from app.routers.data_imports import record_data_import
        await record_data_import(
            db,
            file_type="PO",
            file_name=getattr(upload, "filename", None),
            row_count=len(imported) + len(updated) + len(skipped),
            inserted=len(imported),
            updated=len(updated),
            skipped=len(skipped),
            status="SUCCESS" if not errors else f"PARTIAL ({len(errors)} errors)",
            uploaded_by_code=(payload.get("employee_code") or payload.get("sub")),
            uploaded_by_name=payload.get("name"),
        )
        await db.commit()
    else:
        await db.rollback()

    return {
        "dry_run": dry_run,
        "file_name": getattr(upload, "filename", None),
        "mapping": mapping_report,
        "imported": len(imported),
        "updated": len(updated),
        "skipped": len(skipped),
        "errors": len(errors),
        "imported_rows": imported,
        "updated_rows": updated,
        "skipped_rows": skipped,
        "error_rows": errors,
    }


@router.get("/project-pos/import-hw/logs")
async def list_hw_import_logs(
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    rows = (
        await db.execute(
            select(HWImportLog).order_by(HWImportLog.imported_at.desc()).limit(50)
        )
    ).scalars().all()
    return {
        "data": [
            {
                "id": r.id,
                "file_name": r.file_name,
                "file_size_kb": r.file_size_kb,
                "imported": r.imported,
                "updated": getattr(r, "updated", 0) or 0,
                "skipped": r.skipped,
                "errors": r.errors,
                "imported_by": r.imported_by,
                "imported_at": r.imported_at.isoformat() if r.imported_at else None,
            }
            for r in rows
        ]
    }


class POPlanIn(BaseModel):
    planned_dte_codes: str | None = None
    planned_dte_names: str | None = None
    planned_start_date: str | None = None
    planned_end_date: str | None = None
    # Fractional days (0.3 = quick visit ~2.4h, 1.0 = full day, up to 7.0 = week-long cluster)
    # When set, planned_end_date is auto-computed from start + ceil(duration) - 1.
    planned_duration_days: float | None = None


# Operational limits for Plan DTE
PLAN_MAX_FUTURE_DAYS = 30
PLAN_DURATION_MIN = 0.3
PLAN_DURATION_MAX = 7.0


def _validate_duration(value: float | None) -> float | None:
    """Validate planned_duration_days. None passes through. Returns rounded to 1 decimal."""
    if value is None:
        return None
    try:
        d = float(value)
    except (TypeError, ValueError):
        raise HTTPException(400, f"Invalid planned_duration_days (expected number, got '{value}')")
    if d < PLAN_DURATION_MIN or d > PLAN_DURATION_MAX:
        raise HTTPException(
            400,
            f"planned_duration_days must be between {PLAN_DURATION_MIN} and {PLAN_DURATION_MAX} days (got {d})"
        )
    return round(d, 1)


def validate_plan_dates(start_str: str | None, end_str: str | None, duration: float | None = None):
    """Parse + validate planned_start_date / planned_end_date / duration.

    Rules:
      - Format must be YYYY-MM-DD
      - Both dates must be today or future
      - Both dates must be <= today + PLAN_MAX_FUTURE_DAYS
      - start <= end (if both given)
      - Duration 0.3 ≤ d ≤ 7.0 (if given)
      - If duration given but no end_date, end_date is auto-computed: start + ceil(duration) - 1
      - start + ceil(duration) - 1 must not exceed today + PLAN_MAX_FUTURE_DAYS

    Returns (start_date, end_date, duration) — all may be None. Raises HTTPException(400)
    on any violation.
    """
    import math
    from datetime import date as _date
    today_local = _date.today()
    max_date = today_local + timedelta(days=PLAN_MAX_FUTURE_DAYS)

    def _parse(label: str, value: str | None):
        if not value:
            return None
        try:
            d = datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(400, f"Invalid {label} format (expected YYYY-MM-DD, got '{value}')")
        if d < today_local:
            raise HTTPException(400, f"{label} cannot be in the past (got {d}, today is {today_local})")
        if d > max_date:
            raise HTTPException(400, f"{label} cannot exceed {PLAN_MAX_FUTURE_DAYS} days from today (max: {max_date})")
        return d

    start = _parse("planned_start_date", start_str)
    end = _parse("planned_end_date", end_str)
    dur = _validate_duration(duration)

    # If duration given and no explicit end_date, derive end from start + ceil(dur) - 1
    if start and dur is not None and end is None:
        span_cells = max(1, math.ceil(dur))
        end = start + timedelta(days=span_cells - 1)
        if end > max_date:
            raise HTTPException(
                400,
                f"planned_start_date + duration ({dur}d → end {end}) exceeds the {PLAN_MAX_FUTURE_DAYS}-day window (max: {max_date})"
            )

    if start and end and start > end:
        raise HTTPException(400, f"planned_start_date ({start}) must be <= planned_end_date ({end})")
    return start, end, dur


@router.patch("/project-pos/{po_id}/plan")
async def patch_project_po_plan(
    po_id: int,
    body: POPlanIn,
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    row = (await db.execute(select(ProjectPO).where(ProjectPO.id == po_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "PO line not found")
    from datetime import date as date_type
    start_d, end_d, dur_d = validate_plan_dates(
        body.planned_start_date,
        body.planned_end_date,
        body.planned_duration_days,
    )
    row.planned_dte_codes = body.planned_dte_codes
    row.planned_dte_names = body.planned_dte_names
    row.planned_start_date = start_d
    row.planned_end_date   = end_d
    row.planned_duration_days = dur_d
    row.planned_at = datetime.now(timezone.utc)

    # Sync workflow_status — only adjust early-stage states so we don't
    # rewind PO lines that are already In Progress / Done / Billed / Closed
    _PRE_PLAN = {None, "", "NEW", "AUTO_MAPPED", "NEED_REVIEW", "NEED_MAPPING_REVIEW",
                 "PENDING_SITE_MAP", "SITE_MAPPED"}
    current = (row.workflow_status or "").upper() or None
    if body.planned_dte_codes:
        if current in _PRE_PLAN:
            row.workflow_status = "PLANNED"
    else:
        # plan cleared — drop back to SITE_MAPPED only if it was set by us (PLANNED)
        if current == "PLANNED":
            row.workflow_status = "SITE_MAPPED"

    # ── Auto-sync ClockApp prerequisites (shared with PAC cluster-plan) ──
    sync_summary = await ensure_clockapp_sync(db, row, body.planned_dte_codes,
                                              note_suffix=f"PO {po_id}")

    # ── Seed Pre-Site tracking immediately on plan (SSV + PAC) ──
    # SSV = 1 PO = 1 tracking row (key=po_id). PAC = N POs → 1 cluster tracking
    # row (key=cluster_key=rf_cluster_name). The Plan Board assigns a whole PAC
    # cluster by planning each child PO via this endpoint, so we must seed/update
    # the cluster tracking row here too — otherwise the assigned DTE never sees
    # the work on the Pre-Site board. (seed_tracking_from_po dedups PAC by cluster
    # and updates assigned_dte on the existing row.)
    tracking_seeded = False
    if body.planned_dte_codes and (row.work_type or "").upper() in ("SSV", "PAC"):
        from app.routers.presite_monitor import seed_tracking_from_po
        seeded = await seed_tracking_from_po(db, row)
        tracking_seeded = seeded is not None

    await db.commit()
    return {"ok": True, "id": po_id, "workflow_status": row.workflow_status,
            "tracking_seeded": tracking_seeded, **sync_summary}


class PoBillingIn(BaseModel):
    milestone: str                    # 'ac1' | 'ac2'
    plan_month: str | None = None     # "YYYY-MM" to set, "" to clear, None = leave
    billed: bool | None = None        # True = mark billed (now), False = unmark
    invoice_no: str | None = None


@router.patch("/project-pos/{po_id}/billing")
async def patch_po_billing(
    po_id: int,
    body: PoBillingIn,
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    """Set HW billing plan/actual for one AC milestone of a PO.
      - plan_month: forecast month we expect to invoice ("YYYY-MM")
      - billed: mark/unmark the milestone as invoiced (stamps *_billed_at = now)
      - invoice_no: HW invoice reference
    Keeps the legacy hw_billed_at in sync (set when every applicable AC is billed)."""
    m = (body.milestone or "").lower()
    if m not in ("ac1", "ac2"):
        raise HTTPException(400, "milestone must be 'ac1' or 'ac2'")
    row = (await db.execute(select(ProjectPO).where(ProjectPO.id == po_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "PO not found")

    if body.plan_month is not None:
        setattr(row, f"{m}_plan_month", body.plan_month.strip() or None)
    if body.invoice_no is not None:
        setattr(row, f"{m}_invoice_no", body.invoice_no.strip() or None)
    if body.billed is not None:
        setattr(row, f"{m}_billed_at", datetime.now(timezone.utc) if body.billed else None)

    # Legacy hw_billed_at = billed when all applicable AC milestones are billed.
    has_ac2 = "/" in (row.payment_terms or "")
    ac1_ok = row.ac1_billed_at is not None
    ac2_ok = (row.ac2_billed_at is not None) if has_ac2 else True
    if ac1_ok and ac2_ok:
        if row.hw_billed_at is None:
            row.hw_billed_at = datetime.now(timezone.utc)
            row.hw_billed_by = payload.get("employee_code") or payload.get("sub")
    else:
        row.hw_billed_at = None
        row.hw_billed_by = None

    row.last_action_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(row)
    return po_to_dict(row)


class PoCollectionActionIn(BaseModel):
    action: str                       # bill_ac1|bill_ac2|unbill_ac1|unbill_ac2|mark_dte_paid|unmark_dte_paid
    invoice_no: str | None = None     # for bill_ac1 / bill_ac2
    payment_ref: str | None = None    # for mark_dte_paid


@router.post("/project-pos/{po_id}/collection-action")
async def po_collection_action(
    po_id: int,
    body: PoCollectionActionIn,
    request: Request,
    payload: dict = Depends(require_po_finance_action_user),
    db: AsyncSession = Depends(get_db),
):
    """Quick finance actions from the Collection Tracking dashboard.

    Updates only the close-out money legs (HW billing / DTE payment) and the
    derived hw_billed_at sync — it intentionally does NOT rewind the workflow
    state machine, so it is safe to call from a monitoring view. Every action
    is audit-logged.
    """
    action = (body.action or "").lower().strip()
    row = (await db.execute(select(ProjectPO).where(ProjectPO.id == po_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "PO not found")

    actor = payload.get("employee_code") or payload.get("sub")
    now = datetime.now(timezone.utc)

    if action in ("bill_ac1", "bill_ac2", "unbill_ac1", "unbill_ac2"):
        m = "ac1" if action.endswith("ac1") else "ac2"
        billed = action.startswith("bill_")
        setattr(row, f"{m}_billed_at", now if billed else None)
        if billed and body.invoice_no is not None:
            setattr(row, f"{m}_invoice_no", body.invoice_no.strip() or None)
        # Keep legacy hw_billed_at in sync: billed when all applicable AC done.
        has_ac2 = "/" in (row.payment_terms or "")
        ac1_ok = row.ac1_billed_at is not None
        ac2_ok = (row.ac2_billed_at is not None) if has_ac2 else True
        if ac1_ok and ac2_ok:
            if row.hw_billed_at is None:
                row.hw_billed_at = now
                row.hw_billed_by = actor
        else:
            row.hw_billed_at = None
            row.hw_billed_by = None
    elif action == "mark_dte_paid":
        row.dte_paid_at = now
        row.dte_paid_by = actor
    elif action == "unmark_dte_paid":
        row.dte_paid_at = None
        row.dte_paid_by = None
    else:
        raise HTTPException(400, f"Unknown action '{action}'")

    row.last_action_at = now
    await write_audit_log(
        db, action=f"po_collection_{action}", entity_type="project_po", entity_id=str(row.id),
        payload=payload,
        new_value={"po_number": row.po_number, "po_line": row.po_line,
                   "invoice_no": body.invoice_no, "payment_ref": body.payment_ref},
        request=request, source="PO Collection Dashboard",
    )
    await db.commit()
    await db.refresh(row)
    rec = po_to_dict(row)
    rec["billing_state"] = _po_billing_state(row)
    rec["dte_pay_state"] = _po_dte_pay_state(row)
    return rec


async def ensure_clockapp_sync(db: AsyncSession, row: ProjectPO,
                                dte_code: str | None,
                                *, note_suffix: str = "") -> dict:
    """Ensure assignment + auth_user + clock_site exist so ClockApp can serve this DTE.

    Called from PATCH /project-pos/{id}/plan (SSV) and from POST /presite/cluster-plan (PAC).
    Idempotent — safe to call multiple times. Caller is responsible for db.commit().

    Returns: dict with assignment_created / auth_user_synced / clock_site_created flags.
    """
    from datetime import date as date_type
    summary = {"assignment_created": False, "clock_site_created": False, "auth_user_synced": False}
    if not dte_code:
        return summary
    dte_code = dte_code.strip()
    if not dte_code:
        return summary

    # Resolve the project code. SSV/PAC PO lines imported from HW often carry a NULL
    # ace_project_code (notably PAC SSOA clusters); fall back to the project registered
    # on the DU's clock_site so ClockApp assignment + cluster site still get created.
    du_code = (row.du_id or "").strip()
    project_code = row.ace_project_code
    if not project_code and du_code:
        du_cs = (await db.execute(
            select(ClockSite).where(func.upper(ClockSite.site_code) == du_code.upper())
        )).scalar_one_or_none()
        if du_cs and du_cs.project_code:
            project_code = du_cs.project_code
    if not project_code:
        return summary

    # (a) ensure project assignment (DTE PER_SITE) exists
    existing_assign = (await db.execute(
        select(ProjectAssignment).where(
            ProjectAssignment.project_code == project_code,
            ProjectAssignment.employee_code == dte_code,
            ProjectAssignment.role_in_project == "DTE",
        )
    )).scalar_one_or_none()
    if not existing_assign:
        db.add(ProjectAssignment(
            project_code=project_code,
            employee_code=dte_code,
            role_in_project="DTE",
            clock_type="PER_SITE",
            is_active=True,
            start_date=date_type.today(),
            notes=f"Auto-created by Plan DTE ({note_suffix})" if note_suffix else "Auto-created by Plan DTE",
        ))
        summary["assignment_created"] = True
    elif existing_assign.clock_type != "PER_SITE" or not existing_assign.is_active:
        existing_assign.clock_type = "PER_SITE"
        existing_assign.is_active = True

    # (b) sync auth_user — position_code='DTE' + clock_type='PER_SITE'
    auth_user = (await db.execute(
        select(AuthUser).where(AuthUser.employee_code == dte_code)
    )).scalar_one_or_none()
    if auth_user:
        changed = False
        if auth_user.position_code != "DTE":
            auth_user.position_code = "DTE"; changed = True
        if auth_user.clock_type != "PER_SITE":
            auth_user.clock_type = "PER_SITE"; changed = True
        summary["auth_user_synced"] = changed

    # (c) ensure clock_site exists for the cluster_site root
    if row.cluster_site:
        site_key = row.cluster_site.split("_")[0].strip()
        if site_key:
            existing_cs = (await db.execute(
                select(ClockSite).where(func.upper(ClockSite.site_code) == site_key.upper())
            )).scalar_one_or_none()
            if not existing_cs:
                psite = (await db.execute(
                    select(ProjectSite).where(func.upper(ProjectSite.site_code) == site_key.upper())
                )).scalar_one_or_none()
                # PAC cluster names (e.g. "EAS0049-SSOA-1") have no project_site of their
                # own — borrow coords/name from the cluster's DU so GPS radius works.
                if (psite is None or psite.lat is None) and du_code:
                    du_site = (await db.execute(
                        select(ProjectSite).where(func.upper(ProjectSite.site_code) == du_code.upper())
                    )).scalar_one_or_none()
                    if du_site is not None and du_site.lat is not None:
                        psite = du_site
                db.add(ClockSite(
                    site_code=site_key,
                    site_name=psite.site_name if psite else None,
                    customer=psite.customer if psite else None,
                    project_code=project_code,
                    lat=psite.lat if psite else None,
                    lng=psite.lng if psite else None,
                    gps_radius_m=(psite.gps_radius_m if psite else None) or 500,
                    is_active=True,
                ))
                summary["clock_site_created"] = True
    return summary


@router.get("/project-pos/pipeline")
async def project_po_pipeline(
    ace_project_code: str = Query("HWT2304"),
    work_type: str = Query("SSV"),
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    """Return per-PO pipeline data joining project_pos with clock_sessions.

    For each SSV PO line (or whatever work_type filter), aggregate the
    drive-test sessions logged by the assigned DTE at the matching site.
    Stages: SSV Plan → DT (clock in/out) → Reporting → Done.
    """
    pos = (await db.execute(
        select(ProjectPO).where(
            ProjectPO.ace_project_code == ace_project_code,
            ProjectPO.work_type == work_type,
        )
    )).scalars().all()

    # Map site_code → list of (dte_code, sessions) — bulk-load clock sessions
    # We only care about sessions belonging to the planned DTE for each PO row.
    site_codes = {((p.cluster_site or "").split("_")[0] or "").upper() for p in pos if p.cluster_site}
    dte_codes  = {p.planned_dte_codes for p in pos if p.planned_dte_codes}

    sessions = []
    if site_codes and dte_codes:
        sessions = (await db.execute(
            select(ClockSession).where(
                func.upper(ClockSession.site_code).in_(list(site_codes)),
                ClockSession.employee_code.in_(list(dte_codes)),
            ).order_by(ClockSession.clock_in_at)
        )).scalars().all()

    # Build a lookup: (site_code, employee_code) → [sessions]
    session_index: dict[tuple[str, str], list[ClockSession]] = {}
    for s in sessions:
        key = ((s.site_code or "").upper(), s.employee_code)
        session_index.setdefault(key, []).append(s)

    data = []
    for p in pos:
        site_key = ((p.cluster_site or "").split("_")[0] or "").upper()
        sess = session_index.get((site_key, p.planned_dte_codes or ""), [])
        dt_started = sess[0].clock_in_at if sess else None
        dt_done    = None
        dt_outcome = None  # COMPLETE | ISSUE | STOP | None
        # Determine "DT done": last session with clock_out_at and a terminal outcome
        # COMPLETE → fully done; ISSUE → done with problem; STOP → abandoned
        finished_sess = [s for s in sess if s.clock_out_at and (s.status or "").upper() in {"COMPLETE", "COMPLETED", "DONE", "CLOSED", "ISSUE"}]
        if finished_sess:
            latest = max(finished_sess, key=lambda s: s.clock_out_at)
            dt_done = latest.clock_out_at
            dt_outcome = (latest.status or "").upper()
            if dt_outcome == "COMPLETED":
                dt_outcome = "COMPLETE"

        # Use explicit work_done_at if set, else derive from clock sessions
        effective_dt_done = p.work_done_at.isoformat() if p.work_done_at else (dt_done.isoformat() if dt_done else None)
        data.append({
            "po_id": p.id,
            "site_code": p.cluster_site or "",
            "site_key": site_key,
            "po_number": p.po_number,
            "po_line": p.po_line,
            "item_dis": p.item_dis,
            "workflow_status": p.workflow_status,
            "planned_dte_codes": p.planned_dte_codes,
            "planned_dte_names": p.planned_dte_names,
            "planned_start_date": p.planned_start_date.isoformat() if p.planned_start_date else None,
            "planned_end_date":   p.planned_end_date.isoformat()   if p.planned_end_date   else None,
            "planned_duration_days": float(p.planned_duration_days) if p.planned_duration_days is not None else None,
            "planned_at": p.planned_at.isoformat() if p.planned_at else None,
            "dt_started_at": dt_started.isoformat() if dt_started else None,
            "dt_done_at":    effective_dt_done,
            "dt_outcome":    dt_outcome,
            "dt_sessions":   len(sess),
            # Reporting & Done — reuse leader_checked_at / approved_at
            "reported_at":   p.leader_checked_at.isoformat() if p.leader_checked_at else None,
            "reported_by":   p.leader_code,
            "done_at":       p.approved_at.isoformat() if p.approved_at else None,
        })
    return {"data": data, "total": len(data)}


class POMarkIn(BaseModel):
    note: str | None = None


@router.patch("/project-pos/{po_id}/mark-reported")
async def patch_project_po_mark_reported(
    po_id: int,
    body: POMarkIn,
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    row = (await db.execute(select(ProjectPO).where(ProjectPO.id == po_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "PO line not found")
    row.leader_checked_at = datetime.now(timezone.utc)
    row.leader_code       = payload.get("employee_code") or payload.get("sub") or "system"
    if body.note:
        row.leader_note   = body.note
    # advance workflow if reasonable (not already past)
    if (row.workflow_status or "").upper() in {"PLANNED", "IN_PROGRESS", "WORK_DONE", "SITE_MAPPED"}:
        row.workflow_status = "LEADER_CHECKING"
    await db.commit()
    return {"ok": True, "id": po_id, "workflow_status": row.workflow_status, "reported_at": row.leader_checked_at.isoformat()}


@router.patch("/project-pos/{po_id}/mark-done")
async def patch_project_po_mark_done(
    po_id: int,
    body: POMarkIn,
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    row = (await db.execute(select(ProjectPO).where(ProjectPO.id == po_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "PO line not found")
    row.approved_at = datetime.now(timezone.utc)
    if (row.workflow_status or "").upper() not in {"CLOSED"}:
        row.workflow_status = "LEADER_APPROVED"
    # Auto-seed Pre-Site Monitor tracking row (Day 0 = full_onair_at)
    from app.routers.presite_monitor import seed_tracking_from_po
    try:
        await seed_tracking_from_po(db, row)
    except Exception:
        pass  # don't block mark-done if presite seeding fails
    await db.commit()
    return {"ok": True, "id": po_id, "workflow_status": row.workflow_status, "done_at": row.approved_at.isoformat()}


@router.patch("/project-pos/{po_id}/unmark")
async def patch_project_po_unmark(
    po_id: int,
    stage: str = Query("done", description="reported | done"),
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    row = (await db.execute(select(ProjectPO).where(ProjectPO.id == po_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "PO line not found")
    if stage == "done":
        row.approved_at = None
        # Roll back from LEADER_APPROVED to LEADER_CHECKING (or PLANNED if never reported)
        if (row.workflow_status or "").upper() == "LEADER_APPROVED":
            row.workflow_status = "LEADER_CHECKING" if row.leader_checked_at else "PLANNED"
    elif stage == "reported":
        row.leader_checked_at = None
        row.leader_code = None
        # Roll back from LEADER_CHECKING/APPROVED to PLANNED
        if (row.workflow_status or "").upper() in {"LEADER_CHECKING", "LEADER_APPROVED"}:
            row.workflow_status = "PLANNED" if row.planned_dte_codes else "SITE_MAPPED"
            row.approved_at = None  # also clear done if it was set
    else:
        raise HTTPException(400, "stage must be 'reported' or 'done'")
    await db.commit()
    return {"ok": True, "id": po_id, "workflow_status": row.workflow_status}


@router.post("/project-pos", status_code=201)
async def create_project_po(
    body: POIn,
    request: Request,
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    payload_data = body.model_dump(exclude={
        "on_air",
        "workflow_status",
        "mapping_confidence",
        "mapping_rule",
        "need_mapping_review",
        "expected_release_date",
    })
    row = ProjectPO(
        **payload_data,
        on_air=_date_or_none(body.on_air),
        **_infer_po_workflow_defaults(body),
    )
    db.add(row)
    await db.commit()
    await db.refresh(row)
    await write_audit_log(
        db,
        action="po_created",
        entity_type="project_po",
        entity_id=row.id,
        payload=payload,
        new_value=po_to_dict(row),
        changed_fields=["po_number", "project_code", "workflow_status"],
        request=request,
        source="PO Finance",
    )
    await db.commit()
    return po_to_dict(row)


@router.post("/project-pos/{po_id}/workflow")
async def update_project_po_workflow(
    po_id: int,
    body: POWorkflowIn,
    request: Request,
    payload: dict = Depends(require_project_user),
    db: AsyncSession = Depends(get_db),
):
    row = (await db.execute(select(ProjectPO).where(ProjectPO.id == po_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "PO line not found")
    if row.locked and body.action.upper() not in {"REVISE"}:
        raise HTTPException(400, "Approved PO scope is locked. Create a revision before changing workflow.")

    action = body.action.upper().replace("-", "_")
    old = po_to_dict(row)
    now = _now_utc()

    if action == "ASSIGN_PROJECT":
        # NEED_MAPPING_REVIEW → Project assigns project_code + optional site_code
        if not body.project_code:
            raise HTTPException(400, "project_code is required for ASSIGN_PROJECT.")
        row.project_code = body.project_code.upper()
        if body.site_code:
            row.site_code = body.site_code.upper()
            row.workflow_status = "SITE_MAPPED"
        else:
            row.workflow_status = "PENDING_SITE_MAP"
        row.current_owner_role = "PROJECT"
        row.mapping_confidence = 100
        row.mapping_rule = "Manually assigned by Project"
        row.need_mapping_review = False
    elif action == "MAP_SITE":
        # PENDING_SITE_MAP → Project maps to site_code
        if not body.site_code:
            raise HTTPException(400, "site_code is required for MAP_SITE.")
        row.site_code = body.site_code.upper()
        row.workflow_status = "SITE_MAPPED"
        row.current_owner_role = "PROJECT"
    elif action == "FINANCE_CONFIRM":
        # AUTO_MAPPED / NEED_REVIEW → Finance confirms mapping → send to Project for planning
        if body.project_code:
            row.project_code = body.project_code.upper()
        if body.work_type:
            wt = body.work_type.upper()
            if wt not in PO_WORK_TYPES:
                raise HTTPException(400, f"Invalid work_type '{wt}'. Must be SSV or PAC.")
            row.work_type = wt
        if not row.project_code:
            raise HTTPException(400, "project_code is required before confirming.")
        if not row.work_type:
            raise HTTPException(400, "work_type (SSV/PAC) is required before confirming.")
        row.workflow_status = "PENDING_SITE_MAP"
        row.current_owner_role = "PROJECT"
        row.finance_checked_at = now
        row.need_mapping_review = False
        row.mapping_confidence = 100
        row.mapping_rule = "Confirmed by Finance"
    elif action == "HOLD":
        row.workflow_status = "FINANCE_HOLD"
        row.current_owner_role = "FINANCE"
        row.hold_reason = body.hold_reason or body.note or "Finance hold"
        row.expected_release_date = _date_or_none(body.expected_release_date)
    elif action == "READY":
        row.workflow_status = "READY_TO_SEND_PROJECT"
        row.current_owner_role = "FINANCE"
        row.hold_reason = None
        row.finance_checked_at = now
    elif action == "SEND_PROJECT":
        if not row.project_code:
            raise HTTPException(400, "Cannot send to project without project_code.")
        row.workflow_status = "WAITING_PROJECT_CHECK"
        row.current_owner_role = "PROJECT"
        row.sent_to_project_at = now
    elif action == "PROJECT_ACCEPT":
        row.workflow_status = "PROJECT_ACCEPTED"
        row.current_owner_role = "FINANCE"
        row.project_accepted_at = now
    elif action == "PENDING_APPROVAL":
        row.workflow_status = "PENDING_APPROVAL"
        row.current_owner_role = "APPROVER"
    elif action == "APPROVE":
        row.workflow_status = "APPROVED"
        row.current_owner_role = "PROJECT"
        row.approved_at = now
        row.locked = True
    elif action == "REJECT":
        row.workflow_status = "REJECTED"
        row.current_owner_role = "FINANCE"
        row.hold_reason = body.note or body.hold_reason or "Rejected by approver"
    elif action == "RETURN_FINANCE":
        row.workflow_status = "RETURNED_TO_FINANCE"
        row.current_owner_role = "FINANCE"
        row.hold_reason = body.note or body.hold_reason
    elif action == "RETURN_PROJECT":
        row.workflow_status = "RETURNED_TO_PROJECT"
        row.current_owner_role = "PROJECT"
        row.hold_reason = body.note or body.hold_reason
    elif action == "REVISE":
        row.revision = (row.revision or 1) + 1
        row.locked = False
        row.workflow_status = "FINANCE_RECHECK"
        row.current_owner_role = "FINANCE"
        row.hold_reason = body.note or "Revision opened after approval"

    # ── Plan DTE ──────────────────────────────────────────────────────────────
    elif action == "PLAN_DTE":
        if not body.planned_dte_codes:
            raise HTTPException(400, "planned_dte_codes is required.")
        if not body.planned_start_date:
            raise HTTPException(400, "planned_start_date is required.")
        row.planned_dte_codes = body.planned_dte_codes
        row.planned_dte_names = body.planned_dte_names
        row.planned_start_date = _date_or_none(body.planned_start_date)
        row.planned_end_date = _date_or_none(body.planned_end_date)
        row.planned_at = now
        row.workflow_status = "PLANNED"
        row.current_owner_role = "DTE"
    elif action == "START_WORK":
        row.workflow_status = "IN_PROGRESS"
        row.current_owner_role = "DTE"
        row.work_started_at = now
    elif action == "WORK_DONE":
        row.workflow_status = "WORK_DONE"
        row.current_owner_role = "LEADER"
        row.work_done_at = now

    # ── Leader Review ─────────────────────────────────────────────────────────
    elif action == "LEADER_REVIEW":
        # Leader starts review — moves to LEADER_CHECKING
        row.workflow_status = "LEADER_CHECKING"
        row.current_owner_role = "LEADER"
        row.leader_code = body.leader_code or row.current_owner_user
        row.leader_checked_at = now
    elif action == "LEADER_RETURN":
        # Leader rejects — send back to DTE
        row.workflow_status = "IN_PROGRESS"
        row.current_owner_role = "DTE"
        row.leader_note = body.leader_note or body.note or "Returned by Leader: redo work"
    elif action == "LEADER_APPROVE":
        # Leader approved work quality → move to LEADER_APPROVED
        row.workflow_status = "LEADER_APPROVED"
        row.current_owner_role = "LEADER"
        row.leader_note = body.leader_note or body.note
    elif action == "CONFIRM_TIMESHEET":
        # Leader confirms DTE timesheet → Finance pays DTE
        row.workflow_status = "PENDING_PAYMENT"
        row.current_owner_role = "FINANCE"
    elif action == "CONFIRM_HW_EVIDENCE":
        # Leader confirms HW acceptance docs → Finance bills HW
        if not body.hw_evidence_url and not row.hw_evidence_url:
            raise HTTPException(400, "hw_evidence_url is required.")
        if body.hw_evidence_url:
            row.hw_evidence_url = body.hw_evidence_url
        row.hw_evidence_confirmed_at = now
        row.workflow_status = "PENDING_BILLING"
        row.current_owner_role = "FINANCE"

    # ── Finance Close-out ─────────────────────────────────────────────────────
    elif action == "PAY_DTE":
        # Finance pays DTE team
        row.dte_paid_at = now
        row.dte_paid_by = body.current_owner_user or payload.get("email") or payload.get("name")
        row.workflow_status = "DTE_PAID" if not row.hw_billed_at else "CLOSED"
        row.current_owner_role = "FINANCE" if not row.hw_billed_at else "CLOSED"
        if row.workflow_status == "CLOSED":
            row.locked = True
    elif action == "BILL_HW":
        # Finance issues invoice to HW
        row.hw_billed_at = now
        row.hw_billed_by = body.current_owner_user or payload.get("email") or payload.get("name")
        row.hw_invoice_no = body.hw_invoice_no
        row.workflow_status = "HW_BILLED" if not row.dte_paid_at else "CLOSED"
        row.current_owner_role = "FINANCE" if not row.dte_paid_at else "CLOSED"
        if row.workflow_status == "CLOSED":
            row.locked = True

    else:
        raise HTTPException(400, "Unsupported workflow action.")

    row.current_owner_user = body.current_owner_user or payload.get("email") or payload.get("name")
    row.last_action_at = now
    await db.commit()
    await db.refresh(row)
    new = po_to_dict(row)
    await write_audit_log(
        db,
        action=f"po_workflow_{action.lower()}",
        entity_type="project_po",
        entity_id=row.id,
        payload=payload,
        old_value=old,
        new_value=new,
        changed_fields=[key for key in new if old.get(key) != new.get(key)],
        request=request,
        source="PO Finance",
    )
    await db.commit()
    return new


@router.get("/clock-permissions/{employee_code}")
async def clock_permissions(
    employee_code: str,
    payload: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    require_self_or_admin(payload, employee_code)
    employee = (await db.execute(select(Employee).where(Employee.employee_code == employee_code))).scalar_one_or_none()
    auth_user = (await db.execute(select(AuthUser).where(AuthUser.employee_code == employee_code))).scalar_one_or_none()
    assignments = (await db.execute(
        select(ProjectAssignment)
        .where(ProjectAssignment.employee_code == employee_code, ProjectAssignment.is_active == True)
        .order_by(ProjectAssignment.project_code)
    )).scalars().all()
    project_codes = [row.project_code for row in assignments]
    dte_assignments = [row for row in assignments if row.role_in_project == "DTE"]
    per_site_project_codes = [
        row.project_code for row in dte_assignments
        if _normalize_assignment_clock_type(row.role_in_project, row.clock_type) == "PER_SITE"
    ]
    site_rows = []
    if per_site_project_codes:
        site_rows = (await db.execute(
            select(ClockSite)
            .where(ClockSite.project_code.in_(per_site_project_codes), ClockSite.is_active == True)
            .order_by(ClockSite.site_code)
        )).scalars().all()
    roles = sorted({row.role_in_project for row in assignments})

    # Fallback: if no assignments, use employee's own project_role + project_code
    if not roles and employee and employee.project_role:
        roles = [employee.project_role]
    elif not roles and auth_user and auth_user.position_code:
        roles = [auth_user.position_code]

    if per_site_project_codes:
        clock_type = "PER_SITE"
    elif auth_user and auth_user.clock_type in ("DAILY", "PER_SITE"):
        clock_type = auth_user.clock_type
    else:
        clock_type = "DAILY"

    # Planned sites — PO lines where this employee is the planned DTE and still actionable
    # (Exclude WORK_DONE/LEADER_CHECKING etc. — work is finished, no need to clock again)
    planned_pos = (await db.execute(
        select(ProjectPO).where(
            ProjectPO.planned_dte_codes == employee_code,
            # PLANNED = SSV plan; LEADER_APPROVED = PAC cluster-plan; both still actionable
            ProjectPO.workflow_status.in_(["PLANNED", "IN_PROGRESS", "LEADER_APPROVED"]),
        )
    )).scalars().all()
    planned_site_keys = sorted({(p.cluster_site or "").split("_")[0].upper() for p in planned_pos if p.cluster_site})
    planned_site_rows: list[ClockSite] = []
    planned_meta: dict[str, dict] = {}
    if planned_site_keys:
        planned_site_rows = (await db.execute(
            select(ClockSite)
            .where(func.upper(ClockSite.site_code).in_(planned_site_keys), ClockSite.is_active == True)
            .order_by(ClockSite.site_code)
        )).scalars().all()
        # Index PO context per site for the UI (status, planned_start_date, item_dis, po_id)
        for p in planned_pos:
            if not p.cluster_site:
                continue
            sk = p.cluster_site.split("_")[0].upper()
            meta = planned_meta.setdefault(sk, {
                "po_ids": [], "workflow_statuses": set(),
                "planned_start_date": None, "planned_end_date": None,
                "work_type": None, "item_dis": p.item_dis,
                "cluster_name": p.cluster_site,  # full SSV/PAC name, e.g. "CBR7523_NewSite_East R3"
            })
            meta["po_ids"].append(p.id)
            if p.workflow_status:
                meta["workflow_statuses"].add(p.workflow_status)
            if p.planned_start_date and (meta["planned_start_date"] is None or p.planned_start_date < meta["planned_start_date"]):
                meta["planned_start_date"] = p.planned_start_date
            if p.planned_end_date and (meta["planned_end_date"] is None or p.planned_end_date > meta["planned_end_date"]):
                meta["planned_end_date"] = p.planned_end_date
            if p.work_type and not meta["work_type"]:
                meta["work_type"] = p.work_type

    def _planned_site_dict(s: ClockSite) -> dict:
        d = site_to_dict(s)
        m = planned_meta.get((s.site_code or "").upper(), {})
        d.update({
            "planned_po_ids": m.get("po_ids", []),
            "planned_po_count": len(m.get("po_ids", [])),
            "workflow_statuses": sorted(list(m.get("workflow_statuses", []))),
            "planned_start_date": m["planned_start_date"].isoformat() if m.get("planned_start_date") else None,
            "planned_end_date":   m["planned_end_date"].isoformat()   if m.get("planned_end_date")   else None,
            "work_type": m.get("work_type"),
            "item_description": m.get("item_dis"),
            "cluster_name": m.get("cluster_name"),
        })
        # Show the full plan name (e.g. "CBR7523_NewSite_East R3") in ClockApp when the
        # clock_site itself has no name — keeps board + ClockApp labels consistent.
        if not d.get("site_name"):
            d["site_name"] = m.get("cluster_name")
        return d

    return {
        "employee": employee_to_dict(employee) if employee else None,
        "assignments": [assignment_to_dict(row, employee) for row in assignments],
        "sites": [site_to_dict(row) for row in site_rows],
        "planned_sites": [_planned_site_dict(s) for s in planned_site_rows],
        "roles": roles,
        "clock_type": clock_type,
        "photo_required": any(role == "DTE" for role in roles),
        "gps_required": True,   # GPS required for all roles (office or per-site)
    }


async def seed_employee_database(db: AsyncSession) -> None:
    await ensure_employee_hr_columns(db)
    has_rows = (await db.execute(select(func.count()).select_from(Employee))).scalar_one()
    if has_rows:
        await backfill_employee_hr_columns(db)
        return

    seed_path = Path(__file__).resolve().parents[1] / "employee_seed.json"
    rows = json.loads(seed_path.read_text(encoding="utf-8"))

    projects: dict[str, dict] = {}
    for idx, item in enumerate(rows, start=1):
        code = f"ACE-XLSX-{idx:03d}"
        db.add(
            Employee(
                employee_code=code,
                email=item.get("email") or None,
                full_name=item["name"],
                first_name=str(item["name"]).split()[0] if item.get("name") else None,
                last_name=" ".join(str(item["name"]).split()[1:]) if item.get("name") else None,
                department="Project",
                job_title=item.get("position"),
                employment_type="FULL_TIME",
                contract_type="FULL_TIME",
                nationality="Thai",
                project_team=item["team"],
                project_role=item.get("position"),
                project_code=item.get("project_code"),
                project_name=item.get("project_name"),
                position=item["team"],
                status="ACTIVE",
            )
        )
        project_code = item.get("project_code")
        if project_code:
            entry = projects.setdefault(
                project_code,
                {
                    "project_code": project_code,
                    "project_name": item.get("project_name") or project_code,
                    "team": item["team"],
                    "headcount": 0,
                },
            )
            entry["headcount"] += 1

    for project in projects.values():
        db.add(ProjectCatalog(**project))

    await db.commit()
    await backfill_employee_hr_columns(db)


async def ensure_employee_hr_columns(db: AsyncSession) -> None:
    existing = {
        row[0]
        for row in (
            await db.execute(text("""
                select column_name
                from information_schema.columns
                where table_name = 'employees'
            """))
        ).all()
    }
    for column, column_type in EMPLOYEE_HR_COLUMNS.items():
        if column not in existing:
            await db.execute(text(f"alter table employees add column {column} {column_type}"))
    assignment_existing = {
        row[0]
        for row in (
            await db.execute(text("""
                select column_name
                from information_schema.columns
                where table_name = 'project_assignments_live'
            """))
        ).all()
    }
    for column, column_type in PROJECT_ASSIGNMENT_COLUMNS.items():
        if column not in assignment_existing:
            await db.execute(text(f"alter table project_assignments_live add column {column} {column_type}"))
    po_existing = {
        row[0]
        for row in (
            await db.execute(text("""
                select column_name
                from information_schema.columns
                where table_name = 'project_pos'
            """))
        ).all()
    }
    for column, column_type in PROJECT_PO_COLUMNS.items():
        if column not in po_existing:
            await db.execute(text(f"alter table project_pos add column {column} {column_type}"))
    site_existing = {
        row[0]
        for row in (
            await db.execute(text("""
                select column_name
                from information_schema.columns
                where table_name = 'project_sites'
            """))
        ).all()
    }
    for column, column_type in PROJECT_SITE_COLUMNS.items():
        if column not in site_existing:
            await db.execute(text(f"alter table project_sites add column {column} {column_type}"))
    await db.execute(text("""
        update project_assignments_live
        set clock_type = case when role_in_project = 'DTE' then 'PER_SITE' else 'DAILY' end
        where clock_type is null
    """))
    await db.execute(text("""
        update project_pos
        set
            ace_project_code = case
                when item_dis like 'B\\_%' escape '\\' then 'HWT2304'
                when item_dis like 'A\\_%' escape '\\' then 'HWT2604'
                else ace_project_code
            end,
            work_type = case
                when work_type is null and item_dis ilike '%single site verification%' then 'SSV'
                when work_type is null and (item_dis ilike '%ssoa%' or item_dis ilike '%cluster%') then 'PAC'
                else work_type
            end
        where hw_data is not null and (ace_project_code is null or work_type is null)
    """))
    await db.execute(text("""
        update project_pos
        set
            workflow_status = case
                when workflow_status is null then case when project_code is null then 'NEED_MAPPING_REVIEW' else 'AUTO_MAPPED' end
                when workflow_status = 'NEW' and project_code is not null then 'AUTO_MAPPED'
                else workflow_status
            end,
            mapping_confidence = case
                when mapping_confidence is null or mapping_confidence = 0 then case when project_code is null then 20 else 85 end
                else mapping_confidence
            end,
            mapping_rule = coalesce(mapping_rule, case when project_code is null then 'No project match' else 'Existing project code' end),
            need_mapping_review = case
                when need_mapping_review is null then project_code is null
                when project_code is not null and workflow_status in ('NEW', 'AUTO_MAPPED') then false
                else need_mapping_review
            end,
            current_owner_role = coalesce(current_owner_role, 'FINANCE'),
            revision = coalesce(revision, 1),
            locked = coalesce(locked, false),
            last_action_at = coalesce(last_action_at, updated_at, created_at)
    """))
    await db.commit()


async def backfill_employee_hr_columns(db: AsyncSession) -> None:
    rows = (await db.execute(select(Employee))).scalars().all()
    changed = False
    for row in rows:
        if (not row.first_name or not row.last_name) and row.full_name:
            parts = row.full_name.split()
            row.first_name = row.first_name or (parts[0] if parts else None)
            row.last_name = row.last_name or (" ".join(parts[1:]) if len(parts) > 1 else None)
            changed = True
        if not row.job_title and row.project_role:
            row.job_title = row.project_role
            changed = True
        if not row.employment_type:
            row.employment_type = "FULL_TIME"
            changed = True
        if not row.contract_type:
            row.contract_type = "FULL_TIME"
            changed = True
        if not row.nationality:
            row.nationality = "Thai"
            changed = True
    if changed:
        await db.commit()


async def seed_project_operations(db: AsyncSession) -> None:
    assignment_count = (await db.execute(select(func.count()).select_from(ProjectAssignment))).scalar_one()
    if not assignment_count:
        employees = (await db.execute(select(Employee).where(Employee.department == "Project"))).scalars().all()
        for employee in employees:
            if not employee.project_code:
                continue
            role = _role_from_employee(employee)
            db.add(ProjectAssignment(
                project_code=employee.project_code,
                employee_code=employee.employee_code,
                role_in_project=role,
                clock_type=_normalize_assignment_clock_type(role, None),
                start_date=employee.hire_date,
                allocation_pct=100,
                is_active=True,
            ))

    site_count = (await db.execute(select(func.count()).select_from(ProjectSite))).scalar_one()
    if not site_count:
        seed_sites = [
            ("AIS2601", "SITE-AIS-BKK-0421", "Sukhumvit 21", "AIS", 13.7433, 100.5588, 500, "Bangkok", "Watthana"),
            ("AIS2601", "SITE-AIS-BKK-0422", "Asok BTS Area", "AIS", 13.7456, 100.5602, 300, "Bangkok", "Khlong Toei"),
            ("HWT2306", "SITE-TRUE-NNT-0118", "Nonthaburi Tower", "TRUE", 13.8591, 100.5134, 500, "Nonthaburi", "Mueang"),
            ("NBTC2501", "SITE-AIS-RYG-0055", "Rangsit Tower", "AIS", 14.0233, 100.6177, 400, "Pathum Thani", "Rangsit"),
            ("HWT2301", "SITE-DTAC-BKK-0201", "Silom Center", "DTAC", 13.7234, 100.5260, 400, "Bangkok", "Bang Rak"),
            ("WW2503", "SITE-NT-CNX-0031", "Chiangmai NT Tower", "NT", 18.7883, 98.9853, 600, "Chiang Mai", "Mueang"),
        ]
        for project_code, site_code, site_name, customer, lat, lng, radius, province, district in seed_sites:
            db.add(ProjectSite(
                project_code=project_code,
                site_code=site_code,
                site_name=site_name,
                customer=customer,
                lat=lat,
                lng=lng,
                gps_radius_m=radius,
                province=province,
                district=district,
            ))
            clock_site = (await db.execute(select(ClockSite).where(ClockSite.site_code == site_code))).scalar_one_or_none()
            if clock_site:
                clock_site.project_code = project_code

    po_count = (await db.execute(select(func.count()).select_from(ProjectPO))).scalar_one()
    if not po_count:
        db.add(ProjectPO(
            po_target="RF",
            project_code="AIS2601",
            po_number="PO-RF-2026-001",
            po_line="001",
            du_id="DU-BKK-0421",
            item_dis="Drive Test Service",
            cluster_site="BKK-Cluster-A / SITE-AIS-BKK-0421",
            owner="Peerapol Piamsri",
            lat_long="13.7433, 100.5588",
            on_air=_date_or_none("2026-05-01"),
            cluster_type="RF / Urban",
        ))
        db.add(ProjectPO(
            po_target="RF",
            project_code="AIS2601",
            po_number="PO-RF-2026-001",
            po_line="002",
            du_id="DU-BKK-0422",
            item_dis="Report Preparation",
            cluster_site="BKK-Cluster-A / SITE-AIS-BKK-0422",
            owner="Kannika Phanit",
            lat_long="13.7456, 100.5602",
            on_air=_date_or_none("2026-05-03"),
            cluster_type="RF / Urban",
        ))

    await db.commit()
