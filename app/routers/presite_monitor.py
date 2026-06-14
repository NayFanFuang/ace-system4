"""Pre-Site Monitor — DTE per-site workflow tracker (Phase 1: Steps 1–6).

Auto-seeded from rf-monitor Pipeline when a PO is marked done (LEADER_APPROVED).
State machine:
  FULL_ONAIR → DT_DONE → REPORT_DONE → CHECKING ⇄ (FAIL rework) → ACE_SUBMITTED
    → TL_REVIEWED → PM_REVIEWED → ACE_APPROVED
"""

import asyncio
import os
import re as _re_path
from datetime import datetime, timezone, timedelta
from typing import Optional

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import FileResponse
from pydantic import BaseModel
from sqlalchemy import and_, delete, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.services import ftp_service

from app.database import get_db
from app.deps import get_current_user
import re
from app.models.employee import ProjectPO, ProjectSite
from app.models.presite_tracking import (
    ALL_STAGES,
    STAGE_ACE_APPROVED,
    STAGE_ACE_SUBMITTED,
    STAGE_CHECKING,
    STAGE_DT_DONE,
    STAGE_DT_STARTED,
    STAGE_FULL_ONAIR,
    STAGE_PM_REVIEWED,
    STAGE_REPORT_DONE,
    STAGE_TL_REVIEWED,
    DtePresiteHistory,
    DtePresiteSession,
    DtePresiteTracking,
)

router = APIRouter(prefix="/api/presite", tags=["PreSite Monitor"])

SLA_DAYS = 3


# ────────────────────────────────────────────────────────────────────────────
# Role helpers (action-level)
# ────────────────────────────────────────────────────────────────────────────
ROLE_DTE = {"EMPLOYEE", "PROJECT_ADMIN", "PM", "SUPER_ADMIN", "SYSTEM_ADMIN"}
ROLE_TL  = {"PROJECT_ADMIN", "PM", "SUPER_ADMIN", "SYSTEM_ADMIN"}
ROLE_PM  = {"PM", "SUPER_ADMIN", "SYSTEM_ADMIN"}
ROLE_VIEW = {"EMPLOYEE", "PROJECT_ADMIN", "PM", "DIRECTOR", "HR_ADMIN", "HR_VIEWER", "SUPER_ADMIN", "SYSTEM_ADMIN"}

ACTION_ROLE_MAP: dict[str, set[str]] = {
    "dt-start":      ROLE_DTE,
    "dt-done":       ROLE_DTE,
    "report-done":   ROLE_DTE,
    "check-pass":    ROLE_TL,
    "check-fail":    ROLE_TL,
}


def _can(role: str | None, action: str) -> bool:
    allowed = ACTION_ROLE_MAP.get(action)
    return bool(role and allowed and role in allowed)


# ────────────────────────────────────────────────────────────────────────────
# Serialization
# ────────────────────────────────────────────────────────────────────────────
def _is_sla_breached(row: DtePresiteTracking, now: datetime) -> bool:
    if not row.full_onair_at:
        return False
    # SLA = check_result PASS within 3 days from Day 0
    if (row.check_result or "").upper() == "PASS":
        # Pass on time?
        return bool(row.check_at and (row.check_at - row.full_onair_at) > timedelta(days=SLA_DAYS))
    return (now - row.full_onair_at) > timedelta(days=SLA_DAYS)


def _iso(dt: datetime | None) -> str | None:
    return dt.isoformat() if dt else None


def _row_to_dict(row: DtePresiteTracking, now: datetime) -> dict:
    return {
        "id": row.id,
        "ace_project_code": row.ace_project_code,
        "site_code": row.site_code,
        "po_id": row.po_id,
        "po_number": row.po_number,
        "po_line": row.po_line,
        "assigned_dte_code": row.assigned_dte_code,
        "assigned_dte_name": row.assigned_dte_name,
        "work_type": row.work_type,
        "rf_cluster_name": row.rf_cluster_name,
        "cluster_key": row.cluster_key,
        "cluster_ready_at": _iso(row.cluster_ready_at),
        # Billing handoff fields (added 2026-05-22)
        "du_id": row.du_id,
        "item_dis": row.item_dis,
        "hw_id": row.hw_id,
        "line_amount": float(row.line_amount) if row.line_amount is not None else None,
        "payment_terms": row.payment_terms,
        "billing_sent_at": _iso(row.billing_sent_at),
        "billing_sent_by": row.billing_sent_by,
        "billing_ref": row.billing_ref,
        "site_status": row.site_status,
        "dta_code": row.dta_code,
        "dta_name": row.dta_name,
        "layers": row.layers,
        "total_rounds": row.total_rounds,
        "full_onair_at": _iso(row.full_onair_at),
        "dt_started_at": _iso(row.dt_started_at),
        "dt_started_by": row.dt_started_by,
        "dt_done_at": _iso(row.dt_done_at),
        "dt_done_by": row.dt_done_by,
        "report_done_at": _iso(row.report_done_at),
        "report_done_by": row.report_done_by,
        "check_at": _iso(row.check_at),
        "check_by": row.check_by,
        "check_result": row.check_result,
        "check_notes": row.check_notes,
        "rework_count": row.rework_count,
        "ace_submit_at": _iso(row.ace_submit_at),
        "ace_submit_by": row.ace_submit_by,
        "ace_report_url": row.ace_report_url,
        "tl_review_at": _iso(row.tl_review_at),
        "tl_review_by": row.tl_review_by,
        "pm_review_at": _iso(row.pm_review_at),
        "pm_review_by": row.pm_review_by,
        "ace_approve_at": _iso(row.ace_approve_at),
        "ace_approve_by": row.ace_approve_by,
        # PAC late stages
        "pa_open_at": _iso(row.pa_open_at),
        "pa_open_by": row.pa_open_by,
        "pa_closed_at": _iso(row.pa_closed_at),
        "pa_closed_by": row.pa_closed_by,
        "report_submit_at": _iso(row.report_submit_at),
        "report_submit_by": row.report_submit_by,
        "report_approved_at": _iso(row.report_approved_at),
        "report_approved_by": row.report_approved_by,
        "current_stage": row.current_stage,
        "completed_at": _iso(row.completed_at),
        "sla_breached": _is_sla_breached(row, now),
        # DTE report upload (.rar) — added 2026-05-29
        "report_filename": row.report_filename,
        "report_file_size": row.report_file_size,
        "report_uploaded_at": _iso(row.report_uploaded_at),
        "report_uploaded_by": row.report_uploaded_by,
        "report_version": row.report_version or 0,
        "has_report": bool(row.report_file_path),
        # DTE payment status
        "paid": bool(row.dte_paid_at),
        "dte_paid_at": _iso(row.dte_paid_at),
        "dte_paid_by": row.dte_paid_by,
        "dte_payment_ref": row.dte_payment_ref,
        "created_at": _iso(row.created_at),
        "updated_at": _iso(row.updated_at),
    }


# ────────────────────────────────────────────────────────────────────────────
# Endpoints — read
# ────────────────────────────────────────────────────────────────────────────
@router.get("/tracking")
async def list_tracking(
    ace_project_code: Optional[str] = Query(None),
    dte: Optional[str] = Query(None, description="filter by assigned_dte_code (substring)"),
    stage: Optional[str] = Query(None),
    sla_breach: Optional[bool] = Query(None),
    q: Optional[str] = Query(None, description="search site_code / dte name"),
    payload: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if payload.get("role") not in ROLE_VIEW:
        raise HTTPException(403, "Not allowed")
    stmt = select(DtePresiteTracking)
    if ace_project_code:
        stmt = stmt.where(DtePresiteTracking.ace_project_code == ace_project_code)
    if dte:
        stmt = stmt.where(DtePresiteTracking.assigned_dte_code.ilike(f"%{dte}%"))
    if stage:
        stmt = stmt.where(DtePresiteTracking.current_stage == stage)
    if q:
        like = f"%{q}%"
        stmt = stmt.where(or_(
            DtePresiteTracking.site_code.ilike(like),
            DtePresiteTracking.assigned_dte_name.ilike(like),
            DtePresiteTracking.assigned_dte_code.ilike(like),
        ))
    stmt = stmt.order_by(DtePresiteTracking.full_onair_at.desc().nullslast(), DtePresiteTracking.id.desc())
    rows = (await db.execute(stmt)).scalars().all()
    now = datetime.now(timezone.utc)
    data = [_row_to_dict(r, now) for r in rows]
    if sla_breach is not None:
        data = [d for d in data if d["sla_breached"] == sla_breach]

    # Attach session rounds for each row (SSV + PAC; PAC has dynamic round count)
    tracking_ids = [r.id for r in rows]
    if tracking_ids:
        sess_rows = (await db.execute(
            select(DtePresiteSession)
            .where(DtePresiteSession.tracking_id.in_(tracking_ids))
            .order_by(DtePresiteSession.tracking_id, DtePresiteSession.round_number)
        )).scalars().all()
        sessions_by_tid: dict[int, list[dict]] = {}
        for s in sess_rows:
            sessions_by_tid.setdefault(s.tracking_id, []).append({
                "id": s.id,
                "round": s.round_number,
                "planned_at": _iso(s.planned_at),
                "planned_by": s.planned_by,
                "started_at": _iso(s.started_at),
                "started_by": s.started_by,
                "ended_at": _iso(s.ended_at),
                "ended_by": s.ended_by,
                "check_at": _iso(s.check_at),
                "check_by": s.check_by,
                "check_result": s.check_result,
                "check_notes": s.check_notes,
                "status": s.status,
                "notes": s.notes,
            })
        for d in data:
            d["sessions"] = sessions_by_tid.get(d["id"], [])
    else:
        for d in data:
            d["sessions"] = []
    return {"data": data, "total": len(data)}


@router.get("/summary")
async def summary(
    ace_project_code: Optional[str] = Query(None),
    payload: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if payload.get("role") not in ROLE_VIEW:
        raise HTTPException(403, "Not allowed")
    stmt = select(DtePresiteTracking)
    if ace_project_code:
        stmt = stmt.where(DtePresiteTracking.ace_project_code == ace_project_code)
    rows = (await db.execute(stmt)).scalars().all()
    now = datetime.now(timezone.utc)
    week_ago = now - timedelta(days=7)
    stage_counts: dict[str, int] = {s: 0 for s in ALL_STAGES}
    sla_breach = 0
    approved_this_week = 0
    awaiting_check = 0
    active = 0
    for r in rows:
        st = r.current_stage or STAGE_FULL_ONAIR
        stage_counts[st] = stage_counts.get(st, 0) + 1
        if st != STAGE_ACE_APPROVED:
            active += 1
        if st == STAGE_CHECKING:
            awaiting_check += 1
        if _is_sla_breached(r, now) and st != STAGE_ACE_APPROVED:
            sla_breach += 1
        if r.ace_approve_at and r.ace_approve_at >= week_ago:
            approved_this_week += 1
    return {
        "total": len(rows),
        "active": active,
        "awaiting_check": awaiting_check,
        "sla_breach": sla_breach,
        "approved_this_week": approved_this_week,
        "stage_counts": stage_counts,
    }


@router.get("/tracking/{tracking_id}/history")
async def list_history(
    tracking_id: int,
    payload: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if payload.get("role") not in ROLE_VIEW:
        raise HTTPException(403, "Not allowed")
    rows = (await db.execute(
        select(DtePresiteHistory)
        .where(DtePresiteHistory.tracking_id == tracking_id)
        .order_by(DtePresiteHistory.at.desc(), DtePresiteHistory.id.desc())
    )).scalars().all()
    return {"data": [{
        "id": h.id,
        "stage": h.stage,
        "action": h.action,
        "actor_code": h.actor_code,
        "actor_name": h.actor_name,
        "notes": h.notes,
        "at": _iso(h.at),
    } for h in rows]}


@router.get("/tracking/{tracking_id}/billing-pos")
async def list_billing_pos(
    tracking_id: int,
    payload: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List billing-relevant POs underneath a tracking row.

    For SSV: returns just the single PO referenced by tracking.po_id.
    For PAC: returns all POs in the cluster that match `B_Cluster%` (billable items).
    """
    if payload.get("role") not in ROLE_VIEW:
        raise HTTPException(403, "Not allowed")
    t = (await db.execute(select(DtePresiteTracking).where(DtePresiteTracking.id == tracking_id))).scalar_one_or_none()
    if not t:
        raise HTTPException(404, "Tracking not found")

    if (t.work_type or "").upper() == "PAC":
        if not t.cluster_key:
            return {"data": []}
        pos = (await db.execute(
            select(ProjectPO).where(
                ProjectPO.cluster_site == t.cluster_key,
                ProjectPO.work_type == "PAC",
                ProjectPO.item_dis.op("~*")(r"^[AB]_(Cluster|SSOA) "),
            ).order_by(ProjectPO.du_id)
        )).scalars().all()
    else:
        if not t.po_id:
            return {"data": []}
        pos = (await db.execute(
            select(ProjectPO).where(ProjectPO.id == t.po_id)
        )).scalars().all()

    return {"data": [{
        "po_id": p.id,
        "po_number": p.po_number,
        "po_line": p.po_line,
        "du_id": p.du_id,
        "item_dis": p.item_dis,
        "hw_id": p.hw_id or (p.hw_data.get("ID") if isinstance(p.hw_data, dict) else None),
        "line_amount": float(p.line_amount) if p.line_amount is not None else None,
        "payment_terms": p.payment_terms,
        "work_type": p.work_type,
        "workflow_status": p.workflow_status,
    } for p in pos]}


# ────────────────────────────────────────────────────────────────────────────
# Endpoints — advance / undo
# ────────────────────────────────────────────────────────────────────────────
class AdvanceIn(BaseModel):
    action: str
    notes: Optional[str] = None
    ace_report_url: Optional[str] = None


def _record_history(
    db: AsyncSession,
    *,
    tracking_id: int,
    stage: str,
    action: str,
    actor_code: str | None,
    actor_name: str | None,
    notes: str | None,
) -> None:
    db.add(DtePresiteHistory(
        tracking_id=tracking_id,
        stage=stage,
        action=action,
        actor_code=actor_code,
        actor_name=actor_name,
        notes=notes,
    ))


@router.post("/tracking/{tracking_id}/advance")
async def advance(
    tracking_id: int,
    body: AdvanceIn,
    payload: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    action = (body.action or "").lower().strip()
    role = payload.get("role")
    if not _can(role, action):
        raise HTTPException(403, f"Role {role} cannot perform action '{action}'")
    row = (await db.execute(select(DtePresiteTracking).where(DtePresiteTracking.id == tracking_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Tracking row not found")

    actor_code = payload.get("employee_code") or payload.get("sub")
    actor_name = payload.get("name") or payload.get("preferred_name") or actor_code
    now = datetime.now(timezone.utc)
    stage = row.current_stage

    # Stage transition guards + apply
    if action == "dt-start":
        if stage != STAGE_FULL_ONAIR:
            raise HTTPException(409, f"Cannot dt-start from stage {stage}")
        row.dt_started_at = now
        row.dt_started_by = actor_code
        row.current_stage = STAGE_DT_STARTED

    elif action == "dt-done":
        if stage not in (STAGE_DT_STARTED, STAGE_FULL_ONAIR):
            raise HTTPException(409, f"Cannot dt-done from stage {stage}")
        # Allow dt-done directly (no start was recorded) — backfill start = now
        if not row.dt_started_at:
            row.dt_started_at = now
            row.dt_started_by = actor_code
        row.dt_done_at = now
        row.dt_done_by = actor_code
        row.current_stage = STAGE_DT_DONE

    elif action == "report-done":
        if stage not in (STAGE_DT_DONE, STAGE_CHECKING):
            raise HTTPException(409, f"Cannot report-done from stage {stage}")
        row.report_done_at = now
        row.report_done_by = actor_code
        # Clear prior FAIL when re-submitting after rework
        row.check_result = None
        row.check_at = None
        row.check_by = None
        row.check_notes = None
        row.current_stage = STAGE_REPORT_DONE

    elif action == "check-pass":
        if stage not in (STAGE_REPORT_DONE, STAGE_CHECKING):
            raise HTTPException(409, f"Cannot check-pass from stage {stage}")
        row.check_at = now
        row.check_by = actor_code
        row.check_result = "PASS"
        row.check_notes = body.notes
        row.completed_at = now
        row.current_stage = STAGE_ACE_APPROVED
        # Also mark latest session as DONE+PASS
        latest = (await db.execute(
            select(DtePresiteSession).where(DtePresiteSession.tracking_id == row.id)
            .order_by(DtePresiteSession.round_number.desc()).limit(1)
        )).scalar_one_or_none()
        if latest:
            latest.check_at = now; latest.check_by = actor_code
            latest.check_result = "PASS"; latest.check_notes = body.notes
            latest.status = "DONE"

    elif action == "check-fail":
        if stage not in (STAGE_REPORT_DONE, STAGE_CHECKING):
            raise HTTPException(409, f"Cannot check-fail from stage {stage}")
        if not body.notes:
            raise HTTPException(400, "Notes required for check-fail (rework reason)")
        row.check_at = now
        row.check_by = actor_code
        row.check_result = "FAIL"
        row.check_notes = body.notes
        row.rework_count = (row.rework_count or 0) + 1
        row.current_stage = STAGE_CHECKING
        # Also mark latest session as DONE+FAIL (so next clock-in creates new round)
        latest = (await db.execute(
            select(DtePresiteSession).where(DtePresiteSession.tracking_id == row.id)
            .order_by(DtePresiteSession.round_number.desc()).limit(1)
        )).scalar_one_or_none()
        if latest:
            latest.check_at = now; latest.check_by = actor_code
            latest.check_result = "FAIL"; latest.check_notes = body.notes
            latest.status = "DONE"

    else:
        raise HTTPException(400, f"Unknown action: {action}")

    _record_history(
        db,
        tracking_id=row.id,
        stage=row.current_stage,
        action=action,
        actor_code=actor_code,
        actor_name=actor_name,
        notes=body.notes,
    )
    await db.commit()
    await db.refresh(row)
    return _row_to_dict(row, now)


class UndoIn(BaseModel):
    stage: str  # the stage to undo (e.g. "DT_DONE" undoes dt_done_at, returns to FULL_ONAIR)
    notes: Optional[str] = None


@router.post("/tracking/{tracking_id}/undo")
async def undo(
    tracking_id: int,
    body: UndoIn,
    payload: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    role = payload.get("role")
    if role not in ROLE_TL:
        raise HTTPException(403, "Only TL/PM/Admin may undo")
    row = (await db.execute(select(DtePresiteTracking).where(DtePresiteTracking.id == tracking_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Tracking row not found")
    stage = (body.stage or "").upper()
    actor_code = payload.get("employee_code") or payload.get("sub")
    actor_name = payload.get("name") or actor_code
    now = datetime.now(timezone.utc)

    if stage == STAGE_DT_STARTED:
        row.dt_started_at = None
        row.dt_started_by = None
        row.current_stage = STAGE_FULL_ONAIR
    elif stage == STAGE_DT_DONE:
        row.dt_done_at = None
        row.dt_done_by = None
        row.current_stage = STAGE_DT_STARTED if row.dt_started_at else STAGE_FULL_ONAIR
    elif stage == STAGE_REPORT_DONE:
        row.report_done_at = None
        row.report_done_by = None
        row.current_stage = STAGE_DT_DONE
    elif stage == STAGE_CHECKING:
        row.check_at = None
        row.check_by = None
        row.check_result = None
        row.check_notes = None
        row.current_stage = STAGE_REPORT_DONE
    elif stage == STAGE_ACE_SUBMITTED:
        row.ace_submit_at = None
        row.ace_submit_by = None
        row.ace_report_url = None
        row.current_stage = STAGE_CHECKING
    elif stage == STAGE_TL_REVIEWED:
        row.tl_review_at = None
        row.tl_review_by = None
        row.current_stage = STAGE_ACE_SUBMITTED
    elif stage == STAGE_PM_REVIEWED:
        row.pm_review_at = None
        row.pm_review_by = None
        row.current_stage = STAGE_TL_REVIEWED
    elif stage == STAGE_ACE_APPROVED:
        row.ace_approve_at = None
        row.ace_approve_by = None
        row.completed_at = None
        row.current_stage = STAGE_PM_REVIEWED
    else:
        raise HTTPException(400, f"Cannot undo stage: {stage}")

    _record_history(
        db,
        tracking_id=row.id,
        stage=row.current_stage,
        action=f"undo-{stage.lower()}",
        actor_code=actor_code,
        actor_name=actor_name,
        notes=body.notes,
    )
    await db.commit()
    await db.refresh(row)
    return _row_to_dict(row, now)


# ────────────────────────────────────────────────────────────────────────────
# Auto-seed — called from employees.py:mark-done hook
# ────────────────────────────────────────────────────────────────────────────
async def seed_tracking_from_po(db: AsyncSession, po: ProjectPO) -> DtePresiteTracking | None:
    """Idempotent seed of tracking row from a PO.

      SSV: 1 PO = 1 DU + 1 item_dis = 1 tracking row (key=po_id)
      PAC: 1 cluster = N POs = 1 tracking row (key=cluster_key=cluster_site)
           When called with any PO of the cluster, attaches to (or creates) the cluster row.
    """
    if not po:
        return None
    work_type = (po.work_type or "").upper()
    is_pac = work_type == "PAC"
    ssv_site_code = None  # effective SSV site identity (site_code → cluster root → du_id)

    # Resolve site metadata up-front. ISDP import stores BOTH Col B (Site Code) and
    # Col C (DU ID, e.g. "RYG7235_Flash_RAN_EAS R3") as site_code keys in
    # project_sites — both pointing to the same Col I (RF Cluster Name, e.g.
    # "EAS-FLASH-0012"). Try every candidate so 22 child POs with different
    # cluster_site values can resolve to the same canonical cluster identity.
    candidate_codes: list[str] = []
    for c in (po.cluster_site, po.du_id, po.site_code):
        if c and c not in candidate_codes:
            candidate_codes.append(c)
    site = None
    for code in candidate_codes:
        site = (await db.execute(
            select(ProjectSite).where(ProjectSite.site_code == code)
        )).scalar_one_or_none()
        if site:
            break

    # Canonical RF Cluster Name (from ISDP Col I) wins; cluster_site is fallback.
    canonical_cluster = (site.rf_cluster_name if site else None) or po.cluster_site

    if is_pac:
        cluster_key = canonical_cluster
        if not cluster_key:
            return None
        existing = (await db.execute(
            select(DtePresiteTracking).where(
                and_(DtePresiteTracking.work_type == "PAC",
                     DtePresiteTracking.cluster_key == cluster_key)
            )
        )).scalar_one_or_none()
    else:
        # SSV site identity: real SSV POs often carry only du_id + cluster_site
        # (site_code is blank). Fall back so seeding works at plan-time and clock-in.
        ssv_site_code = (po.site_code or "").strip()
        if not ssv_site_code and po.cluster_site:
            ssv_site_code = po.cluster_site.split("_")[0].strip()
        if not ssv_site_code:
            ssv_site_code = (po.du_id or "").strip()
        if not ssv_site_code:
            return None
        existing = (await db.execute(
            select(DtePresiteTracking).where(DtePresiteTracking.po_id == po.id)
        )).scalar_one_or_none()

    if existing:
        # Keep the assigned DTE in sync when (re)planning — the Plan Board assigns
        # a whole PAC cluster (or an SSV site) to a DTE via this path, so the
        # tracking row must reflect the latest planned DTE.
        planned_code = (po.planned_dte_codes or "").split(",")[0].strip() or None
        planned_name = (po.planned_dte_names or "").split(",")[0].strip() or None
        if planned_code and existing.assigned_dte_code != planned_code:
            existing.assigned_dte_code = planned_code
            existing.assigned_dte_name = planned_name or existing.assigned_dte_name
            await db.flush()
        return existing
    full_onair = po.approved_at or po.work_done_at or datetime.now(timezone.utc)

    cluster_ready_at = None
    if site and site.cluster_ready:
        cluster_ready_at = datetime(site.cluster_ready.year, site.cluster_ready.month, site.cluster_ready.day, tzinfo=timezone.utc)
    rf_cluster_name = canonical_cluster

    # Parse layers from item_dis (e.g. "for 8 layer" or "for 4~8 layer")
    layers = None
    if po.item_dis:
        m = re.search(r"for\s+\d+~(\d+)\s+layer", po.item_dis, re.IGNORECASE)
        if not m:
            m = re.search(r"for\s+(\d+)\s+layer", po.item_dis, re.IGNORECASE)
        if m:
            try: layers = int(m.group(1))
            except (ValueError, TypeError): pass

    # total_rounds is dynamic — grows as PM adds rounds. SSV always 1, PAC starts 0 (none planned yet)
    total_rounds = 1 if (po.work_type or "").upper() == "SSV" else 0

    # Billing handoff fields — pull from PO + hw_data fallback
    hw_id = po.hw_id
    if not hw_id and po.hw_data and isinstance(po.hw_data, dict):
        hw_id = po.hw_data.get("ID")

    # SSV: identity tied to PO (1 PO = 1 DU + 1 item_dis = 1 tracking)
    # PAC: identity tied to cluster (1 cluster = N POs = 1 tracking). du_id/item_dis/hw_id NULL
    #      because they vary per child PO; use /tracking/{id}/billing-pos to list them.
    row = DtePresiteTracking(
        ace_project_code=po.ace_project_code,
        # Display name: PAC shows the canonical RF Cluster Name (ISDP Col I) so the
        # tracking row represents "1 cluster" even when child POs carry different
        # cluster_site values. SSV uses the full cluster_site name (e.g.
        # "CBR7523_NewSite_East R3"), falling back to the GPS-root code only when
        # cluster_site is blank.
        site_code=(canonical_cluster if is_pac else (po.cluster_site or ssv_site_code)),
        po_id=(None if is_pac else po.id),
        po_number=(None if is_pac else po.po_number),
        po_line=(None if is_pac else po.po_line),
        cluster_key=(canonical_cluster if is_pac else None),
        assigned_dte_code=(po.planned_dte_codes or "").split(",")[0].strip() or None,
        assigned_dte_name=(po.planned_dte_names or "").split(",")[0].strip() or None,
        work_type=po.work_type,
        rf_cluster_name=rf_cluster_name,
        cluster_ready_at=cluster_ready_at,
        layers=layers,
        total_rounds=total_rounds,
        full_onair_at=full_onair,
        current_stage=STAGE_FULL_ONAIR,
        # Billing handoff — SSV only at tracking level. For PAC, billing fields are per child PO.
        du_id=(None if is_pac else po.du_id),
        item_dis=(None if is_pac else po.item_dis),
        hw_id=(None if is_pac else hw_id),
        line_amount=(None if is_pac else po.line_amount),
        payment_terms=(None if is_pac else po.payment_terms),
    )
    db.add(row)
    await db.flush()
    _record_history(
        db,
        tracking_id=row.id,
        stage=STAGE_FULL_ONAIR,
        action="auto-seed",
        actor_code="SYSTEM",
        actor_name="Auto-seed from Pipeline",
        notes=f"Seeded from PO {po.po_number}/{po.po_line} (LEADER_APPROVED)",
    )
    return row


# ────────────────────────────────────────────────────────────────────────────
# Cluster-level Plan DTE (PAC) — assign 1 DTE to whole cluster (fanout to all POs)
# ────────────────────────────────────────────────────────────────────────────
class ClusterPlanIn(BaseModel):
    cluster_key: str
    dte_code: str
    dte_name: Optional[str] = None
    # NEW (2026-05-28): support scheduled dates + fractional-day duration
    planned_start_date: Optional[str] = None  # YYYY-MM-DD
    planned_end_date: Optional[str] = None    # YYYY-MM-DD (optional; auto-computed if duration given)
    planned_duration_days: Optional[float] = None  # 0.3 ≤ d ≤ 7.0


@router.post("/cluster-plan")
async def plan_pac_cluster(
    body: ClusterPlanIn,
    payload: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Plan DTE at PAC cluster level: assign 1 DTE to ALL B_Cluster* POs in the cluster,
    then seed (or attach to) the cluster-level tracking row.
    """
    if payload.get("role") not in ROLE_PM:
        raise HTTPException(403, "Only PM/Admin may plan PAC clusters")

    # Validate dates + duration via shared helper from employees router
    from app.routers.employees import validate_plan_dates
    start_d, end_d, dur_d = validate_plan_dates(
        body.planned_start_date,
        body.planned_end_date,
        body.planned_duration_days,
    )

    pos = (await db.execute(
        select(ProjectPO).where(
            func.upper(ProjectPO.cluster_site) == body.cluster_key.upper(),
            ProjectPO.work_type == "PAC",
            ProjectPO.item_dis.op("~*")(r"^[AB]_(Cluster|SSOA) "),
        )
    )).scalars().all()
    if not pos:
        raise HTTPException(404, f"No PAC Cluster/SSOA POs found for cluster '{body.cluster_key}'")

    now = datetime.now(timezone.utc)
    for po in pos:
        po.planned_dte_codes = body.dte_code
        po.planned_dte_names = body.dte_name or body.dte_code
        po.planned_start_date = start_d
        po.planned_end_date = end_d
        po.planned_duration_days = dur_d
        po.planned_at = now
        po.workflow_status = "LEADER_APPROVED"
        if not po.approved_at:
            po.approved_at = now

    await db.flush()
    # Seed the cluster tracking from the first PO (idempotent on cluster_key)
    tracking = await seed_tracking_from_po(db, pos[0])

    # Auto-sync ClockApp prerequisites (assignment + auth_user + clock_site)
    # so DTE sees the cluster site in the ClockApp PER_SITE list immediately.
    # We use the first PO as the representative (cluster_site is the same across all).
    from app.routers.employees import ensure_clockapp_sync
    sync_summary = await ensure_clockapp_sync(
        db, pos[0], body.dte_code, note_suffix=f"cluster {body.cluster_key}"
    )

    await db.commit()
    if tracking:
        await db.refresh(tracking)

    return {
        "cluster_key": body.cluster_key,
        "pos_updated": len(pos),
        "tracking_id": tracking.id if tracking else None,
        "tracking_stage": tracking.current_stage if tracking else None,
        **sync_summary,
    }


class ClusterUnplanIn(BaseModel):
    cluster_key: str
    force: bool = False  # if True, delete tracking even when work has started


@router.post("/cluster-unplan")
async def unplan_pac_cluster(
    body: ClusterUnplanIn,
    payload: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Remove DTE assignment from PAC cluster.

    Safety rule:
      - If cluster tracking is at FULL_ONAIR (no clock-in yet) → safely delete tracking + history
      - If tracking is past FULL_ONAIR (work started) → keep tracking; require `force=True` to delete
      - Always clear planned_dte_codes/names on all PAC POs in cluster
    """
    if payload.get("role") not in ROLE_PM:
        raise HTTPException(403, "Only PM/Admin may unplan PAC clusters")

    # Find cluster tracking
    tracking = (await db.execute(
        select(DtePresiteTracking).where(
            and_(DtePresiteTracking.work_type == "PAC",
                 DtePresiteTracking.cluster_key == body.cluster_key)
        )
    )).scalar_one_or_none()

    deleted_tracking = False
    if tracking:
        if tracking.current_stage == STAGE_FULL_ONAIR or body.force:
            await db.execute(delete(DtePresiteHistory).where(DtePresiteHistory.tracking_id == tracking.id))
            await db.execute(delete(DtePresiteSession).where(DtePresiteSession.tracking_id == tracking.id))
            await db.execute(delete(DtePresiteTracking).where(DtePresiteTracking.id == tracking.id))
            deleted_tracking = True
        else:
            # Work already in progress — block unless force=True
            raise HTTPException(
                409,
                f"Cluster '{body.cluster_key}' tracking is at stage {tracking.current_stage}. "
                f"Pass force=true to unplan anyway (will delete tracking + history)."
            )

    # Clear all PAC POs in cluster (B_Cluster/SSOA + Pre DT items not affected — they're SSV now)
    pos = (await db.execute(
        select(ProjectPO).where(
            func.upper(ProjectPO.cluster_site) == body.cluster_key.upper(),
            ProjectPO.work_type == "PAC",
            ProjectPO.item_dis.op("~*")(r"^[AB]_(Cluster|SSOA) "),
        )
    )).scalars().all()
    for po in pos:
        po.planned_dte_codes = None
        po.planned_dte_names = None
        po.workflow_status = "AUTO_MAPPED"
        po.approved_at = None

    await db.commit()
    return {
        "cluster_key": body.cluster_key,
        "pos_unplanned": len(pos),
        "tracking_deleted": deleted_tracking,
    }


# ────────────────────────────────────────────────────────────────────────────
# PAC endpoints — dynamic rounds + 4 late stages
# ────────────────────────────────────────────────────────────────────────────
class PacSessionIn(BaseModel):
    notes: Optional[str] = None


@router.post("/tracking/{tracking_id}/sessions/add")
async def add_session(
    tracking_id: int,
    body: PacSessionIn,
    payload: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Add a new planned round (PENDING session). No backend cap — UI shows max 5."""
    if payload.get("role") not in ROLE_TL:
        raise HTTPException(403, "Only TL/PM/Admin may add rounds")
    row = (await db.execute(select(DtePresiteTracking).where(DtePresiteTracking.id == tracking_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Tracking row not found")
    existing = (await db.execute(
        select(func.max(DtePresiteSession.round_number)).where(DtePresiteSession.tracking_id == tracking_id)
    )).scalar_one() or 0
    next_round = existing + 1
    sess = DtePresiteSession(tracking_id=tracking_id, round_number=next_round, status="PENDING", notes=body.notes)
    db.add(sess)
    row.total_rounds = next_round
    _record_history(db, tracking_id=tracking_id, stage=row.current_stage, action=f"plan-round-{next_round}",
                    actor_code=payload.get("employee_code"), actor_name=payload.get("name"),
                    notes=body.notes or f"Round {next_round} planned")
    await db.commit()
    return {"ok": True, "round": next_round, "total_rounds": next_round}


@router.post("/tracking/{tracking_id}/sessions/{round_number}/start")
async def start_session(
    tracking_id: int,
    round_number: int,
    body: PacSessionIn,
    payload: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if round_number < 1:
        raise HTTPException(400, "round_number must be >= 1")
    actor = payload.get("employee_code") or payload.get("sub")
    sess = (await db.execute(
        select(DtePresiteSession).where(
            and_(DtePresiteSession.tracking_id == tracking_id, DtePresiteSession.round_number == round_number)
        )
    )).scalar_one_or_none()
    now = datetime.now(timezone.utc)
    if not sess:
        sess = DtePresiteSession(tracking_id=tracking_id, round_number=round_number)
        db.add(sess)
    sess.started_at = now
    sess.started_by = actor
    sess.status = "IN_PROGRESS"
    sess.notes = body.notes
    await db.commit()
    await db.refresh(sess)
    return {"ok": True, "round": sess.round_number, "started_at": _iso(sess.started_at), "status": sess.status}


@router.post("/tracking/{tracking_id}/sessions/{round_number}/end")
async def end_session(
    tracking_id: int,
    round_number: int,
    body: PacSessionIn,
    payload: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    actor = payload.get("employee_code") or payload.get("sub")
    sess = (await db.execute(
        select(DtePresiteSession).where(
            and_(DtePresiteSession.tracking_id == tracking_id, DtePresiteSession.round_number == round_number)
        )
    )).scalar_one_or_none()
    if not sess or not sess.started_at:
        raise HTTPException(409, "Cannot end a session that hasn't started")
    sess.ended_at = datetime.now(timezone.utc)
    sess.ended_by = actor
    sess.status = "DONE"
    if body.notes:
        sess.notes = body.notes
    await db.commit()
    return {"ok": True, "round": sess.round_number, "ended_at": _iso(sess.ended_at), "status": "DONE"}


class PacStageIn(BaseModel):
    action: str   # pa-open | pa-closed | report-submit | report-approve | undo-<action>
    notes: Optional[str] = None


@router.post("/tracking/{tracking_id}/pac-stage")
async def pac_stage(
    tracking_id: int,
    body: PacStageIn,
    payload: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    role = payload.get("role")
    if role not in ROLE_TL:
        raise HTTPException(403, "Only TL/PM/Admin may set PAC stage")
    row = (await db.execute(select(DtePresiteTracking).where(DtePresiteTracking.id == tracking_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Tracking row not found")
    now = datetime.now(timezone.utc)
    actor = payload.get("employee_code") or payload.get("sub")
    actor_name = payload.get("name") or actor
    action = (body.action or "").lower().strip()

    if action == "pa-open":
        row.pa_open_at = now; row.pa_open_by = actor
    elif action == "pa-closed":
        row.pa_closed_at = now; row.pa_closed_by = actor
    elif action == "report-submit":
        row.report_submit_at = now; row.report_submit_by = actor
    elif action == "report-approve":
        row.report_approved_at = now; row.report_approved_by = actor
    elif action == "undo-pa-open":
        row.pa_open_at = None; row.pa_open_by = None
    elif action == "undo-pa-closed":
        row.pa_closed_at = None; row.pa_closed_by = None
    elif action == "undo-report-submit":
        row.report_submit_at = None; row.report_submit_by = None
    elif action == "undo-report-approve":
        row.report_approved_at = None; row.report_approved_by = None
    else:
        raise HTTPException(400, f"Unknown PAC action: {action}")

    _record_history(db, tracking_id=row.id, stage=row.current_stage, action=action,
                    actor_code=actor, actor_name=actor_name, notes=body.notes)
    await db.commit()
    await db.refresh(row)
    return _row_to_dict(row, now)


# ────────────────────────────────────────────────────────────────────────────
# Clock App ↔ Pre-Site integration — auto-promote hooks
# ────────────────────────────────────────────────────────────────────────────
async def _match_pos_for_session(db: AsyncSession, session) -> list[ProjectPO]:
    """Shared SSV/PAC matching used by both auto_link and auto_promote.

      SSV: clock-in by DU code → match ProjectPO.du_id exact, fallback cluster_site prefix
      PAC: clock-in by Cluster Name → match cluster_site prefix + B_Cluster/SSOA items
    """
    raw_site = (session.site_code or "").strip()
    if not raw_site:
        return []
    site_upper = raw_site.upper()
    site_prefix = site_upper.split("_")[0]  # legacy cluster names like "KKIWM_S02..." → "KKIWM"

    # SSV first — du_id exact, or cluster_site prefix (HW data quirk: du_id can be long number)
    pos_rows = (await db.execute(
        select(ProjectPO).where(
            ProjectPO.planned_dte_codes == session.employee_code,
            ProjectPO.work_type == "SSV",
            or_(
                func.upper(ProjectPO.du_id) == site_upper,
                func.upper(func.split_part(ProjectPO.cluster_site, "_", 1)) == site_prefix,
            ),
        )
    )).scalars().all()
    if pos_rows:
        return pos_rows

    # Fallback PAC — cluster_site prefix + B_Cluster/SSOA item_dis filter
    return (await db.execute(
        select(ProjectPO).where(
            ProjectPO.planned_dte_codes == session.employee_code,
            ProjectPO.work_type == "PAC",
            func.upper(func.split_part(ProjectPO.cluster_site, "_", 1)) == site_prefix,
            ProjectPO.item_dis.op("~*")(r"^[AB]_(Cluster|SSOA) "),
        )
    )).scalars().all()


async def auto_link_tracking_on_clockin(db: AsyncSession, session) -> int:
    """Called from clock.py when DTE starts PER_SITE session.

    Clock-in semantics:
      SSV: session.site_code = DU ID (e.g. "CBR0249") → match by ProjectPO.du_id
      PAC: session.site_code = Cluster Name (e.g. "EAS-FLASH-POC") → match by cluster_site
           1 clock-in covers ALL DUs in the cluster (1 cluster tracking row, 1 session)
    """
    if not session or not session.site_code or session.clock_type != "PER_SITE":
        return 0
    if not session.clock_in_at:
        return 0

    pos_rows = await _match_pos_for_session(db, session)
    if not pos_rows:
        return 0

    linked = 0
    # Deduplicate: for PAC we want at most 1 tracking per cluster regardless of how many POs matched
    seen_tracking_ids: set[int] = set()

    for po in pos_rows:
        tracking = await seed_tracking_from_po(db, po)
        if not tracking:
            continue
        if tracking.id in seen_tracking_ids:
            continue
        seen_tracking_ids.add(tracking.id)

        if (tracking.work_type or "").upper() == "PAC":
            # PAC: 1 clock-in = 1 session for the whole cluster (covers all DUs)
            existing = (await db.execute(
                select(DtePresiteSession).where(DtePresiteSession.tracking_id == tracking.id)
                .order_by(DtePresiteSession.round_number)
            )).scalars().all()
            pending = sorted([s for s in existing if s.status == "PENDING"], key=lambda s: s.round_number)
            if pending:
                sess = pending[0]
                next_round = sess.round_number
            else:
                next_round = max([s.round_number for s in existing], default=0) + 1
                sess = DtePresiteSession(tracking_id=tracking.id, round_number=next_round)
                db.add(sess)
                tracking.total_rounds = max(tracking.total_rounds or 0, next_round)
            sess.started_at = session.clock_in_at
            sess.started_by = session.employee_code
            sess.status = "IN_PROGRESS"
            if not tracking.dt_started_at:
                tracking.dt_started_at = session.clock_in_at
                tracking.dt_started_by = session.employee_code
            if tracking.current_stage == STAGE_FULL_ONAIR:
                tracking.current_stage = STAGE_DT_STARTED
            _record_history(
                db, tracking_id=tracking.id, stage=STAGE_DT_STARTED, action=f"pac-round-{next_round}-start",
                actor_code=session.employee_code, actor_name=session.employee_code,
                notes=f"PAC cluster '{tracking.cluster_key}' round {next_round} started from clock_sessions #{session.id}",
            )
            linked += 1
        else:
            # SSV: also use sessions table. Each clock-in cycle = one round (re-test after FAIL = new round)
            existing = (await db.execute(
                select(DtePresiteSession).where(DtePresiteSession.tracking_id == tracking.id)
                .order_by(DtePresiteSession.round_number)
            )).scalars().all()
            # Find pending session (e.g. round 1 after FAIL) or create new round
            pending = sorted([s for s in existing if s.status == "PENDING"], key=lambda s: s.round_number)
            if pending:
                sess = pending[0]
                next_round = sess.round_number
            else:
                # All sessions DONE or no sessions yet — check if need new round
                last = existing[-1] if existing else None
                if last and last.check_result == "FAIL":
                    # Re-plan after FAIL → new round
                    next_round = last.round_number + 1
                    sess = DtePresiteSession(tracking_id=tracking.id, round_number=next_round)
                    db.add(sess)
                    tracking.total_rounds = max(tracking.total_rounds or 0, next_round)
                elif not existing:
                    # First time
                    next_round = 1
                    sess = DtePresiteSession(tracking_id=tracking.id, round_number=next_round)
                    db.add(sess)
                    tracking.total_rounds = max(tracking.total_rounds or 0, 1)
                else:
                    # Already DONE/PASS — don't create more
                    continue

            sess.started_at = session.clock_in_at
            sess.started_by = session.employee_code
            sess.status = "IN_PROGRESS"
            # Mirror to tracking (legacy fields kept for compatibility)
            if not tracking.dt_started_at:
                tracking.dt_started_at = session.clock_in_at
                tracking.dt_started_by = session.employee_code
            if tracking.current_stage == STAGE_FULL_ONAIR:
                tracking.current_stage = STAGE_DT_STARTED
            _record_history(
                db, tracking_id=tracking.id, stage=STAGE_DT_STARTED, action=f"ssv-round-{next_round}-start",
                actor_code=session.employee_code, actor_name=session.employee_code,
                notes=f"SSV round {next_round} started from clock_sessions #{session.id}",
            )
            linked += 1
    return linked


async def auto_promote_to_presite(db: AsyncSession, session, outcome: str) -> int:
    """Called from clock.py when DTE clocks out PER_SITE.

    If outcome=COMPLETE → promote tracking to DT_DONE.
    If outcome=STOP/ISSUE → leave tracking alone; email is sent separately.
    Returns number of tracking rows promoted.
    """
    if not session or not session.site_code or session.clock_type != "PER_SITE":
        return 0
    if not session.clock_out_at:
        return 0
    outcome = (outcome or "").upper()
    if outcome != "COMPLETE":
        return 0  # only COMPLETE promotes

    pos_rows = await _match_pos_for_session(db, session)
    if not pos_rows:
        return 0

    promoted = 0
    seen_tracking_ids: set[int] = set()
    for po in pos_rows:
        # Look up tracking by appropriate identity (PAC uses cluster_key, SSV uses po_id)
        if (po.work_type or "").upper() == "PAC":
            tracking = (await db.execute(
                select(DtePresiteTracking).where(
                    and_(DtePresiteTracking.work_type == "PAC",
                         DtePresiteTracking.cluster_key == po.cluster_site)
                )
            )).scalar_one_or_none()
        else:
            tracking = (await db.execute(
                select(DtePresiteTracking).where(DtePresiteTracking.po_id == po.id)
            )).scalar_one_or_none()
        if not tracking:
            tracking = await seed_tracking_from_po(db, po)
        if not tracking:
            continue
        if tracking.id in seen_tracking_ids:
            continue
        seen_tracking_ids.add(tracking.id)

        if (tracking.work_type or "").upper() == "PAC":
            # PAC: end the latest IN_PROGRESS session
            sess = (await db.execute(
                select(DtePresiteSession)
                .where(and_(DtePresiteSession.tracking_id == tracking.id, DtePresiteSession.status == "IN_PROGRESS"))
                .order_by(DtePresiteSession.round_number.desc())
                .limit(1)
            )).scalar_one_or_none()
            if not sess:
                continue
            sess.ended_at = session.clock_out_at
            sess.ended_by = session.employee_code
            sess.status = "DONE"
            # If all planned rounds done → mark tracking as DT_DONE
            done_count = (await db.execute(
                select(func.count()).select_from(DtePresiteSession)
                .where(and_(DtePresiteSession.tracking_id == tracking.id, DtePresiteSession.status == "DONE"))
            )).scalar_one()
            total_count = (await db.execute(
                select(func.count()).select_from(DtePresiteSession)
                .where(DtePresiteSession.tracking_id == tracking.id)
            )).scalar_one() or 1
            if done_count >= total_count and tracking.current_stage in (STAGE_FULL_ONAIR, STAGE_DT_STARTED):
                tracking.current_stage = STAGE_DT_DONE
                tracking.dt_done_at = session.clock_out_at
                tracking.dt_done_by = session.employee_code
            _record_history(
                db, tracking_id=tracking.id, stage=tracking.current_stage,
                action=f"pac-round-{sess.round_number}-end",
                actor_code=session.employee_code, actor_name=session.employee_code,
                notes=f"PAC round {sess.round_number} ended (done_rounds={done_count}/{total_count}) from clock_sessions #{session.id}",
            )
            promoted += 1
        else:
            # SSV: write to current session + mirror to tracking
            # Find latest IN_PROGRESS session (also gates rework cycle: only end if a round is in-flight)
            sess = (await db.execute(
                select(DtePresiteSession)
                .where(and_(DtePresiteSession.tracking_id == tracking.id, DtePresiteSession.status == "IN_PROGRESS"))
                .order_by(DtePresiteSession.round_number.desc())
                .limit(1)
            )).scalar_one_or_none()
            # Allow rework: if no IN_PROGRESS session, skip silently (preserves prior REPORT_DONE/CHECKING/APPROVED state)
            if not sess and tracking.current_stage in (STAGE_REPORT_DONE, STAGE_CHECKING, STAGE_ACE_APPROVED):
                continue
            if sess:
                sess.ended_at = session.clock_out_at
                sess.ended_by = session.employee_code
                # Don't mark DONE yet — wait for check_result. Status remains IN_PROGRESS or set to AWAITING_CHECK
                sess.status = "IN_PROGRESS"  # still pending check by TL
            if not tracking.dt_started_at:
                tracking.dt_started_at = session.clock_in_at or session.clock_out_at
                tracking.dt_started_by = session.employee_code
            tracking.dt_done_at = session.clock_out_at
            tracking.dt_done_by = session.employee_code
            tracking.current_stage = STAGE_DT_DONE
            _record_history(
                db, tracking_id=tracking.id, stage=STAGE_DT_DONE,
                action=f"ssv-round-{sess.round_number if sess else '?'}-end",
                actor_code=session.employee_code, actor_name=session.employee_code,
                notes=f"SSV DT done from clock_sessions #{session.id}",
            )
            promoted += 1
    return promoted


# ────────────────────────────────────────────────────────────────────────────
# DTE Report Upload (.rar) — added 2026-05-29
# DTE Per-Site uploads a .rar matching the site they tested. PM downloads from
# the SSV/PAC Pre-Site "Report" column.
# ────────────────────────────────────────────────────────────────────────────
REPORTS_DIR = os.getenv("REPORTS_DIR", "/app/reports")
MAX_REPORT_BYTES = 500 * 1024 * 1024  # 500 MB
ALLOWED_REPORT_EXT = ".rar"


def _safe_filename(name: str) -> str:
    """Strip path + keep only safe chars. Preserve extension."""
    base = os.path.basename(name or "report.rar")
    base = _re_path.sub(r"[^A-Za-z0-9._-]", "_", base)
    return base[:200] or "report.rar"


@router.get("/my-sites")
async def my_sites(
    payload: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Return ONLY the sites assigned to the logged-in DTE (by employee_code),
    that are ready for report upload (DT_DONE or later). Used by /project/report-upload.
    """
    employee_code = payload.get("employee_code") or payload.get("sub")
    if not employee_code:
        raise HTTPException(401, "No employee code in token")

    # Match tracking rows where this DTE is assigned. ilike to be tolerant of
    # comma-joined codes, but exact match is the common case.
    rows = (await db.execute(
        select(DtePresiteTracking).where(
            or_(
                DtePresiteTracking.assigned_dte_code == employee_code,
                DtePresiteTracking.assigned_dte_code.ilike(f"%{employee_code}%"),
            )
        ).order_by(DtePresiteTracking.full_onair_at.desc().nullslast(), DtePresiteTracking.id.desc())
    )).scalars().all()

    now = datetime.now(timezone.utc)
    done_stages = {STAGE_DT_DONE, STAGE_REPORT_DONE, STAGE_CHECKING,
                   STAGE_ACE_SUBMITTED, STAGE_TL_REVIEWED, STAGE_PM_REVIEWED, STAGE_ACE_APPROVED}

    # ── DTE income per site (rate table, NOT PO price) ─────────────────────
    # PAC rate is per-site; a cluster may contain N child sites → count them.
    from app.services.dte_rates import compute_income
    # PAC tracking has NULL item_dis (varies per child PO) → pull count + a
    # representative item_dis from the child POs to detect SSOA vs Cluster.
    cluster_keys = [r.cluster_key for r in rows if r.work_type == "PAC" and r.cluster_key]
    cluster_site_count: dict[str, int] = {}
    cluster_item_dis: dict[str, str] = {}
    if cluster_keys:
        crows = (await db.execute(
            select(ProjectPO.cluster_site, func.count(), func.min(ProjectPO.item_dis))
            .where(ProjectPO.cluster_site.in_(cluster_keys))
            .group_by(ProjectPO.cluster_site)
        )).all()
        for cs, cnt, idis in crows:
            cluster_site_count[cs] = int(cnt)
            if idis:
                cluster_item_dis[cs] = idis

    # Resolve approver codes → names (for the Approved column)
    from app.models.auth_user import AuthUser
    approver_codes = set()
    for r in rows:
        for c in (r.ace_approve_by, r.check_by):
            if c:
                approver_codes.add(c)
    name_by_code: dict[str, str] = {}
    if approver_codes:
        urows = (await db.execute(
            select(AuthUser.employee_code, AuthUser.first_name, AuthUser.last_name)
            .where(AuthUser.employee_code.in_(list(approver_codes)))
        )).all()
        for code, fn, ln in urows:
            name_by_code[code] = f"{fn or ''} {ln or ''}".strip() or code

    data = []
    for r in rows:
        d = _row_to_dict(r, now)
        d["dt_done"] = r.current_stage in done_stages
        # Approval info: prefer ace_approve_*, then check_* (PASS), then completed_at
        appr_at = r.ace_approve_at or (r.check_at if r.check_result == "PASS" else None) or r.completed_at
        appr_by = r.ace_approve_by or r.check_by
        d["approved_at"] = _iso(appr_at) if r.current_stage == STAGE_ACE_APPROVED or appr_at else None
        d["approved_by"] = appr_by
        d["approved_by_name"] = name_by_code.get(appr_by, appr_by) if appr_by else None
        if r.work_type == "PAC":
            site_count = cluster_site_count.get(r.cluster_key, 1)
            item_dis = r.item_dis or cluster_item_dis.get(r.cluster_key)
        else:
            site_count = 1
            item_dis = r.item_dis
        inc = compute_income(r.work_type, item_dis, r.layers, site_count)
        d["income"] = inc                 # {dt, report, total, unit_total, category, site_count}
        d["amount"] = inc["total"]        # convenience: DTE income for this site
        basis = r.dt_done_at or r.full_onair_at
        d["revenue_month"] = basis.strftime("%Y-%m") if basis else None
        d["revenue_week"] = basis.strftime("%G-W%V") if basis else None  # ISO week e.g. 2026-W22
        d["revenue_date"] = basis.date().isoformat() if basis else None
        data.append(d)

    # Rollup helper (by month or week key)
    def _rollup(key_field: str, label_field: str):
        buckets: dict[str, dict] = {}
        for d in data:
            k = d.get(key_field)
            if not k:
                continue
            b = buckets.setdefault(k, {label_field: k, "sites": 0, "revenue": 0.0,
                                       "dt": 0.0, "report": 0.0, "approved": 0, "uploaded": 0,
                                       "paid": 0.0, "unpaid": 0.0})
            tot = d["income"]["total"]
            b["sites"] += 1
            b["revenue"] += tot
            b["dt"]      += d["income"]["dt"]
            b["report"]  += d["income"]["report"]
            if d.get("paid"):
                b["paid"] += tot
            else:
                b["unpaid"] += tot
            if d.get("current_stage") == STAGE_ACE_APPROVED:
                b["approved"] += 1
            if d.get("has_report"):
                b["uploaded"] += 1
        return sorted(buckets.values(), key=lambda x: x[label_field], reverse=True)

    monthly = _rollup("revenue_month", "month")
    weekly = _rollup("revenue_week", "week")

    return {
        "data": data,
        "total": len(data),
        "employee_code": employee_code,
        "monthly": monthly,
        "weekly": weekly,
        "currency": "THB",
    }


async def _compute_income_for_tracking(db: AsyncSession, r: DtePresiteTracking) -> dict:
    """Compute the live DTE income for a single tracking row.

    Mirrors the PAC site-count / representative item_dis resolution used by the
    bulk finance builder, so mark-paid freezes the exact same number Finance saw.
    """
    from app.services.dte_rates import compute_income
    if (r.work_type or "").upper() == "PAC" and r.cluster_key:
        crow = (await db.execute(
            select(func.count(), func.min(ProjectPO.item_dis))
            .where(ProjectPO.cluster_site == r.cluster_key)
        )).one()
        sc = int(crow[0]) or 1
        idis = r.item_dis or crow[1]
    else:
        sc = 1
        idis = r.item_dis
    return compute_income(r.work_type, idis, r.layers, sc)


async def _finance_payments_data(db: AsyncSession, status: str, month: str | None,
                                 dte_code: str | None, work_type: str | None) -> dict:
    """Shared builder for finance payment data (used by JSON view + Excel export)."""
    from app.services.dte_rates import compute_income
    from app.models.auth_user import AuthUser

    stmt = select(DtePresiteTracking)
    if work_type:
        stmt = stmt.where(DtePresiteTracking.work_type == work_type.upper())
    rows = (await db.execute(
        stmt.order_by(DtePresiteTracking.dt_done_at.desc().nullslast(), DtePresiteTracking.id.desc())
    )).scalars().all()

    # PAC cluster site-count + representative item_dis
    cluster_keys = [r.cluster_key for r in rows if r.work_type == "PAC" and r.cluster_key]
    cluster_site_count: dict[str, int] = {}
    cluster_item_dis: dict[str, str] = {}
    if cluster_keys:
        crows = (await db.execute(
            select(ProjectPO.cluster_site, func.count(), func.min(ProjectPO.item_dis))
            .where(ProjectPO.cluster_site.in_(cluster_keys)).group_by(ProjectPO.cluster_site)
        )).all()
        for cs, cnt, idis in crows:
            cluster_site_count[cs] = int(cnt)
            if idis:
                cluster_item_dis[cs] = idis

    # Resolve approver + DTE names
    codes = set()
    for r in rows:
        for c in (r.assigned_dte_code, r.ace_approve_by, r.check_by):
            if c:
                codes.add(c)
    name_by_code: dict[str, str] = {}
    if codes:
        urows = (await db.execute(
            select(AuthUser.employee_code, AuthUser.first_name, AuthUser.last_name)
            .where(AuthUser.employee_code.in_(list(codes)))
        )).all()
        for code, fn, ln in urows:
            name_by_code[code] = f"{fn or ''} {ln or ''}".strip() or code

    data = []
    for r in rows:
        # Filters
        if dte_code and (r.assigned_dte_code or "").upper() != dte_code.upper():
            continue
        paid = bool(r.dte_paid_at)
        if status == "paid" and not paid:
            continue
        if status == "unpaid" and paid:
            continue
        basis = r.dt_done_at or r.full_onair_at
        ym = basis.strftime("%Y-%m") if basis else None
        if month and ym != month:
            continue
        if r.work_type == "PAC":
            sc = cluster_site_count.get(r.cluster_key, 1)
            idis = r.item_dis or cluster_item_dis.get(r.cluster_key)
        else:
            sc = 1
            idis = r.item_dis
        inc = compute_income(r.work_type, idis, r.layers, sc)
        # Paid rows show the frozen snapshot captured at mark-paid time, not the
        # live recomputation — so a later rate/cluster/item_dis change can't drift
        # the reported "paid" total. Falls back to live for legacy rows paid
        # before the snapshot columns existed.
        if paid and r.dte_paid_amount is not None:
            disp_total = float(r.dte_paid_amount)
            disp_dt = float(r.dte_paid_income_dt) if r.dte_paid_income_dt is not None else inc["dt"]
            disp_report = float(r.dte_paid_income_report) if r.dte_paid_income_report is not None else inc["report"]
            disp_category = r.dte_paid_category or inc["category"]
            disp_site_count = r.dte_paid_site_count or inc["site_count"]
        else:
            disp_total = inc["total"]
            disp_dt = inc["dt"]
            disp_report = inc["report"]
            disp_category = inc["category"]
            disp_site_count = inc["site_count"]
        appr_at = r.ace_approve_at or (r.check_at if r.check_result == "PASS" else None) or r.completed_at
        appr_by = r.ace_approve_by or r.check_by
        data.append({
            "tracking_id": r.id,
            "site_code": r.cluster_key or r.site_code,
            "work_type": r.work_type,
            "dte_code": r.assigned_dte_code,
            "dte_name": r.assigned_dte_name or name_by_code.get(r.assigned_dte_code, r.assigned_dte_code),
            "category": disp_category,
            "site_count": disp_site_count,
            "income": disp_total,
            "income_dt": disp_dt,
            "income_report": disp_report,
            "current_stage": r.current_stage,
            "dt_done_date": (r.dt_done_at.date().isoformat() if r.dt_done_at else None),
            "month": ym,
            "approved": r.current_stage == STAGE_ACE_APPROVED,
            "approved_at": _iso(appr_at) if appr_at else None,
            "approved_by_name": name_by_code.get(appr_by, appr_by) if appr_by else None,
            "has_report": bool(r.report_file_path),
            "report_version": r.report_version or 0,
            "paid": paid,
            "dte_paid_at": _iso(r.dte_paid_at),
            "dte_paid_by": r.dte_paid_by,
            "dte_payment_ref": r.dte_payment_ref,
            # payable = approved & not paid & (report uploaded when a Report Prep
            # component is owed). PAC (Cluster/SSOA) earns DT only — no report
            # component, log-file checked by DTA — so it needs no .rar to pay.
            "payable": (
                r.current_stage == STAGE_ACE_APPROVED
                and not paid
                and (bool(r.report_file_path) or (inc["report"] or 0) <= 0)
            ),
        })

    # By-DTE summary
    by_dte: dict[str, dict] = {}
    for d in data:
        code = d["dte_code"] or "—"
        b = by_dte.setdefault(code, {
            "dte_code": code, "dte_name": d["dte_name"], "sites": 0,
            "total": 0.0, "paid": 0.0, "unpaid": 0.0, "payable": 0.0,
        })
        b["sites"] += 1
        b["total"] += d["income"]
        if d["paid"]:
            b["paid"] += d["income"]
        else:
            b["unpaid"] += d["income"]
        if d["payable"]:
            b["payable"] += d["income"]
    by_dte_list = sorted(by_dte.values(), key=lambda x: x["unpaid"], reverse=True)

    # Monthly rollup (paid vs unpaid across all DTEs) — for charts
    months: dict[str, dict] = {}
    for d in data:
        m = d.get("month")
        if not m:
            continue
        b = months.setdefault(m, {"month": m, "sites": 0, "total": 0.0, "paid": 0.0, "unpaid": 0.0})
        b["sites"] += 1
        b["total"] += d["income"]
        if d["paid"]:
            b["paid"] += d["income"]
        else:
            b["unpaid"] += d["income"]
    monthly = sorted(months.values(), key=lambda x: x["month"])

    totals = {
        "sites": len(data),
        "dtes": len(by_dte),
        "total": sum(d["income"] for d in data),
        "paid": sum(d["income"] for d in data if d["paid"]),
        "unpaid": sum(d["income"] for d in data if not d["paid"]),
        "payable": sum(d["income"] for d in data if d["payable"]),
    }
    return {"data": data, "by_dte": by_dte_list, "monthly": monthly, "totals": totals, "currency": "THB"}


@router.get("/finance/payments")
async def finance_payments(
    status: str = Query("all"),
    month: str | None = Query(None),
    dte_code: str | None = Query(None),
    work_type: str | None = Query(None),
    payload: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Finance view of ALL DTE payments — who to pay, how much, with details."""
    if payload.get("role") not in ROLE_FINANCE:
        raise HTTPException(403, "Only Finance/PM/Admin may view payments")
    return await _finance_payments_data(db, status, month, dte_code, work_type)


@router.get("/finance/payments/export")
async def finance_payments_export(
    status: str = Query("all"),
    month: str | None = Query(None),
    dte_code: str | None = Query(None),
    work_type: str | None = Query(None),
    payload: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export DTE payments as an Excel workbook (Summary + one sheet per DTE).
    Serves as payment evidence. Role-gated to Finance/PM/Admin.
    """
    if payload.get("role") not in ROLE_FINANCE:
        raise HTTPException(403, "Only Finance/PM/Admin may export payments")
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    res = await _finance_payments_data(db, status, month, dte_code, work_type)
    rows, by_dte, totals = res["data"], res["by_dte"], res["totals"]

    wb = openpyxl.Workbook()
    hdr_fill = PatternFill("solid", fgColor="1E293B")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    money = '#,##0'
    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    # ── Summary sheet ──
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "DTE Payment Evidence — Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Filter: status={status} · month={month or 'all'} · dte={dte_code or 'all'} · type={work_type or 'all'}"
    ws["A2"].font = Font(italic=True, size=9, color="64748B")
    head = ["DTE Code", "DTE Name", "Sites", "Total (THB)", "Paid (THB)", "Unpaid (THB)", "Ready to Pay (THB)"]
    ws.append([])
    ws.append(head)
    style_header(ws, 4, len(head))
    for b in by_dte:
        ws.append([b["dte_code"], b["dte_name"], b["sites"], b["total"], b["paid"], b["unpaid"], b["payable"]])
    ws.append(["", "TOTAL", totals["sites"], totals["total"], totals["paid"], totals["unpaid"], totals["payable"]])
    last = ws.max_row
    for r in range(5, last + 1):
        for c in range(4, 8):
            ws.cell(row=r, column=c).number_format = money
    for c in (2, 3, 4, 5, 6, 7):
        ws.cell(row=last, column=c).font = Font(bold=True)
    widths = [14, 26, 8, 14, 14, 14, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # ── One sheet per DTE ──
    by_code: dict[str, list] = {}
    for d in rows:
        by_code.setdefault(d["dte_code"] or "—", []).append(d)
    det_head = ["Site", "Type", "Rate Category", "DT Date", "Approved", "Approved By",
                "Report", "Income (THB)", "Payment Status", "Paid At", "Payment Ref"]
    for code, items in by_code.items():
        name = items[0]["dte_name"] or code
        safe = (str(code)[:28] or "DTE").replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "").replace(":", "").replace("[", "").replace("]", "")
        ws2 = wb.create_sheet(title=safe[:31])
        ws2["A1"] = f"Payment Voucher — {name} ({code})"
        ws2["A1"].font = Font(bold=True, size=13)
        ws2.append([])
        ws2.append(det_head)
        style_header(ws2, 3, len(det_head))
        tot = paid = 0.0
        for d in items:
            ws2.append([
                d["site_code"], d["work_type"], d["category"], d["dt_done_date"] or "",
                (d["approved_at"] or "")[:10], d["approved_by_name"] or "",
                f"v{d['report_version']}" if d["has_report"] else "—",
                d["income"],
                "PAID" if d["paid"] else ("READY" if d["payable"] else "UNPAID"),
                (d["dte_paid_at"] or "")[:10], d["dte_payment_ref"] or "",
            ])
            tot += d["income"]
            if d["paid"]:
                paid += d["income"]
        ws2.append(["", "", "", "", "", "", "TOTAL", tot, "", "", ""])
        ws2.append(["", "", "", "", "", "", "PAID", paid, "", "", ""])
        ws2.append(["", "", "", "", "", "", "UNPAID", tot - paid, "", "", ""])
        for r in range(4, ws2.max_row + 1):
            ws2.cell(row=r, column=8).number_format = money
        for i, w in enumerate([16, 7, 22, 12, 12, 16, 8, 14, 14, 12, 16], 1):
            ws2.column_dimensions[chr(64 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"DTE_Payments_{month or 'all'}_{status}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/tracking/{tracking_id}/upload-report")
async def upload_report(
    tracking_id: int,
    file: UploadFile = File(...),
    payload: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """DTE uploads a .rar report for a site they tested.
    - Only the assigned DTE (or PM/Admin) may upload.
    - Accepts .rar only, max 500 MB.
    - Saves to /app/reports/{tracking_id}/{ts}_{filename}, bumps version,
      and advances DT_DONE → REPORT_DONE.
    """
    row = (await db.execute(
        select(DtePresiteTracking).where(DtePresiteTracking.id == tracking_id)
    )).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Tracking row not found")

    employee_code = payload.get("employee_code") or payload.get("sub")
    role = payload.get("role")
    is_owner = bool(employee_code and row.assigned_dte_code and
                    employee_code.upper() in str(row.assigned_dte_code).upper())
    if not (is_owner or role in ROLE_TL):
        raise HTTPException(403, "You may only upload reports for sites assigned to you")

    fname = _safe_filename(file.filename or "report.rar")
    if not fname.lower().endswith(ALLOWED_REPORT_EXT):
        raise HTTPException(400, f"Only {ALLOWED_REPORT_EXT} files are allowed (got '{file.filename}')")

    if not ftp_service.is_configured():
        raise HTTPException(503, "FTP storage is not configured on the server")

    # Stream upload to a temp file first (size guard), then push to FTP.
    import tempfile
    fd, tmp_path = tempfile.mkstemp(suffix=".rar")
    os.close(fd)
    total = 0
    try:
        with open(tmp_path, "wb") as out:
            while True:
                chunk = await file.read(1024 * 1024)  # 1 MB
                if not chunk:
                    break
                total += len(chunk)
                if total > MAX_REPORT_BYTES:
                    raise HTTPException(413, f"File exceeds max size of {MAX_REPORT_BYTES // (1024*1024)} MB")
                out.write(chunk)

        # Folder by DT-done date (fallback: today) — matches team's day-folder convention
        dt_date = (row.dt_done_at or datetime.now(timezone.utc)).date()
        try:
            remote_path = await asyncio.to_thread(
                ftp_service.upload_report_sync, tmp_path, dt_date, fname
            )
        except Exception as exc:
            raise HTTPException(502, f"FTP upload failed: {exc}")
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass

    now = datetime.now(timezone.utc)
    row.report_file_path = remote_path          # FTP path (not local)
    row.report_filename = fname
    row.report_file_size = total
    row.report_uploaded_at = now
    row.report_uploaded_by = employee_code
    row.report_version = (row.report_version or 0) + 1

    # Advance stage DT_DONE → REPORT_DONE (only forward; don't rewind later stages)
    if row.current_stage == STAGE_DT_DONE:
        row.current_stage = STAGE_REPORT_DONE
        row.report_done_at = now
        row.report_done_by = employee_code

    _record_history(
        db, tracking_id=tracking_id, stage=row.current_stage, action="upload-report",
        actor_code=employee_code, actor_name=payload.get("name") or employee_code,
        notes=f"Uploaded {fname} (v{row.report_version}, {total} bytes) → FTP {remote_path}",
    )
    await db.commit()
    return {
        "ok": True,
        "tracking_id": tracking_id,
        "filename": fname,
        "size": total,
        "version": row.report_version,
        "current_stage": row.current_stage,
        "remote_path": remote_path,
    }


@router.get("/tracking/{tracking_id}/download-report")
async def download_report(
    tracking_id: int,
    background_tasks: BackgroundTasks,
    payload: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Download the uploaded .rar report (pulled from FTP). Any viewer role may download."""
    if payload.get("role") not in ROLE_VIEW:
        raise HTTPException(403, "Not allowed")
    row = (await db.execute(
        select(DtePresiteTracking).where(DtePresiteTracking.id == tracking_id)
    )).scalar_one_or_none()
    if not row or not row.report_file_path:
        raise HTTPException(404, "No report uploaded for this site")

    remote = row.report_file_path
    # Legacy: a few rows may still hold a local path (pre-FTP). Serve those directly.
    if os.path.exists(remote):
        return FileResponse(
            remote,
            media_type="application/x-rar-compressed",
            filename=row.report_filename or f"report_{tracking_id}.rar",
        )

    if not ftp_service.is_configured():
        raise HTTPException(503, "FTP storage is not configured on the server")
    try:
        tmp_path = await asyncio.to_thread(ftp_service.download_report_sync, remote)
    except Exception as exc:
        raise HTTPException(502, f"FTP download failed: {exc}")
    # Clean up temp file after the response is sent
    background_tasks.add_task(lambda p: os.path.exists(p) and os.remove(p), tmp_path)
    return FileResponse(
        tmp_path,
        media_type="application/x-rar-compressed",
        filename=row.report_filename or f"report_{tracking_id}.rar",
    )


# ────────────────────────────────────────────────────────────────────────────
# DTE payment status — Finance/PM marks a site's DTE income as paid
# ────────────────────────────────────────────────────────────────────────────
ROLE_FINANCE = {"ACCOUNTING", "PM", "PROJECT_ADMIN", "SUPER_ADMIN", "SYSTEM_ADMIN", "HR_ADMIN"}


class MarkPaidIn(BaseModel):
    payment_ref: Optional[str] = None


@router.post("/tracking/{tracking_id}/mark-paid")
async def mark_paid(
    tracking_id: int,
    body: MarkPaidIn,
    payload: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Mark a site's DTE income as paid (Finance/PM/Admin only)."""
    if payload.get("role") not in ROLE_FINANCE:
        raise HTTPException(403, "Only Finance/PM/Admin may mark payments")
    row = (await db.execute(
        select(DtePresiteTracking).where(DtePresiteTracking.id == tracking_id)
    )).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Tracking row not found")
    # Guard: only approved work is payable (cannot pay un-approved / unfinished sites)
    if row.current_stage != STAGE_ACE_APPROVED:
        raise HTTPException(
            409, f"Cannot mark paid: stage is {row.current_stage}, must be {STAGE_ACE_APPROVED}")
    # Guard: idempotency — never silently overwrite an existing payment
    if row.dte_paid_at:
        raise HTTPException(409, "Already marked paid; unmark first to re-pay")
    # Freeze the computed income onto the row so the paid total can't drift later
    inc = await _compute_income_for_tracking(db, row)
    # Guard: a report-bearing item (SSV/Pre-DT, report component > 0) must have its
    # report uploaded before payment. PAC earns DT only (log file checked by DTA)
    # so it has no report requirement.
    if (inc["report"] or 0) > 0 and not row.report_file_path:
        raise HTTPException(409, "Cannot mark paid: report not uploaded for a report-bearing item")
    now = datetime.now(timezone.utc)
    actor = payload.get("employee_code") or payload.get("sub")
    row.dte_paid_at = now
    row.dte_paid_by = actor
    row.dte_payment_ref = body.payment_ref
    row.dte_paid_amount = inc["total"]
    row.dte_paid_income_dt = inc["dt"]
    row.dte_paid_income_report = inc["report"]
    row.dte_paid_category = inc["category"]
    row.dte_paid_site_count = inc["site_count"]
    _record_history(db, tracking_id=tracking_id, stage=row.current_stage, action="mark-paid",
                    actor_code=actor, actor_name=payload.get("name") or actor,
                    notes=f"DTE payment marked paid · {inc['total']:.2f} THB ({inc['category']})"
                          f"{(' · ref ' + body.payment_ref) if body.payment_ref else ''}")
    await db.commit()
    return {"ok": True, "tracking_id": tracking_id, "paid": True, "dte_paid_at": _iso(now),
            "amount": inc["total"], "category": inc["category"]}


@router.post("/tracking/{tracking_id}/unmark-paid")
async def unmark_paid(
    tracking_id: int,
    payload: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Revert a paid mark (Finance/PM/Admin only)."""
    if payload.get("role") not in ROLE_FINANCE:
        raise HTTPException(403, "Only Finance/PM/Admin may change payments")
    row = (await db.execute(
        select(DtePresiteTracking).where(DtePresiteTracking.id == tracking_id)
    )).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Tracking row not found")
    row.dte_paid_at = None
    row.dte_paid_by = None
    row.dte_payment_ref = None
    row.dte_paid_amount = None
    row.dte_paid_income_dt = None
    row.dte_paid_income_report = None
    row.dte_paid_category = None
    row.dte_paid_site_count = None
    actor = payload.get("employee_code") or payload.get("sub")
    _record_history(db, tracking_id=tracking_id, stage=row.current_stage, action="unmark-paid",
                    actor_code=actor, actor_name=payload.get("name") or actor, notes="DTE payment reverted")
    await db.commit()
    return {"ok": True, "tracking_id": tracking_id, "paid": False}


# ────────────────────────────────────────────────────────────────────────────
# Cluster geo — real coordinates from project_sites (for DTA Map View)
# ────────────────────────────────────────────────────────────────────────────
@router.get("/cluster-geo")
async def cluster_geo(
    db: AsyncSession = Depends(get_db),
    payload: dict = Depends(get_current_user),
):
    """Per-cluster centroid coordinates derived from project_sites.lat/lng.

    Keyed by rf_cluster_name (e.g. 'EAS0120-Full-1') so the DTA Map View can
    plot real PAC/cluster locations instead of mock points. Only clusters that
    have at least one geocoded site are returned.
    """
    rows = (await db.execute(
        select(
            ProjectSite.rf_cluster_name,
            func.count().label("sites"),
            func.count(ProjectSite.lat).label("geocoded"),
            func.avg(ProjectSite.lat).label("clat"),
            func.avg(ProjectSite.lng).label("clng"),
            func.min(ProjectSite.lat).label("min_lat"),
            func.max(ProjectSite.lat).label("max_lat"),
            func.min(ProjectSite.lng).label("min_lng"),
            func.max(ProjectSite.lng).label("max_lng"),
        )
        .where(
            ProjectSite.rf_cluster_name.isnot(None),
            ProjectSite.lat.isnot(None),
            ProjectSite.lng.isnot(None),
        )
        .group_by(ProjectSite.rf_cluster_name)
    )).all()

    clusters = {}
    for r in rows:
        if not r.clat or not r.clng:
            continue
        clusters[r.rf_cluster_name] = {
            "cluster": r.rf_cluster_name,
            "lat": round(float(r.clat), 6),
            "lng": round(float(r.clng), 6),
            "sites": int(r.sites or 0),
            "geocoded_sites": int(r.geocoded or 0),
            "bbox": [
                round(float(r.min_lat), 6), round(float(r.min_lng), 6),
                round(float(r.max_lat), 6), round(float(r.max_lng), 6),
            ],
        }
    return {"count": len(clusters), "clusters": clusters}


@router.get("/cluster-geo/{cluster_name}/sites")
async def cluster_geo_sites(
    cluster_name: str,
    db: AsyncSession = Depends(get_db),
    payload: dict = Depends(get_current_user),
):
    """Individual geocoded sites for one cluster (for drill-down on the map)."""
    rows = (await db.execute(
        select(
            ProjectSite.site_code, ProjectSite.site_name,
            ProjectSite.lat, ProjectSite.lng,
            ProjectSite.province, ProjectSite.district,
        )
        .where(
            ProjectSite.rf_cluster_name == cluster_name,
            ProjectSite.lat.isnot(None),
            ProjectSite.lng.isnot(None),
        )
        .order_by(ProjectSite.site_code)
    )).all()
    return {
        "cluster": cluster_name,
        "sites": [
            {
                "site_code": r.site_code, "site_name": r.site_name,
                "lat": round(float(r.lat), 6), "lng": round(float(r.lng), 6),
                "province": r.province, "district": r.district,
            }
            for r in rows
        ],
    }
