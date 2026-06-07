"""
Bill Reader service — อ่านบิล (PDF สแกน) ด้วย Tesseract OCR แล้วแปลงเป็น Payment Voucher

รองรับหลายชนิดบิลผ่าน Bill Profile Registry (app/services/bill_profiles.py):
ค่าโทรศัพท์ / ค่าไฟ-น้ำ / Fleet card / ค่าเช่า / ผู้รับเหมา / อื่น ๆ

Pipeline: PDF -> render (pymupdf) -> OCR (tha+eng) -> classify ชนิดบิล
          -> parse ตาม profile -> คำนวณ WHT/Net ตาม profile -> กรอก PV.03/04

Stateless: ไม่แตะ DB
"""

from __future__ import annotations

import datetime
import io
import re
from dataclasses import dataclass, field
from pathlib import Path

import fitz  # PyMuPDF
import pytesseract
from PIL import Image
import openpyxl

from app.services import bill_profiles
from app.services.bill_profiles import BillProfile

VAT_RATE = 0.07
OCR_LANG = "tha+eng"
OCR_DPI = 300

ASSETS = Path(__file__).resolve().parent.parent / "assets"
PV_TEMPLATES = {
    "pv03": (ASSETS / "pv_template.xlsx", "Template PV.03 (Other Expense)"),
    "pv04": (ASSETS / "pv04_template.xlsx", "Template PV.04 (Subcon TF)"),
}

DEFAULT_HEADER = {
    "project": "NBTC2501 NBTC NSA/SA Benchmarking Project",
    "name": "Ms.Anong Jantaraeng",
    "issued": "Ms.Anong Jantaraeng",
}
THAI_MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

_NUM_RE = re.compile(r"\d{1,3}(?:,\d{3})*\.\d{2}")
_PHONE_RE = re.compile(r"0[689][\s\-]?\d{4}[\s\-]?\d{4}")
_PERIOD_RE = re.compile(
    r"(\d{1,2})/(\d{1,2})/(\d{4})\s*[-–—]\s*(\d{1,2})/(\d{1,2})/(\d{4})")
_MONTH_RE = re.compile(
    r"(January|February|March|April|May|June|July|August|September|October|November|December)"
    r"\s+(\d{4})", re.I)
_METER_RE = re.compile(r"\b\d{2}/\d{2,4}[-/]\w+\b")  # คร่าว ๆ สำหรับเลขผู้ใช้ไฟ


# ---------------------------------------------------------------------------
@dataclass
class BillLine:
    identifier: str = ""        # เบอร์ / meter / สัญญา / site
    period: str = ""
    amount: float = 0.0
    vat: float = 0.0
    vendor: str = ""
    source_page: int = 0
    needs_review: bool = False
    wht: float = 0.0
    net: float = 0.0

    def compute(self, profile: BillProfile) -> "BillLine":
        self.wht = round(self.amount * profile.wht_rate, 2)
        self.net = round(self.amount + self.vat - self.wht, 2)
        return self

    @property
    def id_fmt(self) -> str:
        d = re.sub(r"\D", "", self.identifier)
        if len(d) == 10 and d[0] == "0":
            return f"{d[0:2]}-{d[2:6]}-{d[6:]}"
        return self.identifier

    def brand(self) -> str:
        v = (self.vendor or "").lower()
        for b in ("AIS", "True", "DTAC", "NT", "TOT"):
            if b.lower() in v:
                return b
        if "awn" in v:
            return "AIS"
        return "Bill"

    def desc(self, profile: BillProfile) -> str:
        if profile.key == "telecom":
            return f"{self.brand()} no.{self.id_fmt} of period {self.period}"
        prefix = profile.desc_prefix or "Service"
        idpart = f" no.{self.id_fmt}" if self.identifier else ""
        period = f" of period {self.period}" if self.period else ""
        return f"{prefix}{idpart}{period}".strip()

    def to_dict(self, profile: BillProfile) -> dict:
        return {
            "identifier": self.identifier, "id_fmt": self.id_fmt, "period": self.period,
            "amount": self.amount, "vat": self.vat, "wht": self.wht, "net": self.net,
            "vendor": self.vendor, "desc": self.desc(profile),
            "source_page": self.source_page, "needs_review": self.needs_review,
        }


# ---------------------------------------------------------------------------
def _ocr_pages(pdf_bytes: bytes) -> list[str]:
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    texts = []
    for page in doc:
        pix = page.get_pixmap(dpi=OCR_DPI)
        img = Image.open(io.BytesIO(pix.tobytes("png")))
        texts.append(pytesseract.image_to_string(img, lang=OCR_LANG))
    doc.close()
    return texts


def _to_ad_year(y: int) -> int:
    return y - 543 if y > 2400 else y


def _parse_period(text: str) -> str:
    m = _PERIOD_RE.search(text)
    if m:
        d1, m1, y1, d2, m2, y2 = m.groups()
        return (f"{int(d1):02d}/{int(m1):02d}/{_to_ad_year(int(y1))}-"
                f"{int(d2):02d}/{int(m2):02d}/{_to_ad_year(int(y2))}")
    m = _MONTH_RE.search(text)
    if m:
        return f"{m.group(1).title()} {m.group(2)}"
    return ""


def _parse_phone(text: str) -> str:
    for m in _PHONE_RE.finditer(text):
        d = re.sub(r"\D", "", m.group())
        if len(d) == 10:
            return d
    return ""


def _near_in(value: float, pool: set[float], tol: float = 0.02) -> bool:
    return any(abs(value - p) <= tol for p in pool)


def _parse_amount_vat(text: str, extract_vat: bool) -> tuple[float, float, bool]:
    nums = sorted({float(n.replace(",", "")) for n in _NUM_RE.findall(text)})
    if not nums:
        return 0.0, 0.0, True
    if not extract_vat:
        return max(nums), 0.0, False  # ยอดรวมแล้ว ไม่มี VAT แยก
    present = set(nums)
    best = None
    for amount in nums:
        if amount <= 0:
            continue
        vat = round(amount * VAT_RATE, 2)
        total = round(amount + vat, 2)
        if _near_in(vat, present) and _near_in(total, present):
            if best is None or amount > best[0]:
                best = (amount, vat)
    if best:
        return best[0], best[1], False
    total = max(nums)
    amount = round(total / (1 + VAT_RATE), 2)
    return amount, round(total - amount, 2), True


def _detect_vendor(text: str, profile: BillProfile) -> str:
    # เฉพาะบิลโทรศัพท์เท่านั้นที่จับชื่อค่ายจากข้อความ — ชนิดอื่นใช้ค่าเริ่มต้นของ profile
    if profile.key != "telecom":
        return profile.default_vendor
    up = text.upper()
    if "AIS" in up or "AWN" in up:
        return "Advanced Wireless Network Co., Ltd. (AIS)"
    if "DTAC" in up:
        return "Total Access Communication PCL. (DTAC)"
    if "TRUE" in up:
        return "True Move H Universal Communication Co., Ltd."
    if "TOT" in up or re.search(r"\bNT\b", up):
        return "National Telecom PCL. (NT)"
    return profile.default_vendor


def _parse_identifier(text: str, profile: BillProfile) -> str:
    if profile.id_kind == "phone":
        return _parse_phone(text)
    # meter / contract / site -> ยังอ่านอัตโนมัติไม่แม่น (สแกนหลากหลาย) ให้ผู้ใช้กรอกในตาราง
    return ""


def parse_page(text: str, page_no: int, profile: BillProfile) -> BillLine:
    amount, vat, review = _parse_amount_vat(text, profile.extract_vat)
    line = BillLine(
        identifier=_parse_identifier(text, profile),
        period=_parse_period(text),
        amount=amount, vat=vat,
        vendor=_detect_vendor(text, profile),
        source_page=page_no,
        needs_review=review or amount <= 0,
    )
    # telecom ต้องมีเบอร์+รอบบิล จึงจะมั่นใจ
    if profile.id_kind == "phone" and (not line.identifier or not line.period):
        line.needs_review = True
    return line.compute(profile)


# ---------------------------------------------------------------------------
def extract_bill(pdf_bytes: bytes, filename: str = "", bill_type: str | None = None) -> dict:
    """อ่านบิล -> {bill_type, lines, header, vendor, ...}"""
    texts = _ocr_pages(pdf_bytes)
    full = "\n".join(texts)
    profile = bill_profiles.get_profile(bill_type) if bill_type else bill_profiles.classify(full)

    lines = [parse_page(t, i + 1, profile) for i, t in enumerate(texts)]
    lines = [l for l in lines if l.amount > 0 or l.identifier or l.period]
    vendor = next((l.vendor for l in lines if l.vendor), profile.default_vendor)

    return {
        "bill_type": profile.key,
        "bill_type_label": profile.label,
        "wht_rate": profile.wht_rate,
        "extract_vat": profile.extract_vat,
        "pv_template": profile.pv_template,
        "id_kind": profile.id_kind,
        "vendor": vendor,
        "lines": [l.to_dict(profile) for l in lines],
        "header": build_header(filename),
        "profiles": bill_profiles.public_list(),
    }


def build_header(filename: str, today: datetime.date | None = None) -> dict:
    today = today or datetime.date.today()
    stem = Path(filename).stem if filename else ""
    m = re.search(r"TF-?(\d{1,2})-?(\d{4})", stem, re.I)
    pv_no = f"TF-{int(m.group(1)):02d}/{m.group(2)}" if m else f"TF-XX/{today.year}"
    mi = re.search(r"Item\.?\s*(\d+)", stem, re.I)
    item = mi.group(1) if mi else ""
    date_str = f"{today.day:02d}-{THAI_MONTHS[today.month - 1]}-{today.year}"
    return {"pv_no": pv_no, "item": item, "date": date_str,
            "project": DEFAULT_HEADER["project"], "name": DEFAULT_HEADER["name"],
            "issued": DEFAULT_HEADER["issued"]}


# ---------------------------------------------------------------------------
FIRST_ROW, LAST_ROW = 12, 23


def build_pv_excel(lines: list[dict], header: dict, vendor: str = "",
                   bill_type: str | None = None) -> bytes:
    profile = bill_profiles.get_profile(bill_type)
    tpl_path, sheet = PV_TEMPLATES.get(profile.pv_template, PV_TEMPLATES["pv03"])
    wb = openpyxl.load_workbook(tpl_path)
    ws = wb[sheet]

    if header.get("pv_no"):   ws["M5"] = header["pv_no"]
    if header.get("item"):    ws["M6"] = header["item"]
    if header.get("date"):    ws["M7"] = header["date"]
    if header.get("project"): ws["C7"] = header["project"]
    if header.get("name"):    ws["C8"] = header["name"]
    if header.get("issued"):  ws["C9"] = header["issued"]

    r = FIRST_ROW
    vendor = vendor or (lines[0].get("vendor") if lines else "") or profile.default_vendor
    if vendor:
        ws.cell(row=r, column=2, value=vendor)
        r += 1

    for idx, ln in enumerate(lines, 1):
        if r > LAST_ROW:
            break
        amount = float(ln.get("amount") or 0)
        vat = float(ln.get("vat") or 0)
        wht = round(amount * profile.wht_rate, 2)
        net = round(amount + vat - wht, 2)
        desc = ln.get("desc") or _desc_from_dict(ln, profile)
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=" - " + desc)
        ws.cell(row=r, column=10, value=amount)              # J Amount
        ws.cell(row=r, column=11, value=vat if vat else None)  # K Vat ("-" ถ้า 0)
        ws.cell(row=r, column=12, value=-wht if wht else None) # L WHT ("-" ถ้า 0)
        ws.cell(row=r, column=13, value=net)                 # M Net
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _desc_from_dict(ln: dict, profile: BillProfile) -> str:
    bl = BillLine(identifier=ln.get("identifier", ln.get("phone", "")),
                  period=ln.get("period", ""), vendor=ln.get("vendor", ""))
    return bl.desc(profile)
