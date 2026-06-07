# Generates clusterMockData.js from the real "Cluster Level" sheet.
# Source: Z_PO/TRUE MERGE EAS (Internal) Master Progress Report-5.1_2026-05-28.xlsx
# Run: python clusterMockData.gen.py
import json, datetime, warnings
from openpyxl import load_workbook
warnings.filterwarnings('ignore')

SRC = r'C:\GoogleAppScript\0_NewServer\Z_PO\TRUE MERGE EAS (Internal) Master Progress Report-5.1_2026-05-28.xlsx'
OUT = r'C:\GoogleAppScript\0_NewServer\ACE_System4\frontend\clusterMockData.js'

wb = load_workbook(SRC, data_only=True)
ws = wb['Cluster Level']

def cell(r, c):
    return ws.cell(row=r, column=c).value

def d(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    return None

# column indices (1-based)
C_NAME, C_OWNER, C_MONTH, C_STATUS = 3, 7, 9, 12
C_SITES, C_READY, C_DTGEN, C_DTAPPR, C_RDATE = 14, 15, 17, 20, 23
C_INIT, C_PAOPEN = 29, 33
CR_COLS = [38, 54, 70, 86, 102, 118]  # 1st..6th CR Implemented
C_TUNE, C_PACTEST, C_PACSUB, C_PACAPPR, C_LIFE = 134, 140, 143, 147, 155
# Plan / trigger dates: EV Plan Open PA(40%), EW Plan Closed PA, EX Plan PAC Approved(60%)
C_PLAN_OPEN, C_PLAN_CLOSE, C_PLAN_PACAPPR = 152, 153, 154

# NEW 11-milestone order:
#  1 Site OnAir · 2 Cluster Ready · 3 DT Route Gen · 4 DT Route Appr · 5 Initial Test
#  6 PA Open(40%) · 7 PA Loop · 8 Tuning Closed · 9 PAC Report · 10 PAC Submit · 11 PAC Approved(60%)
MS_DATE_COL = {2: C_RDATE, 3: C_DTGEN, 4: C_DTAPPR, 5: C_INIT,
               6: C_PAOPEN, 8: C_TUNE, 9: C_PACTEST,
               10: C_PACSUB, 11: C_PACAPPR}
PALOOP_MID = 7   # PA Loop milestone id (latest CR date)
DONE_MID = 11    # final milestone (PAC Approved)

def status_phase(status):
    if not status:
        return 1
    pre = str(status).strip()[:2]
    try:
        n = int(pre)
    except ValueError:
        return 1
    return {14:11, 13:11, 12:10, 11:10, 10:10, 9:9, 8:8,
            7:7, 6:7, 5:7, 4:7, 3:7, 2:7, 1:5, 0:1}.get(n, 1)

rows = []
all_dates = []
for r in range(2, ws.max_row + 1):
    name = cell(r, C_NAME)
    if not name:
        continue
    name = str(name).strip()
    owner = (str(cell(r, C_OWNER)).strip() if cell(r, C_OWNER) else 'Unassigned')
    month = cell(r, C_MONTH)
    month = str(month).strip() if month else None
    status = str(cell(r, C_STATUS)).strip() if cell(r, C_STATUS) else '—'
    sites = cell(r, C_SITES) or 0
    ready = cell(r, C_READY)
    ready = round(float(ready), 3) if isinstance(ready, (int, float)) else 0.0

    # per-milestone real dates
    dates = {}
    cr1 = d(cell(r, CR_COLS[0]))
    for mid, col in MS_DATE_COL.items():
        dt = d(cell(r, col))
        if dt:
            dates[mid] = dt
    # PA Loop milestone -> latest CR implemented date
    cr_dates = [d(cell(r, c)) for c in CR_COLS]
    cr_dates = [x for x in cr_dates if x]
    if cr_dates:
        dates[PALOOP_MID] = max(cr_dates)

    # per-round detail (Tuning & Discuss R1-4) — block base = 38 + (N-1)*16
    pa_rounds = []
    for n in range(4):
        base = 38 + n * 16
        cr = d(cell(r, base))           # CR Implemented
        tune = d(cell(r, base + 6))     # Tuning R{n+1} Test Actual Done
        comp = d(cell(r, base + 11))    # PA_COMPARE Report Done
        disc = d(cell(r, base + 13))    # PA Discussion (TRUE) Date
        closed = cell(r, base + 14)     # #PA Closed
        added = cell(r, base + 15)      # #PA Added
        if not (cr or tune or comp or disc):
            continue
        pa_rounds.append({
            'round': n + 1,
            'cr': cr.isoformat() if cr else None,
            'tuning_done': tune.isoformat() if tune else None,
            'compare_done': comp.isoformat() if comp else None,
            'discuss_date': disc.isoformat() if disc else None,
            'pa_closed': int(closed) if isinstance(closed, (int, float)) else None,
            'pa_added': int(added) if isinstance(added, (int, float)) else None,
        })
    # milestone 1 (Site OnAir) -> earliest known date (fallback)
    if 2 in dates:
        dates[1] = dates.get(3) or dates[2]
    for v in dates.values():
        all_dates.append(v)

    # plan dates: 6=PA Open(40%), 8=PA Closed/Tuning, 11=PAC Approved(60%)
    plan = {}
    for mid, col in ((6, C_PLAN_OPEN), (8, C_PLAN_CLOSE), (11, C_PLAN_PACAPPR)):
        pdt = d(cell(r, col))
        if pdt:
            plan[mid] = pdt

    pa_round = len(cr_dates)

    # current phase: max(date-based, status-based)
    date_phase = 1
    if (sites or 0) > 0:
        date_phase = 1
    for mid in sorted([k for k in dates.keys()]):
        date_phase = max(date_phase, mid)
    cur = max(date_phase, status_phase(status), 1)
    cur = min(cur, DONE_MID)

    started = dates.get(2) or dates.get(3) or dates.get(5) or dates.get(1)
    life = cell(r, C_LIFE)
    life = int(life) if isinstance(life, (int, float)) else None

    rows.append({
        'r': r, 'name': name, 'owner': owner, 'month': month, 'status': status,
        'sites': int(sites) if isinstance(sites, (int, float)) else 0,
        'ready': ready, 'dates': dates, 'plan': plan, 'pa_round': pa_round, 'cur': cur,
        'started': started, 'life': life, 'pa_rounds': pa_rounds,
    })

# as-of = most recent actual date in the sheet (approx "today")
as_of = max(all_dates) if all_dates else datetime.date(2026, 5, 28)

def health_of(rec, age_at_phase):
    s = rec['status'][:2]
    if s == '14':
        return 'green'
    if s == '00':
        return 'red'
    if age_at_phase is not None and age_at_phase > 21:
        return 'red'
    if s in ('08', '11', '12', '13'):
        return 'green'
    return 'amber'

out = []
for rec in rows:
    dates = rec['dates']
    last_date = max(dates.values()) if dates else rec['started']
    age_at_phase = (as_of - last_date).days if last_date else 0
    if age_at_phase < 0:
        age_at_phase = 0
    if rec['life'] is not None:
        age_total = rec['life']
    elif rec['started']:
        age_total = (as_of - rec['started']).days
    else:
        age_total = age_at_phase
    health = health_of(rec, age_at_phase)
    # for done clusters, stuck time is small
    if rec['cur'] >= DONE_MID and rec['status'][:2] == '14':
        age_at_phase = min(age_at_phase, 3)

    out.append({
        'code': rec['name'],
        'owner': rec['owner'],
        'target_month': rec['month'] or '—',
        'status': rec['status'],
        'current_phase': rec['cur'],
        'age_total': int(age_total),
        'age_at_phase': int(age_at_phase),
        'site_count': rec['sites'],
        'readiness': rec['ready'],
        'started': rec['started'].isoformat() if rec['started'] else None,
        'health': health,
        'pa_round': rec['pa_round'],
        'last_action': rec['status'],
        'dates': {str(k): v.isoformat() for k, v in sorted(dates.items())},
        'plan': {str(k): v.isoformat() for k, v in sorted(rec['plan'].items())},
        'pa_rounds': rec['pa_rounds'],
    })

header = (
    '// AUTO-GENERATED from Z_PO/TRUE MERGE EAS Master Progress Report v5.1 — sheet "Cluster Level".\n'
    '// Regenerate: python clusterMockData.gen.py\n'
    f'// As-of date: {as_of.isoformat()} · {len(out)} clusters\n'
    'export const CLUSTERS = '
)
with open(OUT, 'w', encoding='utf-8') as f:
    f.write(header + json.dumps(out, ensure_ascii=False, indent=0) + '\n')
    f.write(f"export const AS_OF = '{as_of.isoformat()}'\n")

print(f'Wrote {len(out)} clusters to clusterMockData.js (as-of {as_of.isoformat()})')
from collections import Counter
print('owners:', dict(Counter(o["owner"] for o in out)))
print('health:', dict(Counter(o["health"] for o in out)))
print('phase :', dict(sorted(Counter(o["current_phase"] for o in out).items())))
print('SAMPLE:', json.dumps(out[0], ensure_ascii=False))
