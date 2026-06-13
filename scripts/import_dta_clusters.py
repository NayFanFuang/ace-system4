"""ETL: TRUE-MERGE EAS "Cluster Level" sheet -> dta_clusters + dta_cluster_rounds.

- Reads the Excel master report (sheet "Cluster Level").
- Owner = col 6 "FAC and TUNING Cluster Owner" (the real DTA).
- Joins project_sites (by rf_cluster_name) for centroid lat/lng.
- Upserts by rf_cluster_name so it is safe to re-run.

Usage:
  python scripts/import_dta_clusters.py [path-to-xlsx]
Env (optional):
  PGHOST (default 127.0.0.1) PGPORT (5433) PGUSER (ace_user)
  PGPASSWORD (ace_password) PGDATABASE (ace_system)
"""
import datetime
import os
import sys
import warnings

import psycopg2
import psycopg2.extras
from openpyxl import load_workbook

warnings.filterwarnings("ignore")

DEFAULT_XLSX = r"C:\GoogleAppScript\0_NewServer\POsystem\TRUE MERGE EAS (Internal) Master Progress Report-5.1_2026-06-08.xlsx"
XLSX = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX

DB = dict(
    host=os.getenv("PGHOST", "127.0.0.1"),
    port=int(os.getenv("PGPORT", "5433")),
    user=os.getenv("PGUSER", "ace_user"),
    password=os.getenv("PGPASSWORD", "ace_password"),
    dbname=os.getenv("PGDATABASE", "ace_system"),
)

# ---- column indices (1-based), matching clusterMockData.gen.py ----
C_NAME, C_OWNER, C_MONTH, C_STATUS = 3, 6, 9, 12      # owner = col 6 = FAC/TUNING DTA
C_SITES, C_READY, C_DTGEN, C_DTAPPR, C_RDATE = 14, 15, 17, 20, 23
C_INIT, C_PAOPEN = 29, 33
CR_COLS = [38, 54, 70, 86, 102, 118]
C_TUNE, C_PACTEST, C_PACSUB, C_PACAPPR, C_LIFE = 134, 140, 143, 147, 155
C_PLAN_OPEN, C_PLAN_CLOSE, C_PLAN_PACAPPR = 152, 153, 154

# milestone id -> source column (new 11-milestone order)
MS_DATE_COL = {2: C_RDATE, 3: C_DTGEN, 4: C_DTAPPR, 5: C_INIT,
               6: C_PAOPEN, 8: C_TUNE, 9: C_PACTEST, 10: C_PACSUB, 11: C_PACAPPR}
PALOOP_MID = 7
DONE_MID = 11
# milestone id -> dta_clusters column name
MS_COL_NAME = {1: "site_onair", 2: "cluster_ready", 3: "dt_gen", 4: "dt_approved",
               5: "init_test", 6: "pa_open", 8: "tuning_closed", 9: "pac_report",
               10: "pac_submit", 11: "pac_approved"}


def d(v):
    return v.date() if isinstance(v, datetime.datetime) else None


def status_phase(status):
    if not status:
        return 1
    try:
        n = int(str(status).strip()[:2])
    except ValueError:
        return 1
    return {14: 11, 13: 11, 12: 10, 11: 10, 10: 10, 9: 9, 8: 8,
            7: 7, 6: 7, 5: 7, 4: 7, 3: 7, 2: 7, 1: 5, 0: 1}.get(n, 1)


def health_of(status, age_at_phase):
    s = (status or "")[:2]
    if s == "14":
        return "green"
    if s == "00":
        return "red"
    if age_at_phase is not None and age_at_phase > 21:
        return "red"
    if s in ("08", "11", "12", "13"):
        return "green"
    return "amber"


def main():
    print(f"Reading {XLSX}")
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Cluster Level"]

    def cell(r, c):
        return ws.cell(row=r, column=c).value

    rows, all_dates = [], []
    for r in range(2, ws.max_row + 1):
        name = cell(r, C_NAME)
        if not name:
            continue
        name = str(name).strip()
        owner = str(cell(r, C_OWNER)).strip() if cell(r, C_OWNER) else None
        month = str(cell(r, C_MONTH)).strip() if cell(r, C_MONTH) else None
        status = str(cell(r, C_STATUS)).strip() if cell(r, C_STATUS) else None
        sites = cell(r, C_SITES) or 0
        ready = cell(r, C_READY)
        ready = round(float(ready), 3) if isinstance(ready, (int, float)) else None

        dates = {}
        for mid, col in MS_DATE_COL.items():
            dt = d(cell(r, col))
            if dt:
                dates[mid] = dt
        cr_dates = [d(cell(r, c)) for c in CR_COLS]
        cr_dates = [x for x in cr_dates if x]
        if cr_dates:
            dates[PALOOP_MID] = max(cr_dates)
        if 2 in dates:
            dates[1] = dates.get(3) or dates[2]
        all_dates += list(dates.values())

        plan = {}
        for mid, col in ((6, C_PLAN_OPEN), (8, C_PLAN_CLOSE), (11, C_PLAN_PACAPPR)):
            pdt = d(cell(r, col))
            if pdt:
                plan[mid] = pdt

        # per-round detail
        pa_rounds = []
        for n in range(4):
            base = 38 + n * 16
            cr = d(cell(r, base)); tune = d(cell(r, base + 6))
            comp = d(cell(r, base + 11)); disc = d(cell(r, base + 13))
            closed = cell(r, base + 14); added = cell(r, base + 15)
            if not (cr or tune or comp or disc):
                continue
            pa_rounds.append(dict(
                round_no=n + 1, cr_date=cr, tuning_done=tune, compare_done=comp,
                discuss_date=disc,
                pa_closed=int(closed) if isinstance(closed, (int, float)) else None,
                pa_added=int(added) if isinstance(added, (int, float)) else None,
            ))

        date_phase = max([1] + list(dates.keys()))
        cur = min(max(date_phase, status_phase(status), 1), DONE_MID)
        started = dates.get(2) or dates.get(3) or dates.get(5) or dates.get(1)
        life = cell(r, C_LIFE)
        life = int(life) if isinstance(life, (int, float)) else None

        rows.append(dict(name=name, owner=owner, month=month, status=status,
                         sites=int(sites) if isinstance(sites, (int, float)) else 0,
                         ready=ready, dates=dates, plan=plan, pa_round=len(cr_dates),
                         cur=cur, started=started, life=life, pa_rounds=pa_rounds))

    as_of = max(all_dates) if all_dates else datetime.date.today()
    print(f"Parsed {len(rows)} clusters · as-of {as_of}")

    conn = psycopg2.connect(**DB)
    conn.autocommit = False
    cur = conn.cursor()

    # centroid coords per cluster from project_sites
    cur.execute("""
        SELECT rf_cluster_name, AVG(lat), AVG(lng)
        FROM project_sites
        WHERE rf_cluster_name IS NOT NULL AND lat IS NOT NULL AND lng IS NOT NULL
        GROUP BY rf_cluster_name
    """)
    geo = {name: (float(la), float(ln)) for name, la, ln in cur.fetchall()}
    print(f"Geo from project_sites: {len(geo)} clusters")

    # PAC PO per cluster: bridge cluster → site_code prefix → PAC PO. One PO number per cluster.
    cur.execute("""
        SELECT s.rf_cluster_name, MIN(p.po_number) AS po_number
        FROM project_pos p
        JOIN project_sites s ON s.site_code = split_part(p.cluster_site, '_', 1)
        WHERE p.work_type = 'PAC' AND s.rf_cluster_name IS NOT NULL
        GROUP BY s.rf_cluster_name
    """)
    po_by_cluster = {name: po for name, po in cur.fetchall()}
    print(f"PO link from PAC PO: {len(po_by_cluster)} clusters")

    n_geo = 0
    for rec in rows:
        dates = rec["dates"]
        last_date = max(dates.values()) if dates else rec["started"]
        age_at_phase = (as_of - last_date).days if last_date else None
        if age_at_phase is not None and age_at_phase < 0:
            age_at_phase = 0
        age_total = rec["life"] if rec["life"] is not None else (
            (as_of - rec["started"]).days if rec["started"] else age_at_phase)
        if rec["cur"] >= DONE_MID and (rec["status"] or "")[:2] == "14" and age_at_phase is not None:
            age_at_phase = min(age_at_phase, 3)
        health = health_of(rec["status"], age_at_phase)
        lat, lng = geo.get(rec["name"], (None, None))
        if lat is not None:
            n_geo += 1

        ms = {col: dates.get(mid) for mid, col in MS_COL_NAME.items()}
        po_number = po_by_cluster.get(rec["name"])
        cur.execute("""
            INSERT INTO dta_clusters (
              rf_cluster_name, dta_name, target_month, status, current_phase, health,
              site_count, readiness, pa_round, age_total, age_at_phase,
              site_onair, cluster_ready, dt_gen, dt_approved, init_test, pa_open,
              tuning_closed, pac_report, pac_submit, pac_approved,
              plan_pa_open, plan_pa_closed, plan_pac_approved,
              lat, lng, po_number, owner_source, source, as_of_date, updated_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                    %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                    %s,%s,%s,%s,%s,%s,'EXCEL','EXCEL',%s, now())
            ON CONFLICT (rf_cluster_name) DO UPDATE SET
              -- preserve manually-assigned owner: ETL only updates dta_name when source is EXCEL
              dta_name = CASE WHEN dta_clusters.owner_source = 'MANUAL'
                              THEN dta_clusters.dta_name ELSE EXCLUDED.dta_name END,
              -- size facts stay sheet-owned (not "progress"):
              target_month=EXCLUDED.target_month,
              site_count=EXCLUDED.site_count, readiness=EXCLUDED.readiness,
              -- plan dates are app-ownable too: frozen once the cluster is system-native:
              plan_pa_open      = CASE WHEN dta_clusters.progress_source='MANUAL' THEN dta_clusters.plan_pa_open      ELSE EXCLUDED.plan_pa_open END,
              plan_pa_closed    = CASE WHEN dta_clusters.progress_source='MANUAL' THEN dta_clusters.plan_pa_closed    ELSE EXCLUDED.plan_pa_closed END,
              plan_pac_approved = CASE WHEN dta_clusters.progress_source='MANUAL' THEN dta_clusters.plan_pac_approved ELSE EXCLUDED.plan_pac_approved END,
              -- geo + PO + as_of always sync from the sheet:
              lat=EXCLUDED.lat, lng=EXCLUDED.lng, po_number=EXCLUDED.po_number,
              as_of_date=EXCLUDED.as_of_date, updated_at=now(),
              -- PROGRESS columns frozen once a DTA has edited in-app (progress_source='MANUAL'):
              status        = CASE WHEN dta_clusters.progress_source='MANUAL' THEN dta_clusters.status        ELSE EXCLUDED.status END,
              current_phase = CASE WHEN dta_clusters.progress_source='MANUAL' THEN dta_clusters.current_phase ELSE EXCLUDED.current_phase END,
              health        = CASE WHEN dta_clusters.progress_source='MANUAL' THEN dta_clusters.health        ELSE EXCLUDED.health END,
              pa_round      = CASE WHEN dta_clusters.progress_source='MANUAL' THEN dta_clusters.pa_round      ELSE EXCLUDED.pa_round END,
              age_total     = CASE WHEN dta_clusters.progress_source='MANUAL' THEN dta_clusters.age_total     ELSE EXCLUDED.age_total END,
              age_at_phase  = CASE WHEN dta_clusters.progress_source='MANUAL' THEN dta_clusters.age_at_phase  ELSE EXCLUDED.age_at_phase END,
              site_onair    = CASE WHEN dta_clusters.progress_source='MANUAL' THEN dta_clusters.site_onair    ELSE EXCLUDED.site_onair END,
              cluster_ready = CASE WHEN dta_clusters.progress_source='MANUAL' THEN dta_clusters.cluster_ready ELSE EXCLUDED.cluster_ready END,
              dt_gen        = CASE WHEN dta_clusters.progress_source='MANUAL' THEN dta_clusters.dt_gen        ELSE EXCLUDED.dt_gen END,
              dt_approved   = CASE WHEN dta_clusters.progress_source='MANUAL' THEN dta_clusters.dt_approved   ELSE EXCLUDED.dt_approved END,
              init_test     = CASE WHEN dta_clusters.progress_source='MANUAL' THEN dta_clusters.init_test     ELSE EXCLUDED.init_test END,
              pa_open       = CASE WHEN dta_clusters.progress_source='MANUAL' THEN dta_clusters.pa_open        ELSE EXCLUDED.pa_open END,
              tuning_closed = CASE WHEN dta_clusters.progress_source='MANUAL' THEN dta_clusters.tuning_closed ELSE EXCLUDED.tuning_closed END,
              pac_report    = CASE WHEN dta_clusters.progress_source='MANUAL' THEN dta_clusters.pac_report    ELSE EXCLUDED.pac_report END,
              pac_submit    = CASE WHEN dta_clusters.progress_source='MANUAL' THEN dta_clusters.pac_submit    ELSE EXCLUDED.pac_submit END,
              pac_approved  = CASE WHEN dta_clusters.progress_source='MANUAL' THEN dta_clusters.pac_approved  ELSE EXCLUDED.pac_approved END
        """, (
            rec["name"], rec["owner"], rec["month"], rec["status"], rec["cur"], health,
            rec["sites"], rec["ready"], rec["pa_round"], age_total, age_at_phase,
            ms["site_onair"], ms["cluster_ready"], ms["dt_gen"], ms["dt_approved"],
            ms["init_test"], ms["pa_open"], ms["tuning_closed"], ms["pac_report"],
            ms["pac_submit"], ms["pac_approved"],
            rec["plan"].get(6), rec["plan"].get(8), rec["plan"].get(11),
            lat, lng, po_number, as_of,
        ))

        # replace rounds — but skip when the cluster is system-native (progress edited in-app).
        # Same txn as the upsert above (single commit), so this SELECT sees the just-written value.
        cur.execute("SELECT progress_source FROM dta_clusters WHERE rf_cluster_name=%s", (rec["name"],))
        row = cur.fetchone()
        if not (row and row[0] == "MANUAL"):
            cur.execute("DELETE FROM dta_cluster_rounds WHERE rf_cluster_name=%s", (rec["name"],))
            for rd in rec["pa_rounds"]:
                cur.execute("""
                    INSERT INTO dta_cluster_rounds
                      (rf_cluster_name, round_no, cr_date, tuning_done, compare_done,
                       discuss_date, pa_closed, pa_added)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                """, (rec["name"], rd["round_no"], rd["cr_date"], rd["tuning_done"],
                      rd["compare_done"], rd["discuss_date"], rd["pa_closed"], rd["pa_added"]))

    # Link owners (team-code dta_name like "TH_Natdanai") to a login by stripping the
    # country prefix and matching first_name / full name — unambiguous matches only.
    # Keeps new clusters' owners scoped to /dta/my-clusters without manual assignment.
    cur.execute("""
        UPDATE dta_clusters d SET assigned_employee_code = m.employee_code
        FROM (
          SELECT n.dta_name, MIN(a.employee_code) AS employee_code
          FROM (SELECT DISTINCT dta_name, regexp_replace(dta_name,'^[A-Z]{2}_','') AS stripped
                FROM dta_clusters WHERE dta_name IS NOT NULL) n
          JOIN auth_users a ON a.is_active
                           AND (lower(a.first_name)=lower(n.stripped)
                                OR lower(a.first_name||' '||a.last_name)=lower(n.stripped))
          GROUP BY n.dta_name HAVING count(DISTINCT a.employee_code)=1
        ) m
        WHERE d.dta_name=m.dta_name AND d.assigned_employee_code IS NULL
    """)
    n_linked = cur.rowcount

    conn.commit()
    cur.execute("SELECT count(*) FROM dta_clusters")
    total = cur.fetchone()[0]
    cur.execute("SELECT count(*) FROM dta_cluster_rounds")
    nrounds = cur.fetchone()[0]
    cur.close(); conn.close()
    print(f"Upserted {len(rows)} clusters ({n_geo} with geo) · {nrounds} rounds · table total={total} · owners auto-linked={n_linked}")


if __name__ == "__main__":
    main()
