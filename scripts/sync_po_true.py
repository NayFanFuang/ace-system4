"""
Sync PO True System Test(1).xlsx  Sheet2 into:
  - project_pos      (all 126 rows, po_target='RF')
  - clock_sites      (unique sites by DU ID → DTE clock-in)
  - project_catalog  (HWT2601 upsert)

Run inside the backend container:
  docker exec ace-system-backend python /app/scripts/sync_po_true.py
"""
import asyncio
import sys
from datetime import date, datetime, time as dtime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl
from sqlalchemy import select

from app.database import SessionLocal, engine, Base
from app.models.employee import ProjectCatalog, ProjectPO
from app.models.clock import ClockSite

XLSX_PATH = Path(__file__).resolve().parents[1] / "PO_True.xlsx"

# Sheet2 column indices (0-based)
COL_DU_ID      = 0   # A: Site Code (DU ID)
COL_PO_NO      = 1   # B: PO NO.
COL_PO_LINE    = 2   # C: PO Line
COL_ITEM_DIS   = 3   # D: PO Item Description
# E: Publish Date — skip
COL_TYPE       = 5   # F: Type
COL_CLUSTER    = 6   # G: Cluster Name
COL_OWNER      = 7   # H: Owner
COL_LAT        = 8   # I: Latitude
COL_LNG        = 9   # J: Longitude
COL_ON_AIR     = 10  # K: On-Air
COL_PROJECT    = 11  # L: Project Name


def _clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _to_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _to_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, dtime):
        return None  # time(0,0) sentinel from Excel
    return None


def _parse_project_code(raw):
    """Extract 'HWT2601' from 'HWT2601 : RF TRUE/HWT Flash EAS&BMA Project'"""
    if not raw:
        return None
    s = str(raw).strip()
    if " : " in s:
        return s.split(" : ", 1)[0].strip()
    return s.split()[0] if s else None


def load_sheet2():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    # Find Sheet2
    ws = None
    if "Sheet2" in wb.sheetnames:
        ws = wb["Sheet2"]
    elif len(wb.sheetnames) > 1:
        ws = wb.worksheets[1]
    else:
        ws = wb.active
    print(f"Reading sheet: {ws.title!r}  ({ws.max_row - 1} data rows)")

    po_rows = []
    site_map = {}  # site_code → {site_code, site_name, lat, lng, project_code}

    for row in ws.iter_rows(min_row=2, values_only=True):
        du_id   = _clean(row[COL_DU_ID])
        po_no   = _clean(row[COL_PO_NO])
        if not du_id and not po_no:
            continue  # skip blank rows

        project_code = _parse_project_code(row[COL_PROJECT])
        lat          = _to_float(row[COL_LAT])
        lng          = _to_float(row[COL_LNG])
        cluster      = _clean(row[COL_CLUSTER])
        owner        = _clean(row[COL_OWNER])
        if owner and owner == '0':
            owner = None

        lat_long = f"{lat},{lng}" if lat and lng else None

        po_rows.append({
            "po_target":    "RF",
            "project_code": project_code,
            "po_number":    po_no or "UNKNOWN",
            "po_line":      _clean(row[COL_PO_LINE]),
            "du_id":        du_id,
            "item_dis":     _clean(row[COL_ITEM_DIS]),
            "cluster_site": cluster,
            "owner":        owner,
            "lat_long":     lat_long,
            "on_air":       _to_date(row[COL_ON_AIR]),
            "cluster_type": _clean(row[COL_TYPE]),
        })

        if du_id and du_id not in site_map:
            site_map[du_id] = {
                "site_code":    du_id,
                "site_name":    cluster or du_id,
                "customer":     "TRUE",
                "project_code": project_code,
                "lat":          lat,
                "lng":          lng,
                "gps_radius_m": 500,
                "is_active":    True,
            }
        elif du_id and site_map[du_id]["lat"] is None and lat:
            # Fill in missing lat/lng from later rows
            site_map[du_id]["lat"] = lat
            site_map[du_id]["lng"] = lng

    return po_rows, list(site_map.values())


async def run():
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    po_rows, site_rows = load_sheet2()
    print(f"PO rows: {len(po_rows)}  |  Unique sites: {len(site_rows)}")

    # Collect unique project codes
    project_codes = {r["project_code"] for r in po_rows if r["project_code"]}

    async with SessionLocal() as db:
        # ── 1. Upsert project_catalog ────────────────────────────────────────
        for pc in project_codes:
            ex = (await db.execute(select(ProjectCatalog).where(ProjectCatalog.project_code == pc))).scalar_one_or_none()
            if not ex:
                db.add(ProjectCatalog(project_code=pc, project_name=f"{pc} Project", team="RF", headcount=0))
                print(f"  Added project_catalog: {pc}")
            else:
                print(f"  Already in catalog: {pc} (team={ex.team})")
        await db.commit()

        # ── 2. Upsert project_pos ─────────────────────────────────────────────
        po_inserted = po_updated = 0
        for r in po_rows:
            ex = (await db.execute(
                select(ProjectPO).where(
                    ProjectPO.po_number == r["po_number"],
                    ProjectPO.po_line   == r["po_line"],
                )
            )).scalar_one_or_none()
            if ex:
                for k, v in r.items():
                    if v is not None:
                        setattr(ex, k, v)
                po_updated += 1
            else:
                db.add(ProjectPO(**r))
                po_inserted += 1
        await db.commit()
        print(f"project_pos: {po_inserted} inserted, {po_updated} updated")

        # ── 3. Upsert clock_sites ─────────────────────────────────────────────
        cs_inserted = cs_updated = 0
        for s in site_rows:
            ex = (await db.execute(
                select(ClockSite).where(ClockSite.site_code == s["site_code"])
            )).scalar_one_or_none()
            if ex:
                ex.site_name    = s["site_name"]
                ex.customer     = s["customer"]
                ex.project_code = s["project_code"]
                if s["lat"]:
                    ex.lat = s["lat"]
                    ex.lng = s["lng"]
                cs_updated += 1
            else:
                db.add(ClockSite(**s))
                cs_inserted += 1
        await db.commit()
        print(f"clock_sites: {cs_inserted} inserted, {cs_updated} updated")

    print("Done.")


if __name__ == "__main__":
    asyncio.run(run())
