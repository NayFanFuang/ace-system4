"""
Sync Raw_Clock.xlsx → clock_sessions

Sheet "DTE"   → PER_SITE sessions  (Start Work → Stop Work / Complete)
Sheet "Other" → DAILY sessions     (Clock In → Clock Out)

Run inside backend container:
  docker exec ace-system-backend python /app/scripts/sync_raw_clock.py
"""
import asyncio
import re
import sys
from collections import defaultdict
from datetime import date, datetime, timezone, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl
from sqlalchemy import select, text

from app.database import SessionLocal, engine, Base
from app.models.clock import ClockSession
from app.models.employee import Employee

XLSX_PATH = Path(__file__).resolve().parents[1] / "Raw_Clock.xlsx"
TZ_BKK    = timezone(timedelta(hours=7))
BATCH     = 500

# ─── helpers ──────────────────────────────────────────────────────────────────

def _aware(dt):
    if dt is None:
        return None
    if isinstance(dt, datetime):
        return dt.replace(tzinfo=TZ_BKK) if dt.tzinfo is None else dt
    return None


def _clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _norm_name(raw):
    """'DTE01_Mr.Chinnakrit Krongthong' → 'chinnakrit krongthong'"""
    s = re.sub(r'^[A-Z]+\d+_', '', (raw or '').strip())
    s = re.sub(r'^(Mr\.|Ms\.|Mrs\.|Miss\.|Dr\.)\s*', '', s, flags=re.IGNORECASE)
    return s.lower().strip()


def _norm_full(raw):
    """'Mr.Chinnakrit Krongthong' → 'chinnakrit krongthong'"""
    s = re.sub(r'^(Mr\.|Ms\.|Mrs\.|Miss\.|Dr\.)\s*', '', (raw or '').strip(), flags=re.IGNORECASE)
    return s.lower().strip()


# ─── employee lookup ──────────────────────────────────────────────────────────

async def build_lookup(db) -> tuple[dict, dict]:
    """Returns (email→code, norm_name→code)"""
    rows = (await db.execute(select(Employee))).scalars().all()
    by_email = {}
    by_name  = {}
    for r in rows:
        if r.email:
            by_email[r.email.lower().strip()] = r.employee_code
        if r.personal_email:
            by_email[r.personal_email.lower().strip()] = r.employee_code
        by_name[_norm_full(r.full_name)] = r.employee_code
    return by_email, by_name


def resolve_code(email, name_raw, by_email, by_name, fallback_prefix="LEGACY"):
    """Match to employee_code, or derive a legacy code."""
    if email:
        code = by_email.get(email.lower().strip())
        if code:
            return code
    norm = _norm_name(name_raw) if name_raw else ''
    code = by_name.get(norm)
    if code:
        return code
    # Derive a stable legacy code
    prefix = re.match(r'^([A-Z]+\d+)_', name_raw or '')
    if prefix:
        return f"LEG-{prefix.group(1)}"
    if email:
        local = email.split('@')[0][:15]
        return f"LEG-{local}"
    return f"LEG-UNK"


# ─── load sheets ──────────────────────────────────────────────────────────────

def load_dte_sheet():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    ws = wb["DTE"]
    events = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ts     = _aware(row[0]) if isinstance(row[0], datetime) else None
        name   = _clean(row[1])
        email  = _clean(row[2])
        proj   = _clean(row[3])
        status = _clean(row[4])
        job    = _clean(row[5])
        lat    = row[6] if isinstance(row[6], (int, float)) else None
        lng    = row[7] if isinstance(row[7], (int, float)) else None
        photo1 = _clean(row[8])
        photo2 = _clean(row[9])
        if ts is None or status not in ('Start Work', 'Stop Work', 'Complete', 'Issue'):
            continue
        events.append({
            'ts': ts, 'name': name, 'email': email,
            'project': proj, 'status': status, 'job': job,
            'lat': lat, 'lng': lng,
            'photo1': photo1, 'photo2': photo2,
        })
    wb.close()
    print(f"DTE sheet: {len(events)} events loaded")
    return events


def load_other_sheet():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    ws = wb["Other"]
    events = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ts     = _aware(row[0]) if isinstance(row[0], datetime) else None
        email  = _clean(row[1])
        name   = _clean(row[2])
        lat    = row[3] if isinstance(row[3], (int, float)) else None
        lng    = row[4] if isinstance(row[4], (int, float)) else None
        status = _clean(row[5])
        photo  = _clean(row[6])
        proj   = _clean(row[7])
        loc    = _clean(row[8])
        if ts is None or status not in ('Clock In', 'Clock Out'):
            continue
        events.append({
            'ts': ts, 'email': email, 'name': name,
            'lat': lat, 'lng': lng, 'status': status,
            'photo': photo, 'project': proj, 'location': loc,
        })
    wb.close()
    print(f"Other sheet: {len(events)} events loaded")
    return events


# ─── pairing ──────────────────────────────────────────────────────────────────

def pair_dte(events, by_email, by_name):
    """Returns list of ClockSession dicts."""
    # Group events by employee (name is stable identifier in DTE sheet)
    by_emp = defaultdict(list)
    for e in events:
        by_emp[e['name']].append(e)

    sessions = []
    unmatched_open = 0
    for name, evts in by_emp.items():
        evts.sort(key=lambda x: x['ts'])
        emp_code = resolve_code(evts[0]['email'], name, by_email, by_name)
        pending = {}  # job_key → event

        for evt in evts:
            job = (evt['job'] or 'UNKNOWN')[:50]
            st  = evt['status']

            if st == 'Start Work':
                # Flush existing pending for same job (missed close)
                if job in pending:
                    sessions.append(_make_per_site(pending.pop(job), None, emp_code))
                    unmatched_open += 1
                pending[job] = evt

            elif st in ('Stop Work', 'Complete', 'Issue'):
                if job in pending:
                    sessions.append(_make_per_site(pending.pop(job), evt, emp_code))
                # orphan close → ignore

        # Flush remaining (no close recorded)
        for job, evt in pending.items():
            sessions.append(_make_per_site(evt, None, emp_code))
            unmatched_open += 1

    print(f"DTE pairs: {len(sessions)} sessions  ({unmatched_open} with no clock-out)")
    return sessions


def _make_per_site(open_evt, close_evt, emp_code):
    job = (open_evt['job'] or 'UNKNOWN')[:50]
    return {
        'employee_code': emp_code,
        'user_id':       None,
        'clock_type':    'PER_SITE',
        'work_date':     open_evt['ts'].date(),
        'site_id':       None,
        'site_code':     job,
        'site_name':     (open_evt['job'] or '')[:200],
        'clock_in_at':   open_evt['ts'],
        'lat_in':        open_evt['lat'],
        'lng_in':        open_evt['lng'],
        'photo_in':      open_evt['photo1'],
        'clock_out_at':  close_evt['ts'] if close_evt else None,
        'lat_out':       close_evt['lat'] if close_evt else None,
        'lng_out':       close_evt['lng'] if close_evt else None,
        'photo_out':     close_evt['photo1'] if close_evt else None,
        'status':        'COMPLETED',
    }


def pair_other(events, by_email, by_name):
    """Returns list of ClockSession dicts."""
    # Group by email (most reliable key in Other sheet)
    by_email_key = defaultdict(list)
    for e in events:
        key = (e['email'] or e['name'] or 'unknown').lower()
        by_email_key[key].append(e)

    sessions = []
    unmatched_in = 0
    for key, evts in by_email_key.items():
        evts.sort(key=lambda x: x['ts'])
        first = evts[0]
        emp_code = resolve_code(first['email'], first['name'], by_email, by_name)
        pending_in = None

        for evt in evts:
            if evt['status'] == 'Clock In':
                if pending_in:
                    sessions.append(_make_daily(pending_in, None, emp_code))
                    unmatched_in += 1
                pending_in = evt
            elif evt['status'] == 'Clock Out':
                if pending_in:
                    sessions.append(_make_daily(pending_in, evt, emp_code))
                    pending_in = None
                # orphan clock out → ignore

        if pending_in:
            sessions.append(_make_daily(pending_in, None, emp_code))
            unmatched_in += 1

    print(f"Other pairs: {len(sessions)} sessions  ({unmatched_in} with no clock-out)")
    return sessions


def _make_daily(in_evt, out_evt, emp_code):
    return {
        'employee_code': emp_code,
        'user_id':       None,
        'clock_type':    'DAILY',
        'work_date':     in_evt['ts'].date(),
        'site_id':       None,
        'site_code':     None,
        'site_name':     in_evt.get('location') or in_evt.get('project'),
        'clock_in_at':   in_evt['ts'],
        'lat_in':        in_evt['lat'],
        'lng_in':        in_evt['lng'],
        'photo_in':      in_evt['photo'],
        'clock_out_at':  out_evt['ts'] if out_evt else None,
        'lat_out':       out_evt['lat'] if out_evt else None,
        'lng_out':       out_evt['lng'] if out_evt else None,
        'photo_out':     out_evt['photo'] if out_evt else None,
        'status':        'COMPLETED',
    }


# ─── main ─────────────────────────────────────────────────────────────────────

async def run():
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    dte_events   = load_dte_sheet()
    other_events = load_other_sheet()

    async with SessionLocal() as db:
        by_email, by_name = await build_lookup(db)
        print(f"Employee lookup: {len(by_email)} emails, {len(by_name)} names")

        all_sessions = []
        all_sessions += pair_dte(dte_events, by_email, by_name)
        all_sessions += pair_other(other_events, by_email, by_name)
        print(f"Total sessions to import: {len(all_sessions)}")

        # Build existing key set to skip duplicates
        existing = set()
        rows = (await db.execute(
            text("SELECT employee_code, work_date, clock_in_at FROM clock_sessions")
        )).fetchall()
        for r in rows:
            existing.add((r[0], r[1], r[2]))
        print(f"Existing sessions in DB: {len(existing)}")

        inserted = skipped = 0
        batch = []
        for s in all_sessions:
            key = (s['employee_code'], s['work_date'], s['clock_in_at'])
            if key in existing:
                skipped += 1
                continue
            existing.add(key)
            batch.append(s)
            if len(batch) >= BATCH:
                await db.execute(ClockSession.__table__.insert(), batch)
                await db.commit()
                inserted += len(batch)
                batch = []
                print(f"  Inserted {inserted}…")

        if batch:
            await db.execute(ClockSession.__table__.insert(), batch)
            await db.commit()
            inserted += len(batch)

    print(f"\nDone.  inserted={inserted}  skipped={skipped}")


if __name__ == "__main__":
    asyncio.run(run())
