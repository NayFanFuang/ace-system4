"""
ACE Special Export — desktop tool (Tkinter GUI)
================================================
Pick an employee + month, then export their Attendance Timesheet PDF and/or a
Clock In/Out Excel. Data comes from the production server (read-only); PDFs are
built on the server (same engine as the website) and downloaded locally.

Run:  python tools/special_export.py
Needs: paramiko, openpyxl  (tkinter ships with Python).

SSH credentials are read from deploy.py at runtime (not duplicated here).
Nothing is modified on the server — this only reads data and generates files.
"""
import io
import os
import re
import threading
import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import paramiko

# ---------------------------------------------------------------- config ----
HERE = os.path.dirname(os.path.abspath(__file__))
DEPLOY_PY = os.path.normpath(os.path.join(HERE, "..", "deploy.py"))
PG = "ace-system-postgres"
BE = "ace-system-backend"
DEFAULT_OUT = os.path.join(os.path.expanduser("~"), "Downloads", "ACE_Exports")


def _load_cfg():
    src = open(DEPLOY_PY, encoding="utf-8").read()
    cfg = {}
    for k in ("HOST", "PORT", "USER", "PASSWORD"):
        m = re.search(rf'^{k}\s*=\s*(.+?)\s*$', src, re.M)
        if not m:
            raise RuntimeError(f"{k} not found in deploy.py")
        cfg[k] = eval(m.group(1))
    return cfg


def _ssh(cfg):
    c = paramiko.SSHClient()
    c.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    c.connect(cfg["HOST"], port=cfg["PORT"], username=cfg["USER"],
              password=cfg["PASSWORD"], timeout=30)
    return c


def _psql(ssh, sql, tuples=True):
    """Run a SELECT via docker exec; return list of pipe-split rows."""
    flags = "-tA -F'|'" if tuples else "-c"
    cmd = f"docker exec {PG} psql -U ace_user -d ace_system {flags} -c \"{sql}\""
    _, o, e = ssh.exec_command(cmd, timeout=120)
    out = o.read().decode()
    err = e.read().decode()
    rows = [ln.split("|") for ln in out.splitlines() if ln.strip() and not ln.startswith(("psql:", "ERROR"))]
    return rows, err


# Server-side PDF builder (uploaded & executed inside the backend container).
GEN_PDF_SRC = r'''
import sys, asyncio
from sqlalchemy import select, and_
from app.database import SessionLocal
from app.models.employee import Employee, ProjectAssignment, ProjectCatalog
from app.models.clock import ClockSession
from app.models.worklog import DailyWorkLog
from app.routers.worklog import _timesheet_period, _build_timesheet_pdf

EMP = sys.argv[1]; MONTH = sys.argv[2]; SD = int(sys.argv[3])

async def main():
    async with SessionLocal() as db:
        ps, pe = _timesheet_period(MONTH, SD)
        e = (await db.execute(select(Employee).where(Employee.employee_code==EMP))).scalar_one_or_none()
        mgr = None
        if e and e.manager_code:
            mgr = (await db.execute(select(Employee).where(Employee.employee_code==e.manager_code))).scalar_one_or_none()
        po = None
        if e and not (e.project_code or e.project_name):
            fa = (await db.execute(select(ProjectAssignment, ProjectCatalog).join(ProjectCatalog, ProjectCatalog.project_code==ProjectAssignment.project_code, isouter=True).where(and_(ProjectAssignment.employee_code==EMP, ProjectAssignment.is_active.is_(True))).order_by(ProjectAssignment.project_code.asc()).limit(1))).first()
            if fa:
                pa, pc = fa; po = (pa.project_code, pc.project_name if pc else None)
        ss = (await db.execute(select(ClockSession).where(and_(ClockSession.employee_code==EMP, ClockSession.work_date>=ps, ClockSession.work_date<=pe)).order_by(ClockSession.clock_in_at.asc()))).scalars().all()
        lg = (await db.execute(select(DailyWorkLog).where(and_(DailyWorkLog.employee_code==EMP, DailyWorkLog.work_date>=ps, DailyWorkLog.work_date<=pe)))).scalars().all()
        sbd = {}
        for s in ss: sbd.setdefault(s.work_date, []).append(s)
        lbd = {l.work_date: l for l in lg}
        buf = _build_timesheet_pdf(employee=e, emp_code=EMP, manager=mgr, project_override=po, period_start=ps, period_end=pe, sessions_by_date=sbd, logs_by_date=lbd)
        with open("/tmp/_se_out.pdf","wb") as f: f.write(buf.getvalue())
        print("OK")
asyncio.run(main())
'''


class App:
    def __init__(self, root):
        self.root = root
        self.cfg = None
        self.emps = []            # list of (code, name)
        root.title("ACE Special Export")
        root.geometry("560x440")
        root.resizable(False, False)

        pad = {"padx": 10, "pady": 6}
        frm = ttk.Frame(root, padding=12)
        frm.pack(fill="both", expand=True)

        ttk.Label(frm, text="ACE Special Export", font=("Segoe UI", 14, "bold")).grid(row=0, column=0, columnspan=3, sticky="w")
        ttk.Label(frm, text="Timesheet PDF / Clock Excel — from production (read-only)",
                  foreground="#667").grid(row=1, column=0, columnspan=3, sticky="w", pady=(0, 8))

        ttk.Label(frm, text="Employee:").grid(row=2, column=0, sticky="w", **pad)
        self.emp_cb = ttk.Combobox(frm, width=48, state="normal")
        self.emp_cb.grid(row=2, column=1, columnspan=2, sticky="w", **pad)
        self.emp_cb.bind("<KeyRelease>", self._filter_emps)

        ttk.Label(frm, text="Month:").grid(row=3, column=0, sticky="w", **pad)
        self.month_cb = ttk.Combobox(frm, width=14, state="readonly", values=self._months())
        self.month_cb.grid(row=3, column=1, sticky="w", **pad)
        self.month_cb.set(datetime.date.today().strftime("%Y-%m"))

        ttk.Label(frm, text="Period:").grid(row=4, column=0, sticky="w", **pad)
        self.start_day = tk.IntVar(value=1)
        pf = ttk.Frame(frm); pf.grid(row=4, column=1, columnspan=2, sticky="w", **pad)
        ttk.Radiobutton(pf, text="Calendar (1–end)", variable=self.start_day, value=1).pack(side="left")
        ttk.Radiobutton(pf, text="Payroll (26–25)", variable=self.start_day, value=26).pack(side="left", padx=(12, 0))

        ttk.Label(frm, text="Output:").grid(row=5, column=0, sticky="w", **pad)
        self.out_var = tk.StringVar(value=DEFAULT_OUT)
        ttk.Entry(frm, textvariable=self.out_var, width=40).grid(row=5, column=1, sticky="w", **pad)
        ttk.Button(frm, text="Browse…", command=self._browse).grid(row=5, column=2, sticky="w")

        bf = ttk.Frame(frm); bf.grid(row=6, column=0, columnspan=3, pady=12)
        self.btn_pdf = ttk.Button(bf, text="Export Timesheet PDF", command=lambda: self._run(self.export_pdf))
        self.btn_pdf.pack(side="left", padx=4)
        self.btn_xls = ttk.Button(bf, text="Export Clock Excel", command=lambda: self._run(self.export_xlsx))
        self.btn_xls.pack(side="left", padx=4)
        self.btn_both = ttk.Button(bf, text="Export Both", command=lambda: self._run(self.export_both))
        self.btn_both.pack(side="left", padx=4)

        self.log = tk.Text(frm, height=9, width=64, state="disabled", bg="#f7f8fa", font=("Consolas", 9))
        self.log.grid(row=7, column=0, columnspan=3, sticky="we", pady=(4, 0))

        self._say("Connecting to server and loading employees…")
        threading.Thread(target=self._load_emps, daemon=True).start()

    # ----- helpers -----
    def _months(self):
        today = datetime.date.today()
        out = []
        y, m = today.year, today.month
        for _ in range(24):
            out.append(f"{y:04d}-{m:02d}")
            m -= 1
            if m == 0: m = 12; y -= 1
        return out

    def _say(self, msg):
        self.log.configure(state="normal")
        self.log.insert("end", msg + "\n")
        self.log.see("end")
        self.log.configure(state="disabled")
        self.root.update_idletasks()

    def _browse(self):
        d = filedialog.askdirectory(initialdir=self.out_var.get() or os.path.expanduser("~"))
        if d: self.out_var.set(d)

    def _set_buttons(self, enabled):
        st = "normal" if enabled else "disabled"
        for b in (self.btn_pdf, self.btn_xls, self.btn_both): b.configure(state=st)

    def _load_emps(self):
        try:
            self.cfg = _load_cfg()
            ssh = _ssh(self.cfg)
            rows, err = _psql(ssh, "SELECT employee_code, full_name FROM employees WHERE full_name IS NOT NULL ORDER BY full_name")
            ssh.close()
            self.emps = [(r[0], r[1]) for r in rows if len(r) >= 2]
            self._all_labels = [f"{c} — {n}" for c, n in self.emps]
            self.emp_cb["values"] = self._all_labels
            self._say(f"Loaded {len(self.emps)} employees. Ready.")
        except Exception as ex:
            self._say(f"ERROR loading employees: {ex}")
            messagebox.showerror("Connection error", str(ex))

    def _filter_emps(self, _evt):
        typed = self.emp_cb.get().lower()
        if not typed:
            self.emp_cb["values"] = self._all_labels
            return
        self.emp_cb["values"] = [l for l in self._all_labels if typed in l.lower()]

    def _selected_emp(self):
        val = self.emp_cb.get().strip()
        # match "CODE — Name" or just code/name typed
        for c, n in self.emps:
            if val == f"{c} — {n}" or val == c:
                return c, n
        # fallback: contains
        for c, n in self.emps:
            if val.lower() in f"{c} {n}".lower():
                return c, n
        return None, None

    def _run(self, fn):
        self._set_buttons(False)
        threading.Thread(target=self._wrap, args=(fn,), daemon=True).start()

    def _wrap(self, fn):
        try:
            fn()
        except Exception as ex:
            self._say(f"ERROR: {ex}")
            messagebox.showerror("Export error", str(ex))
        finally:
            self._set_buttons(True)

    def _ctx(self):
        code, name = self._selected_emp()
        if not code:
            raise RuntimeError("Please select an employee.")
        month = self.month_cb.get().strip()
        if not re.match(r"^\d{4}-\d{2}$", month):
            raise RuntimeError("Please select a month.")
        outdir = self.out_var.get().strip() or DEFAULT_OUT
        os.makedirs(outdir, exist_ok=True)
        return code, name, month, int(self.start_day.get()), outdir

    # ----- exports -----
    def export_pdf(self):
        code, name, month, sd, outdir = self._ctx()
        self._say(f"[PDF] {name} {month} (period start={sd}) …")
        ssh = _ssh(self.cfg)
        try:
            sftp = ssh.open_sftp()
            with sftp.open("/home/pn_deploy/_se_gen.py", "w") as f:
                f.write(GEN_PDF_SRC)
            cp = "docker cp /home/pn_deploy/_se_gen.py %s:/app/_se_gen.py" % BE
            ssh.exec_command(cp)[1].channel.recv_exit_status()
            run = f"docker exec {BE} bash -lc 'cd /app && python3 _se_gen.py {code} {month} {sd}'"
            _, o, e = ssh.exec_command(run, timeout=120)
            out = o.read().decode(); err = e.read().decode()
            if "OK" not in out:
                raise RuntimeError("PDF build failed: " + (err or out)[-400:])
            safe = name.replace(" ", "_").replace("/", "_")
            local = os.path.join(outdir, f"TimeSheet_{month}_{safe}.pdf")
            ssh.exec_command(f"docker cp {BE}:/tmp/_se_out.pdf /home/pn_deploy/_se_out.pdf")[1].channel.recv_exit_status()
            sftp.get("/home/pn_deploy/_se_out.pdf", local)
            ssh.exec_command(f"rm -f /home/pn_deploy/_se_gen.py /home/pn_deploy/_se_out.pdf; docker exec {BE} rm -f /app/_se_gen.py /tmp/_se_out.pdf")
            sftp.close()
            self._say(f"[PDF] saved: {local} ({os.path.getsize(local)} bytes)")
        finally:
            ssh.close()

    def export_xlsx(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        code, name, month, sd, outdir = self._ctx()
        self._say(f"[XLSX] {name} {month} …")
        y, m = map(int, month.split("-"))
        ps = datetime.date(y, m, 1)
        pe = (datetime.date(y + (m == 12), (m % 12) + 1, 1) - datetime.timedelta(days=1))
        ssh = _ssh(self.cfg)
        try:
            sql = (
                "SET TIME ZONE 'Asia/Bangkok'; "
                "SELECT work_date, to_char(work_date,'Dy'), to_char(clock_in_at,'HH24:MI'), "
                "to_char(clock_out_at,'HH24:MI'), "
                "round(EXTRACT(EPOCH FROM (clock_out_at-clock_in_at))/3600.0,2), "
                "coalesce(site_code,''), "
                "CASE WHEN clock_out_at IS NULL THEN 'no clock-out' "
                "WHEN EXTRACT(EPOCH FROM (clock_out_at-clock_in_at))=28800 AND photo_out IS NULL THEN 'auto 8h (no out)' "
                "WHEN EXTRACT(EPOCH FROM (clock_out_at-clock_in_at))<360 THEN 'short tap' "
                "WHEN clock_out_at::date>clock_in_at::date THEN 'cross-midnight' ELSE '' END "
                f"FROM clock_sessions WHERE employee_code='{code}' "
                f"AND work_date BETWEEN '{ps}' AND '{pe}' ORDER BY work_date, clock_in_at"
            )
            cmd = f"docker exec {PG} psql -U ace_user -d ace_system -tA -F'|' -c \"{sql}\""
            _, o, _ = ssh.exec_command(cmd, timeout=120)
            rows = [ln.split("|") for ln in o.read().decode().splitlines() if ln.count("|") >= 6]
        finally:
            ssh.close()
        wb = Workbook(); ws = wb.active; ws.title = "Clock Detail"
        HF = Font(bold=True, color="FFFFFF"); HB = PatternFill("solid", fgColor="2447D8")
        warn = PatternFill("solid", fgColor="FFF2CC")
        ws.append([f"{name}  ({code})", f"Month {month}", "", "", "", "", ""])
        ws["A1"].font = Font(bold=True, size=12)
        hdr = ["Date", "Day", "Clock In", "Clock Out", "Hours", "Site", "Note"]
        ws.append(hdr)
        for c in range(1, len(hdr) + 1):
            cell = ws.cell(2, c); cell.font = HF; cell.fill = HB; cell.alignment = Alignment(horizontal="center")
        total = 0.0
        for r in rows:
            d, dow, tin, tout, hrs, site, note = (r + [""] * 7)[:7]
            hv = float(hrs) if hrs else 0.0
            total += hv
            ws.append([d, dow.strip(), tin, tout, hv, site, note])
            if note:
                for c in range(1, 8): ws.cell(ws.max_row, c).fill = warn
        ws.append([])
        ws.append(["", "", "", "TOTAL", round(total, 1), "", ""])
        ws.cell(ws.max_row, 4).font = Font(bold=True); ws.cell(ws.max_row, 5).font = Font(bold=True)
        for i, w in enumerate([12, 6, 10, 10, 8, 10, 18], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A3"
        safe = name.replace(" ", "_").replace("/", "_")
        local = os.path.join(outdir, f"ClockInOut_{month}_{safe}.xlsx")
        wb.save(local)
        self._say(f"[XLSX] saved: {local} ({len(rows)} rows, total {round(total,1)} h)")

    def export_both(self):
        self.export_pdf()
        self.export_xlsx()
        self._say("Done (both).")


if __name__ == "__main__":
    root = tk.Tk()
    try:
        ttk.Style().theme_use("vista")
    except Exception:
        pass
    App(root)
    root.mainloop()
